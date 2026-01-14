import requests
import pandas as pd
import numpy as np
import json
import os
import re
from thefuzz import fuzz, process
import time
from io import StringIO
import psutil
import gc
import openpyxl
import pytz
from openpyxl import load_workbook
from pathlib import Path
import ast
import pandas_gbq
import pytz
import matplotlib.pyplot as plt
import seaborn as sns
import joblib
import warnings
warnings.filterwarnings("ignore")

project_id = "flex-power"
bigquery_import = """
SELECT *
FROM `flex-power.sales.origination_fixings`
"""
df_fixings = pandas_gbq.read_gbq(bigquery_import, project_id=project_id)

print(df_fixings)

def ensure_and_reorder(df, order):
    missing_cols = [col for col in order if col not in df.columns]
    for col in missing_cols:
        df[col] = None
    return df[order]

def check_memory_usage():
    process = psutil.Process(os.getpid())
    memory_info = process.memory_info()
    return memory_info.rss / 1024 ** 2

def ram_check():
    print("memory usage:", check_memory_usage(), "MB")

def convert_date_or_keep_string(date):
    try:
        date_obj = pd.to_datetime(date, dayfirst=True, errors='raise')
        return date_obj.strftime('%Y-%m-%d')
    except ValueError:
        return date

def is_number(val):
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False

hardcoded_map = {
    "V90 MK8 Gridstreamer": "V-90 2.0MW Gridstreamer",
    "V126-3.45MW": "V-126 3.45MW",
    "V-90" : "V-90 2.0MW Gridstreamer",
    "Vestas V90 2,00 MW 2.0MW": "V-90 2.0MW Gridstreamer",
    "Vestas V 90 NH 95m 2,00 MW 2.0MW": "V-90 2.0MW Gridstreamer",
    "Vestas V90 2,00 MW": "V-90 2.0MW Gridstreamer",
    "Vestas V90 2,00 MW": "V-90 2.0MW Gridstreamer",

    "V-112 2.0MW" : "V-112 3.3MW",
    "V136-3.6MW" : "V-136 3.6MW",
    "V112-3,45" : "V-112 3.45MW",
    "V162-5.6 MW" : "V-162 5.6MW",
    "V162-6.2 MW" : "V-162 6.2MW",
    "Vestas V162" : "N-163/6800",
    "V 150-4.2 MW" : "V-150 4.2MW (PO)",
    "Vestas V112-3.3 MW MK2A" : "V-112 3.3MW",
    "V 80 - 2.0MW / Mode 105.1 dB" : "V-80 2.0MW GridStreamer",
    "NTK600/43" : "V-44 0.6MW",
    "NTK 600 - 180" : "V-44 0.6MW",

    "V 80 - 2.0MW / Moder 105.1 dB" : "V-80 2.0MW GridStreamer",

    "Nordex N149-5.7 MW" : "N-149/5700",
    "Nordex N149-5.X" : "N-149/5700",
    "N149-5.7 MW" : "N-149/5700",
    "N175-6.8 MW" : "N-175/6800",
    "N163-6.8 MW" : "N-163/6800",
    "N163-5.7 MW" : "N-163/5700",
    "N163-7.0 MW" : "N-163/7000",
    "N149-5.7 MW" : "N-149/5700",
    "Nordex N149-5.7 MW" : "N-149/5700",
    "Nordex N149-5.7 MW" : "N-149/5700",
    "N163/6.X 6800" : "N-163/6800",
    "Nordex N133-4.8" : "N-133/4800",
    "Nordex N133/4.8 4800" : "N-133/4800",

    "N131/3000 PH134": "N-131/3000",
    "N149/4.0-4.5 NH164": "N-149/4500",
    "N117/2400R91 2400.0": "N-117/2400",

    "N 90-2.5" : "N-90/2500",
    "Nordex N149 4500" : "N-149/4500",
    "Nordex N131 3600" : "N-131/3600",

    "N 117-2.4" : "N-117/2400",
    "NordexN131 3300": "N-131/3300",

    "Vestas V90 2.0MW": "V-90 2.0MW Gridstreamer",
    "V-90 MK1-6": "V-90 2.0MW Gridstreamer",
    "V150-4.2 4.2MW": "V-150 4.2MW (PO)",
    "V150-4.2": "V-150 4.2MW (PO)",
    "N 117-2.4": "N-117/2400",
    "V117-3.3/3.45MWBWC" : "V-117 3.45MW",

    "E-66 1.8MW": "E-66/18.70",
    "E-66": "E-66/18.70",
    "E 53-0.81": "E-53 0.8MW",
    "E-82" :"E-82 E2 2.0MW",
    "E-58 1.0MW":"E-58/10.58",
    "E-58":"E-58/10.58",

    "V-80 2.0MW": "V-80 2.0MW GridStreamer",
    "V-80": "V-80 2.0MW GridStreamer",
    "v80 2.0MW": "V-80 2.0MW GridStreamer",

    "V150 4.2": "V-150 4.2MW (PO)",
    "E-138 EP3 E2-HAT-160-ES-C-01": "E-138 EP3 4.2MW",
    "N 131-3300": "N-131/3300",
    "N163/5.X": "N-163/5700",

    "Nordex N117/3600" : "N-117/3600",
    "N117/3.6" : "N-117/3600",

    "N-117 3150" : "N-117/3000",
    "N133 / 4.8 TS110" : "N-133/4800",
    "N149/5.7" : "N-149/5700",

    "Vensys 77" : "77/1500",
    "Senvion 3.4M104" : "3.4M104",
    "Senvion 3.2M" : "3.2M114",
    "Senvion 3.0M114": "3.2M114",
    "3.2M123" : "3.2M122",

    "REpower 3.4M 104 3.37MW": "3.4M122",
    "REpower 3.4M 104": "3.4M122",
    "Senvion 3.2M": "3.4M122",

    "E-141 EP4 4,2 MW" : "E-141 EP4 4.2MW",
    "E-70 E4-2/CS 82 a 2.3MW" : "E-70 E4 2.3MW",
    "E115 EP3  E3 4.2MW" : "E-115 EP3 4.2MW",
    "E115 EP3  E3" : "E-115 EP3 4.2MW",
    "E115 EP3 E3" : "E-115 EP3 4.2MW",
    "E-53/S/72/3K/02" : "E-53 0.8MW",
    "E82 E 2 2.3MW" :"E-82 E2 2.3MW",
    "E-40 0.5MW" : "E-40/5.40",

    "E 53-0.81": "E-53 0.8MW",
    "E 53-0.81 0.8MW": "E-53 0.8MW",


    "NM48/600" : "NM 48/600",
    "NEG MICON NM 600/48" : "NM 48/600",
    "NM600/48" : "NM 48/600",


    "E-70 E4 2300" : "E-70 E4 2.3MW",
    "E 82 Serrations" : "E-82 E2 2.3MW",
    "E40/540/E1" : "E-40/5.40",
    "TW600": "TW 600-43",


    "MM-92" : "MM 92 2.05MW",
    "MM92 2.05MW" : "MM 92 2.05MW",
    "MM-100" : "MM 100 2.0MW",
    "MM-82" : "MM 82 2.05MW",

    "MD-77" : "MD 77 1.5MW",

    "SWT-3.2" : "SWT-3.2-113",

    "GE-5.5" : "GE 5.5-158",
    "GE-3.6" : "GE 3.6-137",
    "AN62 1,3h": "1300/62",
    "N-117 Gamma": "N-117/2400",

    "AN600": "600/44",
    "AN76 2,0": "2000/76",
    "GE B1500": "GE 1.5sl",
    "Enercon-40": "E-40/5.40",
    "M 1500-600": "M1500-600",
    "GE Wind Energy 1.5 SL": "GE 1.5sl",

    "N-149/4500": "Nordex N149 4500.0",
    "Nordex N131 3600.0": "N-131/3600",
    "NordexN131 3300.0": "N-131/3000",
    "Vestas V150 4.2MW": "V-150 4.2MW (PO)",
    "V112 Mk2": "V-112 3.3MW",
    "MM92":"MM 92 2.05MW",
    "N149/5.7 TCS164":"N149/5.7 TCS164",
    "N149/4.5 TCS164": "N-149/4500",
    "N149-4,5MW": "N-149/4500",
    "V117-3,6MW": "V-117 3.45MW",
    "V117- 3,45 MW": "V-117 3.45MW",
    "Vestas V90 - 2,0 MW":"V-90 2.0MW Gridstreamer",
    "3.4M 98.0":"3.4M104",
    "eno92":"eno 92",
    "E53 800kW":"E-53 0.8MW",
    "Vestas V126 - 3,3 MW": "V-126 3.3MW",
    "Vestas V162-6.2":"V-162 6.2MW",
    "N163/6.X TCS164":"N-163/6800",
    "N163/5.X" : "N-163/5700",
    "Nordex N100" : "N-100/2500",
    "N149/4.0-4.5 Delta 4000": "N-149/4500",
    "N149/4.5 TS 125": "N-149/4500",
    "N149/5.x TS125": "N-149/5700",
    "N117/120 2400.0": "N-117/2400",
    "N117/120 2400": "N-117/2400",
    "E-66 1.8MW": "3.2M114",

    "SE 3.2M-114": "3.2M114",

    "N163/5.X,": "N-163/5700",
    "N149/5.X TS125-04 5700": "N-149/5700",
    "Delta 4000 N149/5.X": "N-149/5700",
    "N149-5.7": "N-149/5700",
    "Delta 4000 N149/5.X": "N-149/5700",
    "V150-4,2": "V-150 4.2MW (PO)",
    "MM92": "MM92/2050",

    "N149/4.0-4.5": "N149/5700",
    "NM48/750": "NM 48/600",
    "V162/6.2MW": "V162/6200",

    #delete below

    "N131" : "N-131/3900",
    "V 150 En Ventus 5.6MW" : "V-150 5.6MW",
    "V172-7.2 7.2MW"  : "V-172 7.2MW",
    "V90/2,0 MW 2.0MW" : "V-90 2.0MW Gridstreamer",
    "V80" : "V-150 5.6MW",
    "E-115 E2" : "E-115 3.2MW",
    "E-70 E4" : "E-70 E4 2.3MW",
    "Repower MD 77" : "MD 77 1.5MW",
    "V90-2.0MW" : "V-90 2.0MW Gridstreamer",
    "E-101" : "E-101 3.05MW",
    "E-115 EP3" : "E-115 3.0MW",
    "V90" : "V-90 2.0MW Gridstreamer",

    "E-115 E2" : "E-115 3.2MW",
    "E-70" : "E-70 E4 2.3MW",
    "MD 77" : "MD 77 1.5MW",
    "V90/2MW 2.0MW" : "V-90 2.0MW Gridstreamer",
    "N163/5.X TS118" : "N-163/5700",
    "E-92" : "E-92 2.35MW",

    "V126": "V-126 3.45MW",
    "N149" : "N-149/4500",
    "E 82" : "E-82 E2 2.3MW",
    "E-115" : "E-115 3.0MW",
    "V 150": "V-150 4.2MW (PO)",

    "V80" : "V-80 2.0MW GridStreamer",
    "V136-4.2": "V-136 4.2MW",
    "V117" : "V-117 3.3MW",
    "V112-3.45" : "V-112 3.45MW",
    "N163/6.X" : "N-163/6800",
    "E66" : "E-66/20.70",
    "E160-5.56" : "E-160 EP5 E3 5.56MW",
    "E66" : "E-66/20.70",
    "E138-4.2" : "E-138 EP3 4.2MW",
}

folder = Path(r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\DWAG\DWAG_curt_testing_v2")
folder_name = folder.name

path = folder / f"{folder_name}_stammdaten.xlsx"
out_path = folder / f"{folder_name}_customerpricing.xlsx"
out_path_highlighted = folder / f"{folder_name}_customerpricing_highlighted.xlsx"

# path = r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\DWAG\DWAG_curt_testing_v2\DWAG_curt_testing_v2_stammdaten.xlsx"
# out_path = r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\ppa_connect\ppa_connect_0925_GLS\ppa_connect_0925_GLS_customerpricing.xlsx"

# export stammdaten ()
df_stamm = pd.read_excel(path,
                        sheet_name='stammdaten',
                        # usecols= ['Projekt', 'malo','Marktstammdatenregister-ID', 'tech', "net_power_mw","longitude", "latitude","hub_height_m", "manufacturer", "turbine_model","geplante Inbetriebnahme"],
                        engine='openpyxl')

df_stamm.columns = df_stamm.columns.str.strip()
df_stamm['malo'] = df_stamm['malo'].apply(lambda x: str(int(x)) if isinstance(x, (float, int)) and pd.notna(x) else str(x)).str.strip()

df_stamm['Marktstammdatenregister-ID'] = df_stamm['Marktstammdatenregister-ID'].astype(str).str.strip()
df_stamm.dropna(subset=("malo"),axis=0, inplace=True)
df_stamm.dropna(subset=("net_power_kw_unit","turbine_model","latitude","longitude"),axis=0, inplace=True)
df_stamm.rename(columns= {'Marktstammdatenregister-ID': 'unit_mastr_id'}, inplace=True)
#df_stamm['net_power_kw'] = df_stamm['net_power_mw'] * 1000

latitude_min = 47.0  # minimum latitude (southern border)
latitude_max = 55.0  # maximum latitude (northern border)
longitude_min = 5.8  # minimum longitude (western border)
longitude_max = 15.0  # maximum longitude (eastern border)


# Create a mask for rows that are outside Germany
mask_invalid_location = ~df_stamm['latitude'].between(latitude_min, latitude_max) | \
                        ~df_stamm['longitude'].between(longitude_min, longitude_max)

mask_integer_coords = (
    np.isclose(df_stamm['latitude'] % 1, 0, atol=1e-10) |
    np.isclose(df_stamm['longitude'] % 1, 0, atol=1e-10)
)

mask_to_remove = mask_invalid_location | mask_integer_coords

# Show rows with invalid coordinates (outside Germany)
invalid_rows = df_stamm[mask_to_remove]
print("Rows with invalid latitude/longitude (outside Germany):")
print(invalid_rows)

print("🪭🪭🪭")
df_stamm = df_stamm[~mask_to_remove]
print(df_stamm.info())

# Print the number of remaining rows
print(f"Remaining rows after dropping invalid locations: {len(df_stamm)}")

#df_stamm = df_stamm.iloc[:500]

# assigning "EEG" and "category"
def set_category_based_on_conditions(df_assets: pd.DataFrame) -> pd.DataFrame:
    """
    df_assets must contain at least:
        - 'malo'
        - 'tech'  (values: 'PV' or 'WIND')
        - 'EEG'         (string, 'no rules' or something else) - will be created/updated
        - "malo_kw"     (capacity per malo in kW)
        - 'INB'         (date in DD/MM/YYYY format)
        - 'net_power_kw_unit' (power in kW)
    """
    df_assets = df_assets.copy()

    df_assets['INB_date'] = pd.to_datetime(df_assets['INB'], dayfirst=True, errors='coerce')
    df_assets['INB_year'] = df_assets['INB_date'].dt.year

    # Define conditions for EEG rules
    conditions = [
        # Condition 1: Empty INB
        df_assets['INB'].isna() | (df_assets['INB'] == ''),

        # Condition 2: WIND with >=3000 kW and year 2016-2020
        (df_assets['tech'] == 'WIND') &
        (df_assets['net_power_kw_unit'] >= 3000) &
        (df_assets['INB_year'] >= 2016) & (df_assets['INB_year'] < 2021),

        # Condition 3: PV with >=500 kW and year 2016-2020
        (df_assets['tech'] == 'PV') &
        (df_assets['net_power_kw_unit'] >= 500) &
        (df_assets['INB_year'] >= 2016) & (df_assets['INB_year'] < 2021),

        # Condition 4: >=500 kW and year 2021-2022
        (df_assets['net_power_kw_unit'] >= 500) &
        (df_assets['INB_year'] >= 2021) & (df_assets['INB_year'] < 2023),

        # Condition 5: >=100 kW and year >=2023
        (df_assets['net_power_kw_unit'] >= 100) &
        (df_assets['INB_year'] >= 2023)
    ]

    choices = [
        'rules',
        '6h rules',
        '6h rules',
        '4h rules',
        '4_3_2_1 rules'
    ]

    df_assets['EEG'] = np.select(conditions, choices, default='no rules')

    df_assets = df_assets.drop(columns=['INB_date', 'INB_year'])

    # Set category based on technology and EEG
    df_assets['category'] = df_assets.apply(
        lambda row: 'PV_no_rules' if row['tech'] == 'PV' and row['EEG'] == 'no rules'
        else ('PV_rules' if row['tech'] == 'PV'
        else ('WIND_no_rules' if row['tech'] == 'WIND' and row['EEG'] == 'no rules'
        else 'WIND_rules')),
        axis=1
    )
    return df_assets

df_stamm = set_category_based_on_conditions(df_stamm)

print(df_stamm)
df_fixings["Tenor"] = df_fixings["Tenor"].astype(str)
print(df_fixings.info())

# pick a single value for (Technology, Variable)
def get_fix_value(df_fixings, tech, variable, year):
    sel = df_fixings[(df_fixings['Technology'] == tech) &
                     (df_fixings['Variable'] == variable) &
                     (df_fixings['Tenor'] == year)]
    # prefer non-zero new_Fixing; otherwise fall back to EUR_MWh
    s_new = sel['new_Fixing'].replace(0, np.nan).dropna()
    if not s_new.empty:
        return float(s_new.iloc[0])
    s_eur = sel['EUR_MWh'].dropna()
    return float(s_eur.iloc[0]) if not s_eur.empty else np.nan

# ---- compute all needed scalars ----
bc_pv   = get_fix_value(df_fixings, 'PV',   'Balancing Cost', '2026')
bc_wind = get_fix_value(df_fixings, 'Wind', 'Balancing Cost', '2026')

tc_pv   = get_fix_value(df_fixings, 'PV',   'Trading Convenience', '2026')
tc_wind = get_fix_value(df_fixings, 'Wind', 'Trading Convenience', '2026')

cv_pv_no   = get_fix_value(df_fixings, 'PV',   'Curtailment Value without Rule', '2026')
cv_pv_yes  = get_fix_value(df_fixings, 'PV',   'Curtailment Value with any Rule', '2026')
cv_w_no    = get_fix_value(df_fixings, 'Wind', 'Curtailment Value without Rule', '2026')
cv_w_yes   = get_fix_value(df_fixings, 'Wind', 'Curtailment Value with any Rule', '2026')

def check_scalar(scalar, name):
    if np.isnan(scalar):
        print(f"Warning: {name} was not successfully fetched.")
    else:
        print(f"{name} successfully fetched: {scalar}")

# Check each scalar value
check_scalar(bc_pv, 'Balancing Cost PV')
check_scalar(bc_wind, 'Balancing Cost Wind')
check_scalar(tc_pv, 'Trading Convenience PV')
check_scalar(tc_wind, 'Trading Convenience Wind')
check_scalar(cv_pv_no, 'Curtailment Value PV (no rule)')
check_scalar(cv_pv_yes, 'Curtailment Value PV (with rule)')
check_scalar(cv_w_no, 'Curtailment Value Wind (no rule)')
check_scalar(cv_w_yes, 'Curtailment Value Wind (with rule)')

# def get_fix_value(df_fixings, tech, variable):
#     sel = df_fixings[(df_fixings['Technology'] == tech) &
#                      (df_fixings['Variable'] == variable)]
#     # prefer non-zero new_Fixing; otherwise fall back to EUR_MWh
#     s_new = sel['new_Fixing'].replace(0, np.nan).dropna()
#     if not s_new.empty:
#         return float(s_new.iloc[0])
#     s_eur = sel['EUR_MWh'].dropna()
#     return float(s_eur.iloc[0]) if not s_eur.empty else np.nan

# # ---- compute all needed scalars ----
# bc_pv   = get_fix_value(df_fixings, 'PV',   'Balancing Cost')
# bc_wind = get_fix_value(df_fixings, 'Wind', 'Balancing Cost')

# tc_pv   = get_fix_value(df_fixings, 'PV',   'Trading Convenience')
# tc_wind = get_fix_value(df_fixings, 'Wind', 'Trading Convenience')

# cv_pv_no   = get_fix_value(df_fixings, 'PV',   'Curtailment Value without Rule')
# cv_pv_yes  = get_fix_value(df_fixings, 'PV',   'Curtailment Value with any Rule')
# cv_w_no    = get_fix_value(df_fixings, 'Wind', 'Curtailment Value without Rule')
# cv_w_yes   = get_fix_value(df_fixings, 'Wind', 'Curtailment Value with any Rule')

# ---- build masks once ----
m_pv   = df_stamm['tech'].str.upper().eq('PV')
m_wind = df_stamm['tech'].str.upper().eq('WIND')
m_no   = df_stamm['EEG'].str.contains('no rules', case=False, na=False)

# ---- assign scalars ----
# Balancing Cost
df_stamm.loc[m_pv,   'Balancing Cost'] = bc_pv
df_stamm.loc[m_wind, 'Balancing Cost'] = bc_wind

# Trading Convenience
df_stamm.loc[m_pv,   'Trading Convenience'] = tc_pv
df_stamm.loc[m_wind, 'Trading Convenience'] = tc_wind

# Curtailment Value
df_stamm.loc[m_pv & m_no,      'Curtailment Value'] = cv_pv_no
df_stamm.loc[m_pv & ~m_no,     'Curtailment Value'] = cv_pv_yes
df_stamm.loc[m_wind & m_no,    'Curtailment Value'] = cv_w_no
df_stamm.loc[m_wind & ~m_no,   'Curtailment Value'] = cv_w_yes


# Calculate weighted Curtailment Value per 'malo'
df_stamm_curt_value = df_stamm.groupby(['malo'], dropna=False).agg(
    # Calculate the sum of weighted Curtailment Value for each 'malo'
    Curtailment_value_weighted = ('Curtailment Value', lambda x: np.average(x, weights=df_stamm.loc[x.index, 'net_power_kw_unit'])),).reset_index()

df_stamm = pd.merge(
    df_stamm,
    df_stamm_curt_value,
    on = "malo",
    how = "left"
    )


see_data = df_stamm[df_stamm['unit_mastr_id'].str.strip().str.lower().str.startswith('see')].copy()
# IF THERE IS SEE DATA . Jumping to 1126 if there is NO SEE data
if not see_data.empty:
    # Filter column (case-insensitive, starts with "see")
    id_list = see_data['unit_mastr_id'].dropna()
    valid_ids = [
        str(id_).strip().upper()
        for id_ in id_list
        if str(id_).strip().lower().startswith("see")
    ]

    #valid_ids = ['SEE999883285843']
    #Get unique values
    unique_see = set(valid_ids)
    count = len(unique_see)
    print(f"🔢 Unique 'see' values in column SEE: {count}")

    del unique_see
    gc.collect()

    # Step 1: Get access token
    headers = {
        'accept': 'text/plain',
        'Content-Type': 'application/json',
    }

    json_data = {
        'email': 'abc@abc.energy',
        'password': '009009009',
    }

    response = requests.post('https://api.blindleister.de/api/v1/authentication/get-access-token', headers=headers, json=json_data)

    if response.status_code != 200:
        print("Failed to get access token:", response.status_code, response.text)
        exit()

    token = response.text.strip('"')  # Remove potential extra quotes
    print("Access token:", token)


    # Step 2: Use the token to query market price
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOiJvcHNAZmxleC1wb3dlci5lbmVyZ3kifQ.Q1cDDds4fzzYFbW59UuZ4362FnmvBUQ8FY4UNhWp2a0'
    }

    # Fetch blindleister price
    print("🍚🍚")
    # === Years to fetch ===
    years = [2021, 2023, 2024]
    records = []

    # === Loop through each ID and fetch data for each year ===
    for site_id in valid_ids:
        print(f"Processing: {site_id}")

        for year in years:
            payload = {
                'ids': [site_id],
                'year': year
            }

            response = requests.post(
                'https://api.blindleister.de/api/v1/market-price-atlas-api/get-market-price', # market price atlas blindleister API
                headers = headers,
                json=payload
            )

            if response.status_code != 200:
                print(f"  Year {year}: Failed ({response.status_code}) - {response.text}")
                continue

            try:
                result = response.json()
                for entry in result:
                    entry['year'] = year
                    records.append(entry)
            except Exception as e:
                print(f"  Year {year}: Error parsing response - {e}")
                continue

    df_flat = pd.DataFrame(records)

    #if 'df_flat' in locals() and not merge_a3_see.empty:

    df_flat = pd.json_normalize(
        records,
        record_path="months",
        meta=[
            "year",
            "unit_mastr_id",
            "gross_power_kw",
            "energy_source",
            "annual_generated_energy_mwh",
            "benchmark_market_price_eur_mwh",
        ],
        errors="ignore"  # in case some records lack "months"
    )

    cols = [
        "year",
        "unit_mastr_id",
        "gross_power_kw",
        "energy_source",
        "annual_generated_energy_mwh",
        "benchmark_market_price_eur_mwh",
        "month",
        "monthly_generated_energy_mwh",
        "monthly_energy_contribution_percent",
        "monthly_market_price_eur_mwh",
        "monthly_reference_market_price_eur_mwh",
    ]

    df_flat = df_flat[cols]
    df_all_flat = df_flat.copy()

    print("🥨🥨🥨🥨🥨🥨")

    del df_flat
    gc.collect()

    df_all_flat['spot_rmv_EUR_monthly_ytd'] = (
    ((df_all_flat['monthly_generated_energy_mwh'] * df_all_flat['monthly_market_price_eur_mwh']) -
        (df_all_flat['monthly_generated_energy_mwh'] * df_all_flat['monthly_reference_market_price_eur_mwh']))
    )

    permalo_yearly_blind = df_all_flat.groupby(['year', 'unit_mastr_id'], dropna=False).agg(
        spot_rmv_EUR_yearly = ('spot_rmv_EUR_monthly_ytd', 'sum'),
        sum_prod_yearly = ('monthly_generated_energy_mwh', 'sum')
    ).assign(
        blind_yearly = lambda x: x['spot_rmv_EUR_yearly'] / x['sum_prod_yearly']
    ).reset_index()


    permalo_blind = df_all_flat.groupby('unit_mastr_id', dropna=False).agg(
        spot_rmv_EUR_ytd=('spot_rmv_EUR_monthly_ytd', 'sum'),
        sum_prod_ytd=('monthly_generated_energy_mwh', 'sum')
    ).assign(
        average_weighted_eur_mwh_blindleister=lambda x: x['spot_rmv_EUR_ytd'] / x['sum_prod_ytd']
    ).reset_index()

    print(permalo_blind)

    print("🥨🥨🥨")
    print(permalo_yearly_blind)
    ram_check()

    # year_agg_per_unit_new = permalo_yearly_blind.groupby(['year', 'unit_mastr_id'])['weighted_per_mwh_monthly'].mean().reset_index(name='weighted_year_agg_per_unit_eur_mwh')
    # weighted_years_pivot_new = pd.DataFrame(year_agg_per_unit_new)

    weighted_years_pivot_new = permalo_yearly_blind.pivot(
        index='unit_mastr_id',
        columns='year',
        values='blind_yearly'
    ).reset_index()


    weighted_years_pivot_new.columns.name = None  # remove the axis name
    weighted_years_pivot_new = weighted_years_pivot_new.rename(columns={
        2021: 'weighted_2021_eur_mwh_blindleister',
        2023: 'weighted_2023_eur_mwh_blindleister',
        2024: 'weighted_2024_eur_mwh_blindleister'
    })

    cols_to_round = [
        'weighted_2021_eur_mwh_blindleister',
        'weighted_2023_eur_mwh_blindleister',
        'weighted_2024_eur_mwh_blindleister',
        "average_weighted_eur_mwh_blindleister"
        ]


    final_weighted_blindleister_new = weighted_years_pivot_new[[
        'unit_mastr_id',
        'weighted_2021_eur_mwh_blindleister',
        'weighted_2023_eur_mwh_blindleister',
        'weighted_2024_eur_mwh_blindleister',
    ]]

    final_weighted_blindleister = pd.merge(
        final_weighted_blindleister_new,
        permalo_blind[["unit_mastr_id", "average_weighted_eur_mwh_blindleister"]],
        on= 'unit_mastr_id',
        how='left'
    )

    final_weighted_blindleister[cols_to_round] = final_weighted_blindleister[cols_to_round].round(2)

    del weighted_years_pivot_new
    gc.collect()

    print("🚦🚦🚦🚦🚦")
    print(final_weighted_blindleister)

    merge_a1 = pd.merge(
        df_stamm,
        final_weighted_blindleister,
        on = 'unit_mastr_id',
        how='left'
        )

    if pd.api.types.is_numeric_dtype(merge_a1["malo"]):
        merge_a1["malo"] = (
            merge_a1["malo"]
            .astype(float)
            .astype("Int64")   # nullable int (keeps NaN)
            .astype(str)
    )
    else:
        # already string → just ensure dtype string
        merge_a1["malo"] = merge_a1["malo"].astype(str)

    merge_a1.to_excel(f"{folder}\\merge_a1.xlsx", index=False)
    print(merge_a1.info())
    df_combined = merge_a1

    #=== Years to fetch ===
    years_dummy = [2024]
    all_records_enervis = []

    #=== Loop through each ID and fetch data for each year ===
    for site_id in valid_ids:
        print(f"Processing: {site_id}")
        records = []

        for year in years_dummy:
            payload = {
                'ids': [site_id],
                'year': year
            }

            response = requests.post(
                'https://api.blindleister.de//api/v1/mastr-api/get-generator-details', # MaStR Data blindleister
                headers = headers,
                json=payload
            )

            if response.status_code != 200:
                print(f"  Year {year}: Failed ({response.status_code}) - {response.text}")
                continue

            try:
                result = response.json()
                for entry in result:
                    entry['year'] = year
                    records.append(entry)
            except Exception as e:
                print(f"  Year {year}: Error parsing response - {e}")
                continue


        if records:
            all_records_enervis.extend(records)
            print(f"  ✅ Appended {len(records)} records from {site_id}")
        else:
            print(f"  ⚠️ No data found for {site_id}")

    df_blind_fetch = pd.DataFrame(all_records_enervis)
    print(df_blind_fetch.info())
    df_blind_fetch.to_excel(f"{folder}\\df_blind_fetch.xlsx", index=False)

    del all_records_enervis
    gc.collect()

    solar  = df_blind_fetch[df_blind_fetch["energy_source"] != "wind"].copy()
    wind   = df_blind_fetch[df_blind_fetch["energy_source"] == "wind"].copy()

    # if the tech is WIND (with SEE), start calculating enervis wind
    if not wind.empty:
        columns_to_use = [
        'unit_mastr_id', 'windpark', 'manufacturer', 'turbine_model',
        'hub_height_m', 'energy_source', 'latitude', 'longitude', 'net_power_kw'
        ]

        df_blind_fetchturbine = df_blind_fetch[columns_to_use]
        mw_to_map = df_blind_fetch[['unit_mastr_id','net_power_kw']]

        del df_blind_fetch
        gc.collect()

        df_combined = df_blind_fetchturbine[df_blind_fetchturbine["energy_source"] == 'wind']
        df_combined['net_power_mw'] = df_blind_fetchturbine['net_power_kw'] / 1000

        df_combined = df_combined.dropna(subset=['turbine_model'])

        del df_blind_fetchturbine
        gc.collect()

        #Load the turbine reference turbine id from enervis
        df_ref_enervis = pd.read_excel(
            r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\Enervis template\turbine_types_id_enervis_eraseMW.xlsx")
        df_ref_enervis['id'] = df_ref_enervis['id'].astype(str).str.strip()


        def clean_name(name):
            name = str(name).lower()
            for remove_word in [
                'gmbh', 'se', 'deutschland', 'central europe', 'energy',
                'gmbh & co. kg', 'energy gmbh', 'deutschland gmbh', "GmbH", "GmbH & Co. KG", "Deutschland GmbH", "AG"
            ]:
                name = name.replace(remove_word.lower(), '')
            return name.strip()

        # 2. Clean turbine model for some manufacturers
        def clean_name_turbine(name):
            if not isinstance(name, str):
                return ''
            for remove_word in ['senvion', 'Senvion', 'nercon', 'nercon', 'mit Serrations', 'Vensys', 'NEG MICON', 'Delta4000', 'Delta 4000']:
                name = name.replace(remove_word, '')
            return name.strip()

        df_combined['clean_manufacturer'] = df_combined['manufacturer'].apply(clean_name)
        #df_combined = df_combined.dropna(subset=['turbine_model', 'hub_height_m'])
        df_combined['turbine_model_clean'] = df_combined['turbine_model'].apply(clean_name_turbine)
        df_combined['add_turbine'] = df_combined['turbine_model']

        df_combined.loc[
            df_combined['manufacturer'].isin(['Vestas Deutschland GmbH', 'Senvion Deutschland GmbH', 'ENERCON GmbH', 'VENSYS Energy AG', 'Enron Wind GmbH', 'NEG Micon Deutschland GmbH']),
            'add_turbine'
        ] = df_combined['turbine_model_clean'].astype(str).str.strip() + ' ' + df_combined['net_power_mw'].round(3).astype(str) + 'MW'

        df_combined.loc[
            df_combined['manufacturer'].isin(['Nordex Energy GmbH' , 'REpower Systems SE', 'Nordex Germany GmbH', "eno energy GmbH"]),
            'add_turbine'
        ] = df_combined['turbine_model_clean'].astype(str).str.strip() + ' ' + df_combined['net_power_kw'].astype(str)

        df_combined.loc[
            df_combined['manufacturer'].isin(['REpower Systems SE']),
            'add_turbine'
        ] = df_combined['turbine_model_clean'].astype(str).str.strip() + ' ' + df_combined['hub_height_m'].astype(str)

        # Apply fuzzy match only if not in hardcoded list
        def match_add_turbine_with_exceptions(row, choices, threshold=73):
            turbine_model = row['turbine_model']
            add_turbine = row['add_turbine']

            # Check for hardcoded rule first
            if turbine_model in hardcoded_map:
                return hardcoded_map[turbine_model]

            if add_turbine in hardcoded_map:
                return hardcoded_map[add_turbine]

            # Fallback to fuzzy match
            if isinstance(add_turbine, str) and add_turbine.strip():
                #match, score = process.extractOne(add_turbine, choices, scorer=fuzz.token_sort_ratio)
                match, score = process.extractOne(add_turbine, choices, scorer=fuzz.token_set_ratio)
                if score >= threshold:
                    return match
            return None

        # Apply the function to each row
        df_combined['Matched_Turbine_Name'] = df_combined.apply(
            lambda row: match_add_turbine_with_exceptions(row, df_ref_enervis['name'].dropna().unique()),
            axis=1
        )

        name_to_id_map = df_ref_enervis.set_index('name')['id'].to_dict()
        df_combined['Matched_Turbine_ID'] = df_combined['Matched_Turbine_Name'].map(name_to_id_map)

        del df_ref_enervis, name_to_id_map
        gc.collect()

        print("🍹🍹🍹")
        print(df_combined.head(25))

        print("🍲🍲🍲🍲🍲🍲🍲🍲")

        df_combined.to_excel(f"{folder}\\test_NaN_turbineID.xlsx", index=False)
        df_combined = df_combined.dropna(subset=["Matched_Turbine_ID"])
        print(df_combined.head(20))

        # 5. Show result
        df_combined['hub_height_m'] = df_combined['hub_height_m'].fillna(104).replace(0, 104)
        df_combined["hub_height_m"] = df_combined["hub_height_m"].astype(int)
        df_combined["Matched_Turbine_ID"] = df_combined["Matched_Turbine_ID"].astype(int)

        print("🥨🥨🥨")

        df_stamm_malo = df_stamm[["malo","unit_mastr_id"]]
        df_fuzzy_excel = pd.merge(df_stamm_malo, df_combined, on='unit_mastr_id', how='right')

        #df_fuzzy_excel["hub_height_m"] = df_fuzzy_excel["hub_height_m"].apply(lambda x: int(x) if pd.notnull(x) else np.nan)

        df_fuzzy_excel["malo"] = df_fuzzy_excel["malo"].astype(str)
        df_fuzzy_excel.to_excel(f"{folder}\\df_fuzzy_excel.xlsx", index=False)

        gc.collect()

        print("🥕🥕🍛🍛🍛🥕🥕🥕🥕")
        print(df_fuzzy_excel)
        print("🍛🍛🍛")

        EMAIL = "abc@abc.energy"
        PASSWORD = "999999"

        # --- Step 1: Get access token ---
        def get_token():
            url = "https://keycloak.anemosgmbh.com/auth/realms/awis/protocol/openid-connect/token"
            data = {
                'client_id': 'webtool_vue',
                'grant_type': 'password',
                'username': EMAIL,
                'password': PASSWORD
            }
            response = requests.post(url, data=data)

            # Debug output
            print("Status code:", response.status_code)
            print("Response:", response.text)

            response.raise_for_status()
            return response.json()['access_token']


        # --- Step 2: Get historical On-Demand product ID ---
        def get_historical_product_id(token):
            url = "https://api.anemosgmbh.com/products_mva"
            headers = {"Authorization": f"Bearer {token}"}
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            products = response.json()

            print("\n📦 Available products:")
            for p in products:
                product_type_name = p["mva_product_type"]["name"]
                print(f"- ID: {p['id']}, Name: {product_type_name}")
                if "hist-ondemand" in product_type_name.lower():
                    print(f"✅ Found 'hist-ondemand' product.")
                    return p["id"]

            raise Exception("❌ No 'hist-ondemand' product found.")

        def list_turbine_types(token):
            url = "https://api.anemosgmbh.com/turbine_types"
            headers = {"Authorization": f"Bearer {token}"}
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            types = response.json()
            print("\n🌀 Available turbine types:")
            for t in types:
                print(f"- ID: {t['id']}, Name: {t.get('name', 'n/a')}")

            df = pd.DataFrame(types)
            filename = r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\Enervis template\turbine_types_id_enervis.xlsx"
            #df.to_excel(filename, index=False)
            #print(f"✅ Turbine types saved to {filename}")

            return types

        def start_historical_job_from_df(token, product_id, df_input):
            url = "https://api.anemosgmbh.com/jobs"
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }

            # Build parkinfo list from DataFrame rows
            parkinfo_list = []
            for _, row in df_input.iterrows():
                if pd.notnull(row["Matched_Turbine_ID"]):
                    hub_height = row["hub_height_m"]
                    if pd.isna(hub_height):
                        hub_height = 104
                    parkinfo_list.append({
                        "id": int(row["malo"]),
                        "lat": str(row["latitude"]),
                        "lon": str(row["longitude"]),
                        "turbine_type_id": int(row["Matched_Turbine_ID"]),
                        "hub_height": int(hub_height)  # Now this will always be a number
                    })

            if not parkinfo_list:
                print("⚠️  Warning: No valid parkinfo entries found in input DataFrame. Skipping job.")
                return None

            payload = {
                "mva_product_id": product_id,
                "parameters": {
                    "parkinfo": parkinfo_list
                }
            }

            print("\n📤 Sending payload with multiple turbines:")
            print(payload)

            response = requests.post(url, headers=headers, json=payload)

            if response.status_code != 200:
                print("\n❌ API responded with:")
                print(response.text)
                response.raise_for_status()

            resp_json = response.json()
            return resp_json["uuid"]

        # --- Step 4: Poll the job status until complete ---
        def wait_for_job_completion(token, job_uuid, poll_interval=10):
            url = f"https://api.anemosgmbh.com/jobs/{job_uuid}"

            while True:
                headers = {"Authorization": f"Bearer {token}"}
                response = requests.get(url, headers=headers)

                if response.status_code == 401:
                    # Token expired - get new token and retry once
                    print("Token expired, refreshing token...")
                    token = get_token()
                    headers = {"Authorization": f"Bearer {token}"}
                    response = requests.get(url, headers=headers)

                response.raise_for_status()
                job_info = response.json()

                if isinstance(job_info, list):
                    job_info = job_info[0]

                status = job_info.get("status")
                print(f"Job status: {status}")

                if status == "DONE" or status == "COMPLETED":
                    return job_info
                elif status in ["FAILED", "CANCELED"]:
                    raise Exception(f"Job ended with status: {status}")

                time.sleep(poll_interval)

        # --- Step 5: Download result files and load them into DataFrames ---
        def download_result_files(job_info, token):
            headers = {"Authorization": f"Bearer {token}"}
            files = job_info.get("files")

            if files:
                dfs = []
                for f in files:
                    file_url = f.get("url")
                    print(f"📥 Downloading result file: {file_url}")
                    df = download_and_load_csv(file_url, token)
                    dfs.append(df)
                return dfs
            else:
                print("❌ No result files found — checking 'info' field...")
                results = job_info.get("info", {}).get("results", [])

                if results:
                    dfs = []
                    for result in results:
                        turbine_id = result.get("id")
                        year_data = result.get("Marktwertdifferenzen")
                        if year_data:
                            df = pd.DataFrame.from_dict(year_data, orient="index", columns=["Marktwertdifferenz"])
                            df.index.name = "Year"
                            df = df.reset_index()
                            df["id"] = turbine_id
                            dfs.append(df)
                            # print("📊 Extracted results from 'info':")
                            # print(df)
                        else:
                            print("❌ No 'Marktwertdifferenzen' found in result.")
                    return dfs if dfs else None
                else:
                    print("❌ No usable results found in 'info'.")
                    return None

        df_enervis_pivot = pd.DataFrame()
        # --- Main workflow ---
        if __name__ == "__main__":
            try:
                print("🔐 Getting access token...")
                token = get_token()

                print("📦 Getting historical product ID...")
                product_id = get_historical_product_id(token)

                print("🚀 Starting historical On-Demand job...")
                job_uuid = start_historical_job_from_df(token, product_id, df_fuzzy_excel)

                if job_uuid is None:
                    print("⏭️  No valid turbines to process. Skipping job execution and download.")
                    # You can choose to set an empty DataFrame or take other actions here
                    all_df = pd.DataFrame()  # Create an empty DataFrame to avoid errors later
                else:
                    print(f"✅ Job started with UUID: {job_uuid}")

                    print("⏳ Waiting for job completion...")
                    job_info = wait_for_job_completion(token, job_uuid)

                    print("📁 Job finished! Getting result files...")
                    dfs = download_result_files(job_info, token)

                    all_df = pd.concat(dfs, ignore_index=True)
                    del dfs
                    gc.collect()

                    # Convert 'Year' to string once for consistent processing
                    all_df["Year"] = all_df["Year"].astype(str)
                    print("Unique years in raw data:", all_df["Year"].unique())

                    target_years = ["2021", "2023", "2024"]
                    existing_years = all_df["Year"].unique().tolist()

                    # Filter target_years that actually exist in data
                    valid_years = [y for y in target_years if y in existing_years]
                    print("Valid years to process:", valid_years)

                    if not valid_years:
                        raise ValueError("❌ No target years found in data")

                    # Filter rows to keep only valid years
                    all_df = all_df[all_df["Year"].isin(valid_years)].copy()
                    # print("\nData after year filtering:")
                    # print(all_df[["id", "Year", "Marktwertdifferenz"]].head(20))
                    # print(all_df[["id", "Year", "Marktwertdifferenz"]].tail(20))
                    all_df.to_excel(f"{folder}\\all_df_1.xlsx", index=False)


                    # Step 1: Filter to keep only the minimum Marktwertdifferenz per (id, Year)
                    df_filtered = all_df.loc[
                        all_df.groupby(["id", "Year"])["Marktwertdifferenz"].idxmin()
                    ].copy()

                    df_filtered['Marktwertdifferenz'] = df_filtered['Marktwertdifferenz'].round(2)

                    del all_df
                    gc.collect()

                    # Step 2: Pivot to wide format
                    df_enervis_pivot = df_filtered.pivot(
                        index="id",
                        columns="Year",
                        values="Marktwertdifferenz"
                    ).rename_axis(None, axis=1).reset_index()

                    del df_filtered
                    gc.collect()

                    # Step 3: Ensure all year columns are present
                    for year in target_years:
                        if year not in df_enervis_pivot.columns:
                            df_enervis_pivot[year] = np.nan

                    # Check which valid_years are really in the pivot columns
                    available_columns = [col for col in valid_years if col in df_enervis_pivot.columns]
                    print("Available columns:", available_columns)

                    if not available_columns:
                        raise ValueError("❌ No valid year columns found in pivot table")

                    # Calculate average across available year columns
                    df_enervis_pivot["avg_enervis"] = df_enervis_pivot[target_years].mean(axis=1, skipna=True)
                    columns_to_keep = ["id"] + target_years + ["avg_enervis"]
                    df_enervis_pivot_filter = df_enervis_pivot[columns_to_keep]

                    del df_enervis_pivot
                    gc.collect()
                    print("🥥🥥 df_enervis_pivot_filter")
                    print(df_enervis_pivot_filter)
                    df_enervis_pivot_filter.to_excel(f"{folder}\\df_enervis_pivot_filter_0_500.xlsx", index=False)

            except Exception as e:
                print(f"❌ Error: {e}")
                raise

    df_stamm["Power in MW"] = df_stamm["net_power_kw_unit"] / 1000
    merge_a1["Power in MW"] = merge_a1["net_power_kw_unit"] / 1000

    # merge_a1 = merge_a1.loc[merge_a1.groupby(["malo"])["average_weighted_eur_mwh_blindleister"].idxmin()].copy()
    # print("🥓🥓🥓")
    # print(merge_a1)
    # merge_a1.to_excel(f"{folder}\\merge_a1_after_idxmin.xlsx", index=False)

    merge_a1_agg = merge_a1.groupby(['malo'], dropna=False).agg({
                    'unit_mastr_id': 'first',
                    'Projekt': 'first',
                    'tech': 'first',
                    "Power in MW": 'sum',
                    "INB": lambda x: [convert_date_or_keep_string(date) for date in x],
                    #"EEG": list,
                    "EEG": lambda x: list(x.unique()),  # Get distinct EEG values
                    #'Vermarktung': lambda x: list(x.unique()),
                    'AW in EUR/MWh': lambda x: [round(float(val), 2) for val in x if is_number(val)],
                    #"Gesetzliche Vergütung (EEG) inkl. Managementprämie" : lambda x: [round(float(val), 2) for val in x if is_number(val)],
                    "weighted_2021_eur_mwh_blindleister": "min",
                    "weighted_2023_eur_mwh_blindleister": "min",
                    "weighted_2024_eur_mwh_blindleister": "min",
                    "average_weighted_eur_mwh_blindleister": "min",
                    "Curtailment & redispatch included": 'first',
                    "Balancing Cost": 'first',
                    "Curtailment_value_weighted" : 'first',
                    "Trading Convenience": 'first',
                }).reset_index()

    # IF THERE IS WIND SEE RESULT ENERVIS
    if 'df_enervis_pivot_filter' in locals() and not df_enervis_pivot_filter.empty:

        df_enervis_pivot_filter['id'] = df_enervis_pivot_filter['id'].astype(str)

        merge_a3 = pd.merge(
            merge_a1_agg,
            df_enervis_pivot_filter,
            left_on='malo',
            right_on='id',
            how='left'
        )
        merge_a3.drop(columns=['id'], inplace=True)

        desired_order = [
            'malo',
            'Projekt',
            'tech',
            'Power in MW',
            'INB',
            #'Vermarktung',
            'EEG',
            'AW in EUR/MWh',
            #"Gesetzliche Vergütung (EEG) inkl. Managementprämie",
            'weighted_2021_eur_mwh_blindleister',
            'weighted_2023_eur_mwh_blindleister',
            'weighted_2024_eur_mwh_blindleister',
            'average_weighted_eur_mwh_blindleister',
            "2021",
            "2023",
            "2024",
            "avg_enervis",
            "Curtailment & redispatch included",
            "Balancing Cost",
            "Curtailment_value_weighted",
            "Trading Convenience"
        ]

        merge_a3_see = merge_a3[desired_order]
        merge_a3_see.to_excel(f"{folder}\\merge_a3_see.xlsx", index=False)

        gc.collect()
        print("🍞🍞🍞")
        print("done calculating enervis with available SEE")
        print("🍞🍞🍞")

        # if there is no result from enervis, or only PV with SEE
    else:
        print("DataFrame df_enervis_pivot_filter does not exist or is empty. THERE IS NO result WIND ENERVIS, return grouping stammdaten")

        #IF THERE IS NO SEE WIND RESULT ENERVIS, return grouping stammdaten
        if 'merge_a3' not in locals():
            merge_a3_see = merge_a1_agg

            print("no result from enervis or only PV")


# IF THERE IS NO SEE DATA, automatically use wind stammdaten for enervis
no_see_data = df_stamm[pd.isna(df_stamm['unit_mastr_id']) | (df_stamm['unit_mastr_id'].str.strip() == "") | (df_stamm['unit_mastr_id'].str.lower() == "nan")].copy()
wind_stammdaten = df_stamm[
    ~pd.isna(df_stamm['manufacturer']) &
    ~pd.isna(df_stamm['turbine_model']) &
    ~(df_stamm['manufacturer'] == "") &
    ~(df_stamm['turbine_model'] == "")
].copy()


if 'merge_a3_see' in locals() and not merge_a3_see.empty and not wind.empty:
    s = merge_a3_see["avg_enervis"]

    # if there is no SEE data
    if s is None or s.empty:
        no_see_data = no_see_data

    # if there is SEE but no result from enervis
    else:
        s_str = s.astype("string")
        mask_empty = s_str.isna() | s_str.str.strip().str.lower().isin(["", "nan"])
        print(mask_empty)

        malo_need = (merge_a3_see.loc[mask_empty, "malo"].astype(str).str.strip())
        malo_need = malo_need[malo_need.ne("")].unique()
        print(malo_need)

        stamm_for_empty_enervis = df_stamm[df_stamm["malo"].astype(str).str.strip().isin(malo_need)].copy()

        if not stamm_for_empty_enervis.empty:
            no_see_data = pd.concat(
                [no_see_data, stamm_for_empty_enervis],
                ignore_index=True
            )

print(df_stamm[["malo", "unit_mastr_id", "tech"]])
print(no_see_data)
print("🥚🥚🥚")

# IF THERE IS NO SEE DATA
if not no_see_data.empty and not wind_stammdaten.empty:
    print("🥚🥚")
    solar = no_see_data[(no_see_data["tech"] == "PV")].copy()
    wind = no_see_data[(no_see_data["tech"] == "WIND")].copy()

    if not solar.empty:
        solar = solar

    if not wind.empty:
        columns_to_use = [
        "malo",
        'unit_mastr_id', 'Projekt', 'manufacturer', 'turbine_model',
        'hub_height_m', 'tech', 'latitude', 'longitude', 'net_power_kw_unit'
        ]

        df_combined = wind[columns_to_use]
        df_combined['net_power_mw'] = df_combined['net_power_kw_unit'] / 1000

        df_combined["net_power_kw"] = df_combined["net_power_kw_unit"]

        #Load the turbine reference turbine id from enervis
        df_ref_enervis = pd.read_excel(
            r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\Enervis template\turbine_types_id_enervis_eraseMW.xlsx")
        df_ref_enervis['id'] = df_ref_enervis['id'].astype(str).str.strip()

        def clean_name(name):
            name = str(name).lower()
            for remove_word in [
                'gmbh', 'se', 'deutschland', 'central europe', 'energy',
                'gmbh & co. kg', 'energy gmbh', 'deutschland gmbh', "GmbH", "GmbH & Co. KG", "Deutschland GmbH", "AG"
            ]:
                name = name.replace(remove_word.lower(), '')
            return name.strip()

        # 2. Clean turbine model for some manufacturers
        def clean_name_turbine(name):
            if not isinstance(name, str):
                return ''
            for remove_word in ['senvion', 'Senvion', 'nercon', 'nercon', 'mit Serrations', 'Vensys', 'NEG MICON', 'Delta4000', 'Delta 4000']:
                name = name.replace(remove_word, '')
            return name.strip()

        df_combined['clean_manufacturer'] = df_combined['manufacturer'].apply(clean_name)
        #df_combined = df_combined.dropna(subset=['turbine_model', 'hub_height_m'])
        df_combined['turbine_model_clean'] = df_combined['turbine_model'].apply(clean_name_turbine)
        df_combined['add_turbine'] = df_combined['turbine_model']

        df_combined.loc[
            df_combined['manufacturer'].isin(['Vestas Deutschland GmbH', 'Senvion Deutschland GmbH', 'ENERCON GmbH', 'VENSYS Energy AG', 'Enron Wind GmbH', 'NEG Micon Deutschland GmbH']),
            'add_turbine'
        ] = df_combined['turbine_model_clean'].astype(str).str.strip() + ' ' + df_combined['net_power_mw'].round(3).astype(str) + 'MW'

        df_combined.loc[
            df_combined['manufacturer'].isin(['Nordex Energy GmbH' , 'REpower Systems SE', 'Nordex Germany GmbH', "eno energy GmbH"]),
            'add_turbine'
        ] = df_combined['turbine_model_clean'].astype(str).str.strip() + ' ' + df_combined['net_power_kw'].astype(str)

        df_combined.loc[
            df_combined['manufacturer'].isin(['REpower Systems SE']),
            'add_turbine'
        ] = df_combined['turbine_model_clean'].astype(str).str.strip() + ' ' + df_combined['hub_height_m'].astype(str)

        # Apply fuzzy match only if not in hardcoded list
        def match_add_turbine_with_exceptions(row, choices, threshold=76):
            turbine_model = row['turbine_model']
            add_turbine = row['add_turbine']

            # Check for hardcoded rule first
            if turbine_model in hardcoded_map:
                return hardcoded_map[turbine_model]

            if add_turbine in hardcoded_map:
                return hardcoded_map[add_turbine]

            # Fallback to fuzzy match
            if isinstance(add_turbine, str) and add_turbine.strip():
                #match, score = process.extractOne(add_turbine, choices, scorer=fuzz.token_sort_ratio)
                match, score = process.extractOne(add_turbine, choices, scorer=fuzz.token_set_ratio)
                if score >= threshold:
                    return match
            return None

        # Apply the function to each row
        df_combined['Matched_Turbine_Name'] = df_combined.apply(
            lambda row: match_add_turbine_with_exceptions(row, df_ref_enervis['name'].dropna().unique()),
            axis=1
        )

        name_to_id_map = df_ref_enervis.set_index('name')['id'].to_dict()
        df_combined['Matched_Turbine_ID'] = df_combined['Matched_Turbine_Name'].map(name_to_id_map)

        del df_ref_enervis, name_to_id_map
        gc.collect()

        print("🍹🍹🍹")
        print(df_combined.head(50))

        print("🍲🍲🍲🍲🍲🍲🍲🍲")

        df_combined.to_excel(f"{folder}\\test_NaN_turbineID.xlsx", index=False)

        df_fuzzy_excel = df_combined.dropna(subset=["Matched_Turbine_ID", "hub_height_m"])

        #print(df_combined.sample(n=50))
        #print(df_combined.sample(n=50))

        print(df_fuzzy_excel.head(20))

        # 5. Show result
        df_fuzzy_excel['hub_height_m'] = df_fuzzy_excel['hub_height_m'].fillna(104).replace(0, 104)
        df_fuzzy_excel["hub_height_m"] = df_fuzzy_excel["hub_height_m"].astype(int)
        df_fuzzy_excel["Matched_Turbine_ID"] = df_fuzzy_excel["Matched_Turbine_ID"].astype(int)

        df_fuzzy_excel["hub_height_m"] = df_fuzzy_excel["hub_height_m"].apply(lambda x: max(x, 50))


        print("🥨🥨🥨")

        if pd.api.types.is_numeric_dtype(df_fuzzy_excel["malo"]):
            df_fuzzy_excel["malo"] = (
                df_fuzzy_excel["malo"]
                .astype(float)
                .astype("Int64")   # nullable int (keeps NaN)
                .astype(str)
            )
        else:
            # already string → just ensure dtype string
            df_fuzzy_excel["malo"] = df_fuzzy_excel["malo"].astype(str)

        df_fuzzy_excel.to_excel(f"{folder}\\df_fuzzy_excel.xlsx", index=False)

        #del df_fuzzy
        gc.collect()

        print("🥕🥕🍛🍛🍛🥕🥕🥕🥕")
        print(df_fuzzy_excel)
        print("🍛🍛🍛")

        EMAIL = "amani@flex-power.energy"
        PASSWORD = "ypq_CZE2wpg*jgu7hfk"

        # --- Step 1: Get access token ---
        def get_token():
            url = "https://keycloak.anemosgmbh.com/auth/realms/awis/protocol/openid-connect/token"
            data = {
                'client_id': 'webtool_vue',
                'grant_type': 'password',
                'username': EMAIL,
                'password': PASSWORD
            }
            response = requests.post(url, data=data)

            # Debug output
            print("Status code:", response.status_code)
            print("Response:", response.text)

            response.raise_for_status()
            return response.json()['access_token']

        # --- Step 2: Get historical On-Demand product ID ---
        def get_historical_product_id(token):
            url = "https://api.anemosgmbh.com/products_mva"
            headers = {"Authorization": f"Bearer {token}"}
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            products = response.json()

            print("\n📦 Available products:")
            for p in products:
                product_type_name = p["mva_product_type"]["name"]
                print(f"- ID: {p['id']}, Name: {product_type_name}")
                if "hist-ondemand" in product_type_name.lower():
                    print(f"✅ Found 'hist-ondemand' product.")
                    return p["id"]

            raise Exception("❌ No 'hist-ondemand' product found.")

        def list_turbine_types(token):
            url = "https://api.anemosgmbh.com/turbine_types"
            headers = {"Authorization": f"Bearer {token}"}
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            types = response.json()
            print("\n🌀 Available turbine types:")
            for t in types:
                print(f"- ID: {t['id']}, Name: {t.get('name', 'n/a')}")

            df = pd.DataFrame(types)
            # filename = r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\Enervis template\turbine_types_id_enervis.xlsx"
            print(f"✅ Turbine types saved to {filename}")

            return types

            #token = get_token()
            #list_turbine_types(token)

        def start_historical_job_from_df(token, product_id, df_input):
            url = "https://api.anemosgmbh.com/jobs"
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }

            # Build parkinfo list from DataFrame rows
            parkinfo_list = []
            for _, row in df_input.iterrows():
                if pd.notnull(row["Matched_Turbine_ID"]):
                    hub_height = row["hub_height_m"]
                    if pd.isna(hub_height):
                        hub_height = 104

                    parkinfo_list.append({
                        "id": int(row["malo"]),
                        "lat": str(row["latitude"]),
                        "lon": str(row["longitude"]),
                        "turbine_type_id": int(row["Matched_Turbine_ID"]),
                        "hub_height": int(hub_height)
                    })

            if not parkinfo_list:
                print("⚠️  Warning: No valid parkinfo entries found in input DataFrame. Skipping job.")
                return None

            payload = {
                "mva_product_id": product_id,
                "parameters": {
                    "parkinfo": parkinfo_list
                }
            }

            print("\n📤 Sending payload with multiple turbines:")
            print(payload)

            response = requests.post(url, headers=headers, json=payload)

            if response.status_code != 200:
                print("\n❌ API responded with:")
                print(response.text)
                response.raise_for_status()

            resp_json = response.json()
            return resp_json["uuid"]

        # --- Step 4: Poll the job status until complete ---
        def wait_for_job_completion(token, job_uuid, poll_interval=10):
            url = f"https://api.anemosgmbh.com/jobs/{job_uuid}"

            while True:
                headers = {"Authorization": f"Bearer {token}"}
                response = requests.get(url, headers=headers)

                if response.status_code == 401:
                    # Token expired - get new token and retry once
                    print("Token expired, refreshing token...")
                    token = get_token()
                    headers = {"Authorization": f"Bearer {token}"}
                    response = requests.get(url, headers=headers)

                response.raise_for_status()
                job_info = response.json()

                if isinstance(job_info, list):
                    job_info = job_info[0]

                status = job_info.get("status")
                print(f"Job status: {status}")

                if status == "DONE" or status == "COMPLETED":
                    return job_info
                elif status in ["FAILED", "CANCELED"]:
                    raise Exception(f"Job ended with status: {status}")

                time.sleep(poll_interval)


        # --- Step 5: Download result files and load them into DataFrames ---
        def download_result_files(job_info, token):
            headers = {"Authorization": f"Bearer {token}"}
            files = job_info.get("files")

            if files:
                dfs = []
                for f in files:
                    file_url = f.get("url")
                    print(f"📥 Downloading result file: {file_url}")
                    df = download_and_load_csv(file_url, token)
                    dfs.append(df)
                return dfs
            else:
                print("❌ No result files found — checking 'info' field...")
                results = job_info.get("info", {}).get("results", [])

                if results:
                    dfs = []
                    for result in results:
                        turbine_id = result.get("id")
                        year_data = result.get("Marktwertdifferenzen")
                        if year_data:
                            df = pd.DataFrame.from_dict(year_data, orient="index", columns=["Marktwertdifferenz"])
                            df.index.name = "Year"
                            df = df.reset_index()
                            df["id"] = turbine_id
                            dfs.append(df)
                            print("📊 Extracted results from 'info':")
                            print(df)
                        else:
                            print("❌ No 'Marktwertdifferenzen' found in result.")
                    return dfs if dfs else None
                else:
                    print("❌ No usable results found in 'info'.")
                    return None


        df_enervis_pivot = pd.DataFrame()

        # --- Main workflow ---
        if __name__ == "__main__":
            try:
                print("🔐 Getting access token...")
                token = get_token()

                print("📦 Getting historical product ID...")
                product_id = get_historical_product_id(token)

                print("🚀 Starting historical On-Demand job...")
                job_uuid = start_historical_job_from_df(token, product_id, df_fuzzy_excel)

                if job_uuid is None:
                    print("⏭️  No valid turbines to process. Skipping job execution and download.")
                    # You can choose to set an empty DataFrame or take other actions here
                    all_df = pd.DataFrame()  # Create an empty DataFrame to avoid errors later
                else:
                    print(f"✅ Job started with UUID: {job_uuid}")

                    print("⏳ Waiting for job completion...")
                    job_info = wait_for_job_completion(token, job_uuid)

                    print("📁 Job finished! Getting result files...")
                    dfs = download_result_files(job_info, token)

                    all_df = pd.concat(dfs, ignore_index=True)
                    del dfs
                    gc.collect()

                    # Convert 'Year' to string once for consistent processing
                    all_df["Year"] = all_df["Year"].astype(str)
                    print("Unique years in raw data:", all_df["Year"].unique())

                    target_years = ["2021", "2023", "2024"]
                    existing_years = all_df["Year"].unique().tolist()

                    # Filter target_years that actually exist in data
                    valid_years = [y for y in target_years if y in existing_years]
                    print("Valid years to process:", valid_years)

                    if not valid_years:
                        raise ValueError("❌ No target years found in data")

                    # Filter rows to keep only valid years
                    all_df = all_df[all_df["Year"].isin(valid_years)].copy()
                    print("\nData after year filtering:")
                    print(all_df[["id", "Year", "Marktwertdifferenz"]].head(20))
                    print(all_df[["id", "Year", "Marktwertdifferenz"]].tail(20))
                    all_df.to_excel(f"{folder}\\all_df.xlsx", index=False)


                    # Step 1: Filter to keep only the minimum Marktwertdifferenz per (id, Year)
                    df_filtered = all_df.loc[
                        all_df.groupby(["id", "Year"])["Marktwertdifferenz"].idxmin()
                    ].copy()

                    df_filtered['Marktwertdifferenz'] = df_filtered['Marktwertdifferenz'].round(2)

                    del all_df
                    gc.collect()

                    # Step 2: Pivot to wide format
                    df_enervis_pivot = df_filtered.pivot(
                        index="id",
                        columns="Year",
                        values="Marktwertdifferenz"
                    ).rename_axis(None, axis=1).reset_index()

                    del df_filtered
                    gc.collect()

                    # Step 3: Ensure all year columns are present
                    for year in target_years:
                        if year not in df_enervis_pivot.columns:
                            df_enervis_pivot[year] = np.nan

                    # Check which valid_years are really in the pivot columns
                    available_columns = [col for col in valid_years if col in df_enervis_pivot.columns]
                    print("Available columns:", available_columns)

                    if not available_columns:
                        raise ValueError("❌ No valid year columns found in pivot table")

                    # Calculate average across available year columns
                    df_enervis_pivot["avg_enervis"] = (df_enervis_pivot[target_years].mean(axis=1, skipna=True)).round(2)
                    columns_to_keep = ["id"] + target_years + ["avg_enervis"]
                    df_enervis_pivot_filter = df_enervis_pivot[columns_to_keep]

                    del df_enervis_pivot
                    gc.collect()
                    print("🥥🥥 df_enervis_pivot_filter")
                    print(df_enervis_pivot_filter)
                    df_enervis_pivot_filter.to_excel(f"{folder}\\df_enervis_pivot_filter_0_500.xlsx", index=False)

            except Exception as e:
                print(f"❌ Error: {e}")
                raise

    # IF THERE IS ENERVIS RESULT FROM MANUAL SPEC
    if 'df_enervis_pivot_filter' in locals() and "merge_a1" in locals() and not df_enervis_pivot_filter.empty:
        df_enervis_pivot_filter['id'] = df_enervis_pivot_filter['id'].astype(str)

        df_stamm["Power in MW"] = df_stamm["net_power_kw_unit"] / 1000
        merge_a1_agg = merge_a1.groupby(['malo'], dropna=False).agg({
                'unit_mastr_id': 'first',
                'Projekt': 'first',
                'tech': 'first',
                "Power in MW": 'sum',
                "INB": lambda x: [convert_date_or_keep_string(date) for date in x],
                #"EEG": list,
                "EEG": lambda x: list(x.unique()),  # Get distinct EEG values
                #'Vermarktung': lambda x: list(x.unique()),
                'AW in EUR/MWh': lambda x: [round(float(val), 2) for val in x if is_number(val)],
                #"Gesetzliche Vergütung (EEG) inkl. Managementprämie" : lambda x: [round(float(val), 2) for val in x if is_number(val)],
                "weighted_2021_eur_mwh_blindleister": "min",
                "weighted_2023_eur_mwh_blindleister": "min",
                "weighted_2024_eur_mwh_blindleister": "min",
                "average_weighted_eur_mwh_blindleister": "min",
                "Curtailment & redispatch included": 'first',
                "Balancing Cost": 'first',
                "Curtailment_value_weighted" : 'first',
                "Trading Convenience": 'first',
            }).reset_index()

        merge_a3_no_see = pd.merge(
            merge_a1_agg,
            df_enervis_pivot_filter,
            left_on = ('malo'),
            right_on = ('id'),
            how='left'
        )
        merge_a3_no_see.drop(columns=['id'], inplace=True)
        merge_a3_no_see = merge_a3_no_see.dropna(subset=['2021', '2023', "2024", "avg_enervis"])

        #del merge_a2, df_enervis_pivot_filter
        gc.collect()

        print(merge_a3_no_see)
        merge_a3_no_see.to_excel(f"{folder}\\merge_a3_no_see.xlsx", index=False)


    # IF ONLY MANUAL ENERVIS RESULT, without SEE, no blindleister result
    if 'df_enervis_pivot_filter' in locals() and not df_enervis_pivot_filter.empty and "merge_a1" not in locals():
        df_enervis_pivot_filter['id'] = df_enervis_pivot_filter['id'].astype(str)

        df_stamm["Power in MW"] = df_stamm["net_power_kw_unit"] / 1000

        merge_a1_agg = df_stamm.groupby(['malo'], dropna=False).agg({
                'unit_mastr_id': 'first',
                'Projekt': 'first',
                'tech': 'first',
                "Power in MW": 'sum',
                "INB": lambda x: [convert_date_or_keep_string(date) for date in x],
                "EEG": lambda x: list(x.unique()),  # Get distinct EEG values
                'AW in EUR/MWh': lambda x: [round(float(val), 2) for val in x if is_number(val)],
                "Curtailment & redispatch included": 'first',
                "Balancing Cost": 'first',
                "Curtailment_value_weighted" : 'first',
                "Trading Convenience": 'first',
            }).reset_index()

        merge_a3_no_see = pd.merge(
            merge_a1_agg,
            df_enervis_pivot_filter,
            left_on = ('malo'),
            right_on = ('id'),
            how='left'
        )
        merge_a3_no_see.drop(columns=['id'], inplace=True)
        merge_a3_no_see = merge_a3_no_see.dropna(subset=['2021', '2023', "2024", "avg_enervis"])

        #del merge_a2, df_enervis_pivot_filter
        gc.collect()

        print(merge_a3_no_see)
        merge_a3_no_see.to_excel(f"{folder}\\merge_a3_no_see_0_500.xlsx", index=False)


# Check if both DataFrames are not empty before merging
## merging see and no see result
if 'merge_a3_see' in locals() and not merge_a3_see.empty and 'merge_a3_no_see' in locals() and not merge_a3_no_see.empty:
    print(merge_a3_see)
    print(merge_a3_no_see)
    merge_a3 = pd.merge(
        merge_a3_see,
        merge_a3_no_see[['malo', '2021', '2023', '2024', 'avg_enervis']],
        on='malo',
        how='left',
        suffixes=('_see', '_no_see')  # Add suffixes to differentiate columns
    )

    # Fill missing values in 'merge_a3_see' columns with corresponding values from 'merge_a3_no_see'
    columns_to_fill = ['2021', '2023', '2024', 'avg_enervis']

    for column in columns_to_fill:
        # Fill NaN values in the '_see' columns with values from the '_no_see' columns directly into the same column
        merge_a3[column] = merge_a3[column + '_see'].fillna(merge_a3[column + '_no_see'])

    # Drop the extra columns created during the merge (columns with suffixes '_see' and '_no_see')
    merge_a3.drop(columns=[col + '_see' for col in columns_to_fill] + [col + '_no_see' for col in columns_to_fill], inplace=True)

    merge_a3.reset_index(drop=True, inplace=True)

    print("🥑")
    print(merge_a3)
    merge_a3.to_excel(f"{folder}\\merge_a3_merge.xlsx", index=False)

elif 'merge_a3_see' in locals() and not merge_a3_see.empty:
    # If only merge_a3_see is not empty, assign it to merge_a3
    merge_a3 = merge_a3_see
    print("🥑🥑")
    print(merge_a3)

elif 'merge_a3_no_see' in locals() and not merge_a3_no_see.empty:
    # If only merge_a3_no_see is not empty, assign it to merge_a3
    merge_a3 = merge_a3_no_see
    print("🥑🥑🥑")
    print(merge_a3)
else:
    print("🧮🧮🧮 Both DataFrames are empty or not defined. Merge not performed, continue to only prod calculation 🧮🧮🧮")


xls = pd.ExcelFile(path, engine='openpyxl')
sheet_names = xls.sheet_names

# if only stammdaten data, without production
if len(sheet_names) == 1 and sheet_names[0].lower() == 'stammdaten':

    merge_a3 = merge_a3.rename(columns={'tech': 'Technology'})
    sheet1_order = [
                'malo',
                'Projekt',
                'Technology',
                'Power in MW',
                'INB',
                #"Vermarktung",
                'EEG',
                'AW in EUR/MWh',
                'weighted_2021_eur_mwh_blindleister',
                'weighted_2023_eur_mwh_blindleister',
                'weighted_2024_eur_mwh_blindleister',
                'average_weighted_eur_mwh_blindleister',
                "2021",
                "2023",
                "2024",
                "avg_enervis",
                "Curtailment & redispatch included",
                "Balancing Cost",
                "Curtailment_value_weighted",
                "AMW/RWM Delta",
                "Fair Value",
                "Trading Convenience"
            ]

    sheet2_order = [
                'malo',
                'Projekt',
                'Technology',
                'Power in MW',
                'INB',
                #"Vermarktung",
                'EEG',
                'AW in EUR/MWh',
                'average_weighted_eur_mwh_blindleister',
                "avg_enervis",

                "Curtailment & redispatch included",

                "Balancing Cost",
                "Curtailment_value_weighted",
                "AMW/RWM Delta",
                "Fair Value",
                "Trading Convenience",
                "Fee EUR/MWh"
            ]

    sheet3_order = [
                'malo',
                'Projekt',
                "Technology",
                'Power in MW',
                'AW in EUR/MWh',
                "Fee EUR/MWh"
            ]


    # Prepare each sheet
    sheet1_df = ensure_and_reorder(merge_a3.copy(), sheet1_order)
    sheet2_df = ensure_and_reorder(merge_a3.copy(), sheet2_order)
    sheet3_df = ensure_and_reorder(merge_a3.copy(), sheet3_order)

    customer_name = os.path.basename(os.path.dirname(out_path))

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sheet1_df.to_excel(writer, sheet_name=f"{customer_name}_1", index=False)
        sheet2_df.to_excel(writer, sheet_name=f"{customer_name}_2", index=False)
        sheet3_df.to_excel(writer, sheet_name=f"{customer_name}_3", index=False)

    print("🍚🦐🦐🍚")

# if both stammdaten and production data
elif len(sheet_names) > 1 and sheet_names[0].lower() == 'stammdaten':
    print("🦐🦐🦐🦐")
    print("production calculation start")
    print("🦐🦐🦐🦐")

    # 1. find all sheets except “stammdaten”
    sheets = [s for s in xls.sheet_names if s.lower() not in ["stammdaten", "curtailment", "redispatch_wind", "redispatch_pv", "redispatch"]]

    batch_size = 5
    merged = []

    for i in range(0, len(sheets), batch_size):
        batch = sheets[i : i + batch_size]
        dfs = []

        for sh in batch:
            df_tmp = pd.read_excel(path, sheet_name=sh, engine="openpyxl")
            df_tmp.columns = df_tmp.columns.str.strip()  # clean column names

            if "malo" in df_tmp.columns:
                df_tmp["malo"] = (
                    df_tmp["malo"].astype("string")
                    .str.replace(r"\D+", "", regex=True)  # keep digits only
                    .fillna("")
                    .str.strip()
                )
            dfs.append(df_tmp)

        merged.append(pd.concat(dfs, ignore_index=True))
        del dfs
        gc.collect()

    df = pd.concat(merged, ignore_index=True)
    del merged
    gc.collect()

    df.columns = df.columns.str.strip()
    df['malo'] = df['malo'].apply(lambda x: str(int(x)) if isinstance(x, (float, int)) and pd.notna(x) else str(x)).str.strip()

    def process_time_column(df):
        """
        Process time column automatically based on column name:
        - If 'time_berlin': parse as Berlin time and remove timezone
        - If 'time_utc': parse as UTC, convert to Berlin time, then drop UTC column
        """
        df_result = df.copy()

        if 'time_berlin' in df_result.columns:
            # Process as Berlin time
            df_result['time_berlin'] = pd.to_datetime(df_result['time_berlin'], dayfirst=True, errors='coerce')
            df_result['time_berlin'] = df_result['time_berlin'].dt.tz_localize(None)
            print("🍠🍠 Processed time_berlin column (removed timezone)")

        elif 'time_utc' in df_result.columns:
            # Process as UTC and convert to Berlin time
            german_tz = pytz.timezone("Europe/Berlin")
            df_result['time_utc'] = pd.to_datetime(df_result['time_utc'], errors='coerce', dayfirst=True, utc=True)
            df_result['time_berlin'] = df_result['time_utc'].dt.tz_convert(german_tz).dt.tz_localize(None)
            df_result = df_result.drop(columns='time_utc')
            print("🍠🍠🍠🍠 Processed time_utc column (converted to Berlin time and dropped UTC column)")

        else:
            print("No time_berlin or time_utc column found")

        return df_result

    df = process_time_column(df)

    print(df)
    print(df.info())

    # Number of rows per full year of quarter-hourly data
    rows_per_full_year = 35000  # 35000

    df_filtered = df[df['time_berlin'].dt.year.isin([2021, 2023, 2024, 2025])]

    filtered_data = []

    # Iterate over each malo group
    for malo, group in df_filtered.groupby('malo'):
        # Filter the group again to include only the years 2021, 2023, 2024, and 2025
        group_filtered = group[group['time_berlin'].dt.year.isin([2021, 2023, 2024, 2025])]

        # Calculate the number of rows (quarter-hourly data points) for each year
        rows_per_year = group_filtered.groupby(group_filtered['time_berlin'].dt.year).size()
        counting_month = group_filtered.groupby(group_filtered['time_berlin'].dt.to_period('M')).size()
        total_rows_across_years = rows_per_year.sum()

        # Print the rows per year to check if they match the full year requirement
        print(f"malo: {malo}, rows per year: {rows_per_year}")

        # Get the years in the current group
        years_in_data = group_filtered['time_berlin'].dt.year.unique()
        unique_months = group_filtered['time_berlin'].dt.to_period('M').unique()

        valid_months = counting_month[counting_month >= 2592].index  # Get months with at least 2592 rows, 27 days
        group_filtered_valid_months = group_filtered[group_filtered['time_berlin'].dt.to_period('M').isin(valid_months)]

        filtered_group = None

        # If the group has exactly 1 year of data, keep it all (no filtering needed)
        if len(years_in_data) == 1:
            print(f"🍠 malo: {malo} has only 1 year of data, keeping all data.\n")
            filtered_group = group_filtered_valid_months

        elif len(years_in_data) > 2 and any(rows_per_year[rows_per_year.index.isin(years_in_data)] >= rows_per_full_year):
            # Check if each year has a full year's worth of data (based on the row count)
            full_years_available = rows_per_year[rows_per_year >= rows_per_full_year].index.tolist()

            # Debugging: Print which full years are available
            print(f"🔑🔑🔑 malo: {malo}, full years available: {full_years_available}\n")
            # If full year data exists, keep the group and filter for the full year data
            filtered_group = group_filtered_valid_months[group_filtered_valid_months['time_berlin'].dt.year.isin(full_years_available)]

            available_month_after = filtered_group.shape[0]
            available_years = filtered_group['time_berlin'].dt.year.unique().tolist()
            filtered_group['available_years'] = ', '.join(map(str, available_years))
            filtered_data.append(filtered_group)

        elif len(years_in_data) >= 2 and any(rows_per_year[rows_per_year.index.isin(years_in_data)] < rows_per_full_year):
            # if malo has 2 years with less or equal than 12 months data
            print(f"🍠🍠 malo: {malo} has >= 2 years of data, with any of them does not have full year\n")

            # Get unique months and sort them
            unique_months_sorted = sorted(group_filtered_valid_months['time_berlin'].dt.to_period('M').unique())

            # Find ALL continuous periods (12-month, 24-month, 36-month periods)
            continuous_periods = []

            # Check for continuous periods of various lengths (12, 24, 36 months, etc.)
            for period_length in [12, 24, 36]:
                for i in range(len(unique_months_sorted) - period_length + 1):
                    start_month = unique_months_sorted[i]
                    # Check if we have exactly 'period_length' consecutive months
                    if all((start_month + j) in unique_months_sorted for j in range(period_length)):
                        continuous_periods.append([start_month + j for j in range(period_length)])

            if continuous_periods:
                # Prefer the MOST RECENT continuous period
                most_recent_period = max(continuous_periods, key=lambda x: x[0])

                print(f"✅✅ 🍠🍠  malo: {malo} has {len(continuous_periods)} continuous period(s)")
                print(f"Using most recent: {most_recent_period[0]} to {most_recent_period[-1]}\n")

                filtered_group = group_filtered_valid_months[
                    group_filtered_valid_months['time_berlin'].dt.to_period('M').isin(most_recent_period)
                ]
            else:
                print(f"⚠️⚠️ 🍠🍠 malo: {malo} has no continuous periods, keeping all available months\n")
                filtered_group = group_filtered_valid_months

        elif all(rows_per_year < rows_per_full_year):
            print(f"🚨🚨🚨🚨🚨🚨 malo: {malo} has missing rows for each year, keeping all data with full months\n")
            print(valid_months)

            # Get unique months and sort them
            unique_months_sorted = sorted(group_filtered_valid_months['time_berlin'].dt.to_period('M').unique())

            # Find ALL continuous periods (12-month, 24-month, 36-month periods)
            continuous_periods = []

            # Check for continuous periods of various lengths (12, 24, 36 months, etc.)
            for period_length in [12, 24, 36]:
                for i in range(len(unique_months_sorted) - period_length + 1):
                    start_month = unique_months_sorted[i]
                    # Check if we have exactly 'period_length' consecutive months
                    if all((start_month + j) in unique_months_sorted for j in range(period_length)):
                        continuous_periods.append([start_month + j for j in range(period_length)])

            if continuous_periods:
                # Prefer the MOST RECENT continuous period
                most_recent_period, chosen_period_length = max(continuous_periods, key=lambda x: x[0][0])

                print(f"✅✅  malo: {malo} has {len(continuous_periods)} continuous period(s)")
                print(f"Using most recent: {most_recent_period[0]} to {most_recent_period[-1]}")
                print(f"Chosen period length: {chosen_period_length} months")

                filtered_group = group_filtered_valid_months[
                    group_filtered_valid_months['time_berlin'].dt.to_period('M').isin(most_recent_period)
                ]
            else:
                print(f"⚠️⚠️  malo: {malo} has no continuous periods, keeping all available months")
                filtered_group = group_filtered_valid_months

        if filtered_group is not None and not filtered_group.empty:
            # Add available_years information
            available_years = filtered_group['time_berlin'].dt.year.unique().tolist()
            filtered_group = filtered_group.copy()  # Avoid SettingWithCopyWarning
            filtered_group['available_years'] = ', '.join(map(str, available_years))
            filtered_data.append(filtered_group)
            # print(f"✅ Appended malo: {malo} with {len(filtered_group)} rows")
        else:
            print(f"❌ No data to append for malo: {malo}")

    if filtered_data:
        df_source = pd.concat(filtered_data, ignore_index=True)

        # Add available_month column: count unique months per malo
        df_source['month'] = df_source['time_berlin'].dt.to_period('M')
        month_counts = df_source.groupby('malo')['month'].nunique().reset_index(name='available_months')

        # Merge back to df_source
        df_source = df_source.merge(month_counts, on='malo', how='left')

        # Drop helper column
        df_source.drop(columns='month', inplace=True)

        print(f"✅ Final concatenated data shape: {df_source.shape}")
    else:
        print("No data to concatenate. All groups were discarded.")

    print("🦐🦐🦐🦐")
    print(df["malo"].nunique())
    print(df_source["malo"].nunique())
    print(df_source["malo"].unique())
    # print(df_source)

    # SUMMING REDISPATCH
    # Load the workbook to check available sheets
    wb = load_workbook(path, read_only=True)
    available_sheets = wb.sheetnames
    wb.close()

    df_source.rename(columns={'power_kwh': 'infeed_kwh'}, inplace=True)
    df_source['__adj_kwh'] = 0.0


    # Check if 'redispatch' sheet exists
    if 'redispatch' in available_sheets:
        df_redispatch = pd.read_excel(path, sheet_name="redispatch", engine='openpyxl')
        df_redispatch.columns = df_redispatch.columns.str.strip()
        df_redispatch['malo'] = df_redispatch['malo'].apply(lambda x: str(int(x)) if isinstance(x, (float, int)) and pd.notna(x) else str(x)).str.strip()

        # # df_redispatch['time_berlin'] = pd.to_datetime(df_redispatch['time_berlin'], dayfirst=True, errors='coerce')

        # german_tz = pytz.timezone("Europe/Berlin")
        # df_redispatch['time_utc'] = pd.to_datetime(df_redispatch['time_utc'], errors='coerce', dayfirst=True, utc = True)
        # df_redispatch['time_berlin'] = df_redispatch['time_utc'].dt.tz_convert(german_tz).dt.tz_localize(None)
        # df_redispatch = df_redispatch.drop(columns='time_utc')

        df_redispatch = process_time_column(df_redispatch)

        df_source = pd.merge(
            df_source,
            df_redispatch[['malo', 'time_berlin', 'redispatch_kwh']],
            on=['malo', 'time_berlin'],
            how='left'
        )
        df_source['__adj_kwh'] = df_source['__adj_kwh'] + df_source['redispatch_kwh'].fillna(0)

    if 'curtailment' in available_sheets:
        df_curtailment = pd.read_excel(path, sheet_name="curtailment", engine='openpyxl')
        df_curtailment.columns = df_curtailment.columns.str.strip()
        df_curtailment['malo'] = df_curtailment['malo'].apply(lambda x: str(int(x)) if isinstance(x, (float, int)) else str(x)).str.strip()

        # df_curtailment['time_berlin'] = pd.to_datetime(df_curtailment['time_berlin'], dayfirst=True, errors='coerce')

        # german_tz = pytz.timezone("Europe/Berlin")
        # df_curtailment['time_utc'] = pd.to_datetime(df_curtailment['time_utc'], errors='coerce', dayfirst=True, utc = True)
        # df_curtailment['time_berlin'] = df_curtailment['time_utc'].dt.tz_convert(german_tz).dt.tz_localize(None)
        # df_curtailment = df_curtailment.drop(columns='time_utc')

        df_curtailment = process_time_column(df_curtailment)

        df_source = pd.merge(
            df_source,
            df_curtailment[['malo', 'time_berlin', 'curtailment_kwh']],
            on=['malo', 'time_berlin'],
            how='left'
        )
        df_source['curtailment_kwh'] = df_source['curtailment_kwh'].fillna(0)
        df_source['__adj_kwh'] = df_source['__adj_kwh'] + df_source['curtailment_kwh']

        # Check which malo has curtailment data
        df_source['curtailment'] = df_source['malo'].isin(df_curtailment['malo'])

    df_source['infeed_kwh'] = pd.to_numeric(df_source['infeed_kwh'], errors='coerce')
    df_source['__adj_kwh'] = pd.to_numeric(df_source['__adj_kwh'], errors='coerce')

    df_source['power_kwh'] = df_source['infeed_kwh'] + df_source['__adj_kwh']
    print(df_source.info())

    # excel_file_path = f"{folder}\\df_source_avg.xlsx"
    # rows_per_sheet = 1000000
    # with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    #     for i in range(0, len(df_source), rows_per_sheet):
    #         chunk = df_source.iloc[i:i + rows_per_sheet]
    #         sheet_name = f'Sheet_{i//rows_per_sheet + 1}'
    #         chunk.to_excel(writer, sheet_name=sheet_name, index=False)


    print("🦐🦐")

    df_rmv = pd.read_csv(r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\DA price\rmv_price.csv")
    df_dayahead = pd.read_csv(r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\DA price\DA_price.csv")

    print("🦐🦑🦑🦐")

    grouped = df_source.groupby(["malo", "time_berlin", "available_years", "available_months"])

    def custom_power_mwh(group):
        if group.nunique() == 1:
            return group.mean()
        else:
            return group.sum()

    df_source_avg = grouped["power_kwh"].apply(custom_power_mwh).reset_index()
    # df_source_avg["power_kwh"] = df_source_avg["power_mw"] * 1000 / 4
    # df_source_avg = df_source_avg.drop("power_mw", axis='columns')

    print("Columns of df_source_avg:", df_source_avg.columns)
    print("🥟🥟🥟🥟🥟")

    df_dayahead['delivery_start__utc_'] = pd.to_datetime(df_dayahead['delivery_start__utc_'], utc=True)
    # Convert from UTC to Europe/Berlin
    df_dayahead['time_berlin'] = df_dayahead['delivery_start__utc_'].dt.tz_convert('Europe/Berlin')
    # Remove timezone info to make it naive (e.g., for grouping)
    df_dayahead['naive_time'] = df_dayahead['time_berlin'].dt.tz_localize(None)
    # Group by the naive Berlin-local time and average
    df_dayahead_avg = df_dayahead.groupby('naive_time', as_index=False)['dayaheadprice'].mean()
    # Rename for clarity
    df_dayahead_avg = df_dayahead_avg.rename(columns={'naive_time': 'time_berlin'})

    print(df_dayahead_avg)
    print("🧋🧋🧋")

    df_dayahead_avg = df_dayahead_avg.drop_duplicates(subset=["time_berlin","dayaheadprice"])
    df_dayahead_avg_indexed = df_dayahead_avg.set_index('time_berlin')

    del df_dayahead_avg

    print("🍌🍌🍌")


    pattern = r'\d+.*rules'
    # Create temporary sort priority column
    df_temp = df_stamm.copy()
    df_temp['_sort_priority'] = df_temp['category'].str.contains(pattern, na=False)

    # Sort by malo and sort_priority (rules first within each malo)
    df_sorted = df_temp.sort_values(['malo', '_sort_priority'], ascending=[True, False])

    df_assets_mapping = df_sorted.groupby(['malo'], dropna=False).agg({
                'tech': 'first',
                'net_power_kw_unit': 'sum',
                "category": "first",
            }).reset_index()

    df_assets_mapping = df_assets_mapping.drop(columns=['_sort_priority'], errors='ignore')

    df_source_avg = df_source_avg.merge(
        df_assets_mapping,
        on='malo', how='left')

    del df_sorted
    gc.collect()
    print("🫚 after merge")


    print("🫚🫚 after filtering complete months")

    def expand_by_date_threshold(df, cutoff_date='2025-10-01'):
        """
        Expand hourly data before cutoff date, leave quarter-hourly data after cutoff as-is
        """
        cutoff = pd.to_datetime(cutoff_date)

        # Split data
        hourly_data = df[df.index < cutoff]
        quarter_hourly_data = df[df.index >= cutoff]

        # Only expand the hourly part
        if not hourly_data.empty:
            hourly_expanded = hourly_data.resample('15T').ffill()
            # Combine results
            result = pd.concat([hourly_expanded, quarter_hourly_data])
        else:
            result = quarter_hourly_data

        return result.sort_index()

    # Apply date-based expansion
    df_dayahead_avg = expand_by_date_threshold(df_dayahead_avg_indexed, '2025-10-01').reset_index()

    dayaheadprice_production_merge = pd.merge(
                df_source_avg,
                df_dayahead_avg,
                on= 'time_berlin',
                how='inner'
            )

    print("🌶️🌶️🌶️")
    print(dayaheadprice_production_merge.sample(n=20))
    print(dayaheadprice_production_merge.head())
    print("🫚🫚🫚 after joining day‐ahead")

    del df_dayahead_avg
    gc.collect()


    dayaheadprice_production_merge['year']  = dayaheadprice_production_merge['time_berlin'].dt.year.astype('int16')
    dayaheadprice_production_merge['month'] = dayaheadprice_production_merge['time_berlin'].dt.month.astype('int8')

    dayaheadprice_production_merge["tech"] = dayaheadprice_production_merge["tech"].str.strip().str.upper().astype("category")
    df_rmv["tech"] = df_rmv["tech"].str.strip().str.upper().astype("category")

    merge_prod_rmv_dayahead = dayaheadprice_production_merge.merge(
                df_rmv,
                on=["tech","year","month"],
                how="left",
            )

    print("🫚🫚🫚🫚 after joining rmv")
    del dayaheadprice_production_merge, df_rmv
    gc.collect()

    print("🫕🫕🫕")
    df_ts = merge_prod_rmv_dayahead[merge_prod_rmv_dayahead['time_berlin'].dt.year.isin([2023, 2024, 2025])].copy()

    if "curtailment" in df_source.columns:
        df_ts = df_ts[~df_ts['malo'].isin(df_source[df_source['curtailment']]['malo'])]

    from sklearn.metrics import (
        accuracy_score, precision_score, recall_score, f1_score,
        roc_auc_score, average_precision_score, confusion_matrix,
        precision_recall_curve, roc_curve,
        mean_squared_error, mean_absolute_error,
        mean_absolute_percentage_error, r2_score
    )

    project_id = "flex-power"
    forecast_table = """
    SELECT
        *
    FROM `flex-power.sales.price_volume_data_for_curtailment_forecast_table`
    ORDER BY delivery_start_berlin
    """


    df_forecast_table = pandas_gbq.read_gbq(forecast_table, project_id=project_id)
    df_forecast_table['delivery_start_berlin'] = pd.to_datetime(df_forecast_table['delivery_start_berlin'], errors='coerce')

    df_ts = df_ts.merge(
        df_forecast_table,
        left_on=["time_berlin","tech"],
        right_on=["delivery_start_berlin","tech"],
        how="left",
        )

    print("🫕🫕🫕")
    print(df_ts.info())
    print(df_ts)
    print("🫕🫕🫕")

    # =============================================================================
    # SMALL HELPERS
    # =============================================================================

    def print_header(title: str):
        print("\n" + "=" * 80)
        print(title)
        print("=" * 80)


    # =============================================================================
    # CLASSIFICATION IMPLEMENTATION
    # =============================================================================

    def feature_engineering_classification(df: pd.DataFrame, feature_names):
        """Feature engineering for classification – must match training."""
        df = df.copy()

        # volume__mw_imbalance
        if "volume__mw_imbalance" in df.columns:
            df["volume__mw_imbalance"] = pd.to_numeric(df["volume__mw_imbalance"], errors="coerce").fillna(0)
        else:
            df["volume__mw_imbalance"] = 0.0

        # Target flag if actuals present (for metrics)
        if "curtailment_kWh_per_kw" in df.columns:
            df["curtailment_flag"] = (df["curtailment_kWh_per_kw"] > 0).astype(int)
            has_actual_values = True
        else:
            has_actual_values = False

        # price-related flags
        df = df.rename(columns={'dayaheadprice': 'dayaheadprice_eur_mwh'})

        for col in ["dayaheadprice_eur_mwh", "rebap_euro_per_mwh"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")


        df["quarterly_energy_kWh_per_kw"] = df["power_kwh"] / df["net_power_kw_unit"]

        df["DA_negative_flag"] = (df["dayaheadprice_eur_mwh"] < 0).astype(int)
        df["DA_negative_flag_lag_1"] = df["DA_negative_flag"].shift(1)

        df["rebap_negative_flag"] = (df["rebap_euro_per_mwh"] < 0).astype(int)
        df["rebap_negative_flag_lag_1"] = df["rebap_negative_flag"].shift(1)

        # which features are available
        available_features = [f for f in feature_names if f in df.columns]
        missing_features = [f for f in feature_names if f not in df.columns]

        if missing_features:
            print(f"⚠️ Missing classification features: {missing_features}")
            if not available_features:
                raise ValueError("No required classification features available.")

        # drop rows with missing in features
        df_clean = df.dropna(subset=available_features).copy()
        if df_clean.empty:
            raise ValueError("No valid rows after classification cleaning (NaNs in features).")

        # ensure numeric
        for f in available_features:
            df_clean[f] = pd.to_numeric(df_clean[f], errors="coerce")

        return df_clean, available_features, has_actual_values


    def predict_curtailment_classification(
        df_new_prediction: pd.DataFrame,
        model_path: str,
        metadata_path: str,
        threshold_path: str,
        plot: bool = False,
    ):
        """Run classification model on new data."""
        print_header("CLASSIFICATION – LOADING MODEL & METADATA")
        try:
            best_model = joblib.load(model_path)
            #model_metadata = joblib.load(metadata_path)
            with open(threshold_path, "r") as f:
                threshold_info = json.load(f)
        except FileNotFoundError as e:
            print(f"❌ Error loading classification files: {e}")
            return None

        feature_names = threshold_info["feature_names"]
        average_optimal_threshold = threshold_info["average_optimal_threshold"]

        print(f"Using optimal threshold: {average_optimal_threshold:.4f}")

        print_header("CLASSIFICATION – FEATURE ENGINEERING")
        df_clean, available_features, has_actual_values = feature_engineering_classification(
            df_new_prediction, feature_names
        )

        X_new = df_clean[available_features]
        print(f"Classification rows: {len(X_new)}, features used: {available_features}")

        print_header("CLASSIFICATION – PREDICTION")
        y_proba = best_model.predict_proba(X_new)[:, 1]
        y_pred = (y_proba >= average_optimal_threshold).astype(int)

        df_clean["predicted_curtailment_probability"] = y_proba
        df_clean["predicted_curtailment_flag"] = y_pred
        df_clean["prediction_timestamp_cls"] = pd.Timestamp.now()

        print(
            f"Predicted curtailment == 1 for {y_pred.sum():,} rows "
            f"({y_pred.mean()*100:.1f}% of classified rows)."
        )

        # metrics if actuals
        if has_actual_values and "curtailment_flag" in df_clean.columns:
            y_actual = df_clean["curtailment_flag"]
            accuracy = accuracy_score(y_actual, y_pred)
            precision = precision_score(y_actual, y_pred, zero_division=0)
            recall = recall_score(y_actual, y_pred, zero_division=0)
            f1 = f1_score(y_actual, y_pred, zero_division=0)
            roc_auc = roc_auc_score(y_actual, y_proba)
            avg_precision = average_precision_score(y_actual, y_proba)

            print_header("CLASSIFICATION – METRICS (ACTUALS AVAILABLE)")
            print(f"Accuracy:      {accuracy:.4f}")
            print(f"Precision:     {precision:.4f}")
            print(f"Recall:        {recall:.4f}")
            print(f"F1-Score:      {f1:.4f}")
            print(f"ROC AUC:       {roc_auc:.4f}")
            print(f"Avg Precision: {avg_precision:.4f}")
        else:
            accuracy = precision = recall = f1 = roc_auc = avg_precision = None
            print("ℹ️ No actual curtailment available for classification metrics.")

        # simple visualization if requested
        if plot:
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))

            # Linear scale
            sns.histplot(y_proba, bins=30, kde=True, ax=ax1)
            ax1.axvline(average_optimal_threshold, color="red", linestyle="--", label="Threshold")
            ax1.set_title("Linear Scale")
            ax1.set_xlabel("P(curtailment=1)")
            ax1.set_ylabel("Frequency")
            ax1.legend()

            # Log scale
            sns.histplot(y_proba, bins=30, kde=True, ax=ax2)
            ax2.axvline(average_optimal_threshold, color="red", linestyle="--", label="Threshold")
            ax2.set_yscale('log')
            ax2.set_title("Log Scale")
            ax2.set_xlabel("P(curtailment=1)")
            ax2.set_ylabel("Frequency (log scale)")
            ax2.legend()

            plt.suptitle("Predicted Probability Distribution", fontsize=14)
            plt.tight_layout()
            plt.show()

        results = {
            "predictions": df_clean,
            "model": best_model,
            "features_used": available_features,
            "optimal_threshold": average_optimal_threshold,
            "prediction_metrics": {
                "accuracy": accuracy,
                "precision": precision,
                "recall": recall,
                "f1": f1,
                "roc_auc": roc_auc,
                "avg_precision": avg_precision,
            },
            "prediction_stats": {
                "positive_predictions": int(y_pred.sum()),
                "negative_predictions": int(len(y_pred) - y_pred.sum()),
                "positive_rate": float(y_pred.mean()),
                "mean_probability": float(y_proba.mean()),
                "std_probability": float(y_proba.std()),
                "total_predictions": int(len(y_pred)),
            },
        }
        return results


    # =============================================================================
    # REGRESSION IMPLEMENTATION
    # =============================================================================

    def feature_engineering_regression(df: pd.DataFrame, reg_features):
        """Feature engineering for regression – must match training."""
        df = df.copy()

        # Define exogenous features (as at training)
        exo_features = [
            "quarterly_energy_kWh_per_kw",
            "enwex_percentage",
            "dayaheadprice_eur_mwh",
            "rebap_euro_per_mwh",
            "volume__mw_imbalance",
            "id500_eur_mwh"
        ]

        for col in exo_features:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").ffill().bfill()
            else:
                print(f"⚠️ Regression: missing feature {col} – filled with 0.")
                df[col] = 0.0

        # Lag features (not used in reg_features currently, but safe to create)
        if "curtailment_kWh_per_kw" in df.columns:
            df["curt_lag_1"] = df["curtailment_kWh_per_kw"].shift(1)
            df["curt_lag_2"] = df["curtailment_kWh_per_kw"].shift(2)

        available_features = [f for f in reg_features if f in df.columns]
        missing_features = [f for f in reg_features if f not in df.columns]

        if missing_features:
            print(f"⚠️ Missing regression features: {missing_features}")
            if not available_features:
                raise ValueError("No required regression features available.")

        df_clean = df.dropna(subset=available_features).copy()
        if df_clean.empty:
            raise ValueError("No valid rows after regression cleaning (NaNs in features).")

        X_new = df_clean[available_features].apply(pd.to_numeric, errors="coerce")

        return df_clean, X_new, available_features


    def plot_regression_predictions(df_clean):
        """
        Simple regression prediction distribution + time plot,
        including actual curtailment if available.

        Expects:
            - df_clean["predicted_curtailment_kWh_per_kw"]
            - optionally df_clean["curtailment_kWh_per_kw"]
        """
        # Predicted values
        y_pred = df_clean["predicted_curtailment_kWh_per_kw"].values

        # Actual values (if available)
        has_actual = "curtailment_kWh_per_kw" in df_clean.columns
        y_actual = df_clean["curtailment_kWh_per_kw"].values if has_actual else None

        if has_actual:
            y_actual_plot = df_clean[df_clean["curtailment_kWh_per_kw"] > 0]["curtailment_kWh_per_kw"].values

            fig, axes = plt.subplots(1, 2, figsize=(12, 5))

        # -------------------------------------------------------------------------
        # 1) Distribution panel
        # -------------------------------------------------------------------------
        if has_actual:
            axes[0].hist(
                y_actual_plot,
                bins=30,
                alpha=0.6,
                color="blue",
                edgecolor="black",
                label="Actual",
            )

            axes[0].hist(
                    y_pred,
                    bins=30,
                    alpha=0.6 if has_actual else 0.7,
                    color="green",
                    edgecolor="black",
                    label="Predicted",
            )

            axes[0].set_xlabel("Curtailment (kWh/kW)")
            axes[0].set_ylabel("Frequency")
            axes[0].set_title("Curtailment Distribution")
            axes[0].grid(True, alpha=0.3)
            if has_actual:
                axes[0].legend()

        # -------------------------------------------------------------------------
        # 2) Time / index panel
        # -------------------------------------------------------------------------

        if has_actual:
            time_col = None
            for candidate in ["delivery_start_berlin", "time_berlin", "timestamp"]:
                if candidate in df_clean.columns:
                    time_col = candidate
                    break

            if time_col:
                df_sorted = df_clean.sort_values(time_col)
                x_vals = df_sorted[time_col]
                y_pred_sorted = df_sorted["predicted_curtailment_kWh_per_kw"]
                axes[1].plot(
                    x_vals,
                    y_pred_sorted,
                    color="red",
                    linewidth=1,
                    label="Predicted",
                )
                if has_actual:
                    axes[1].plot(
                        x_vals,
                        df_sorted["curtailment_kWh_per_kw"],
                        color="blue",
                        linewidth=0.8,
                        alpha=0.4,
                        label="Actual",
                    )
                axes[1].set_xlabel("Time")
                axes[1].tick_params(axis="x", rotation=45)
            else:
                x_vals = np.arange(len(y_pred))
                axes[1].plot(
                    x_vals,
                    y_pred,
                    color="red",
                    linewidth=1.2,
                    label="Predicted",
                )
                if has_actual:
                    axes[1].plot(
                        x_vals,
                        y_actual,
                        color="blue",
                        linewidth=1.2,
                        alpha=0.8,
                        label="Actual",
                    )
                axes[1].set_xlabel("Sample Index")

            axes[1].set_ylabel("Curtailment (kWh/kW)")
            axes[1].set_title("Curtailment Over Time")
            if has_actual:
                axes[1].legend()
            axes[1].grid(True, alpha=0.3)

            plt.tight_layout()
            plt.show()


    def predict_curtailment_regression(
        df_reg_input: pd.DataFrame,
        model_path: str,
        params_path: str,
        plot: bool = False,
    ):
        """
        Run regression model on subset of rows (already filtered by classification).
        """
        #print_header("REGRESSION – LOADING MODEL & PARAMS")
        try:
            best_model = joblib.load(model_path)
            with open(params_path, "r") as f:
                best_params = json.load(f)
            _ = best_params  # not used but loaded for completeness
        except FileNotFoundError as e:
            print(f"❌ Error loading regression files: {e}")
            return None

        # regression features (must match training)
        reg_features = [
            "quarterly_energy_kWh_per_kw",
            "enwex_percentage",
            "dayaheadprice_eur_mwh",
            "rebap_euro_per_mwh",
            "volume__mw_imbalance",
            "id500_eur_mwh",
        ]

        if df_reg_input.empty:
            print("ℹ️ No rows passed to regression (no predicted curtailment = 1).")
            return {
                "predictions": df_reg_input.assign(predicted_curtailment_kWh_per_kw=np.nan),
                "model": best_model,
                "features_used": reg_features,
                "prediction_metrics": {"mse": None, "mae": None, "mape": None, "r2": None},
                "prediction_stats": {"mean": None, "std": None, "min": None, "max": None, "count": 0},
            }

        print_header("REGRESSION – FEATURE ENGINEERING ON FILTERED ROWS")
        df_clean, X_new, used_features = feature_engineering_regression(df_reg_input, reg_features)

        print(f"Regression rows: {len(X_new)}, features used: {used_features}")

        #print_header("REGRESSION – PREDICTION")
        y_pred = best_model.predict(X_new)
        df_clean["predicted_curtailment_kWh_per_kw"] = y_pred
        df_clean["prediction_timestamp_reg"] = pd.Timestamp.now()

        # metrics if actual curtailment present
        if "curtailment_kWh_per_kw" in df_clean.columns:
            y_actual = df_clean["curtailment_kWh_per_kw"]
            mse = mean_squared_error(y_actual, y_pred)
            mae = mean_absolute_error(y_actual, y_pred)
            mape = mean_absolute_percentage_error(y_actual, y_pred)
            r2 = r2_score(y_actual, y_pred)

            print_header("REGRESSION – METRICS (ACTUALS AVAILABLE)")
            print(f"MSE:  {mse:.4f}")
            print(f"MAE:  {mae:.4f}")
            print(f"MAPE: {mape:.4f}")
            print(f"R²:   {r2:.4f}")
        else:
            mse = mae = mape = r2 = None
            print("ℹ️ No actual curtailment available for regression metrics.")

        if plot:
            plot_regression_predictions(df_clean)

        results = {
            "predictions": df_clean,
            "model": best_model,
            "features_used": used_features,
            "prediction_metrics": {
                "mse": mse,
                "mae": mae,
                "mape": mape,
                "r2": r2,
            },
            "prediction_stats": {
                "mean": float(np.mean(y_pred)),
                "std": float(np.std(y_pred)),
                "min": float(np.min(y_pred)),
                "max": float(np.max(y_pred)),
                "count": int(len(y_pred)),
            },
        }
        return results


    # =============================================================================
    # FULL PIPELINE: CLASSIFICATION -> REGRESSION (FILTERED) FOR ONE CATEGORY
    # =============================================================================

    def run_curtailment_forecast(
        df_new_prediction: pd.DataFrame,
        cls_model_path: str,
        cls_meta_path: str,
        cls_thresh_path: str,
        reg_model_path: str,
        reg_params_path: str,
        plot_class: bool = False,
        plot_reg: bool = False,
    ):
        """
        Full pipeline for a *single* category:
        1) Classification on all rows
        2) Filter rows with predicted_curtailment_flag == 1
        3) Regression only on that filtered subset
        4) Merge regression predictions back into main dataframe
        """
        # 1. Classification
        cls_results = predict_curtailment_classification(
            df_new_prediction,
            model_path=cls_model_path,
            metadata_path=cls_meta_path,
            threshold_path=cls_thresh_path,
            plot=plot_class,
        )
        if cls_results is None:
            return None

        df_cls = cls_results["predictions"].copy()

        # 2. Filter rows with predicted_curtailment_flag == 1
        if "predicted_curtailment_flag" not in df_cls.columns:
            print("❌ Classification result missing 'predicted_curtailment_flag'.")
            return {"classification": cls_results, "regression": None, "combined": df_cls}

        df_for_reg = df_cls[df_cls["predicted_curtailment_flag"] == 1].copy()
        print_header("PIPELINE – ROWS FOR REGRESSION")
        print(f"Rows flagged as curtailment (1): {len(df_for_reg)}")

        # 3. Regression on filtered rows
        reg_results = predict_curtailment_regression(
            df_for_reg,
            model_path=reg_model_path,
            params_path=reg_params_path,
            plot=plot_reg,
        )

        # 4. Merge regression predictions back to classification df
        df_combined = df_cls.copy()
        df_combined["predicted_curtailment_kWh_per_kw"] = np.nan

        if reg_results is not None and not reg_results["predictions"].empty:

            df_reg_pred = reg_results["predictions"].copy()

            # 💡 MUST keep these keys for merge
            merge_keys = ["malo", "delivery_start_berlin"]

            # Only keep keys + prediction column
            df_reg_pred = df_reg_pred[
                merge_keys + ["predicted_curtailment_kWh_per_kw"]
            ]

            # SAFE two-key merge
            df_combined = df_combined.merge(
                df_reg_pred,
                on=merge_keys,
                how="left",
                suffixes=("", "_reg"),
            )

            # Overwrite correct column
            df_combined["predicted_curtailment_kWh_per_kw"] = \
                df_combined["predicted_curtailment_kWh_per_kw_reg"]

            df_combined.drop(columns=["predicted_curtailment_kWh_per_kw_reg"], inplace=True)

            df_combined["predicted_curtailment_kWh_per_kw"] = pd.to_numeric(
                df_combined["predicted_curtailment_kWh_per_kw"],
                errors='coerce'
            ).fillna(0)

        return {
            "classification": cls_results,
            "regression": reg_results,
            "combined": df_combined,
        }


    # =============================================================================
    # PATH CONFIG
    # =============================================================================

    def set_paths_for_category(category: str):
        # Define the base path and the category name
        base_path = r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\curtailment_prediction\curtailment_model"

        # Define a mapping of category names to paths
        category_mapping = {
            "PV_rules": "PV_rules",
            "PV_no_rules": "PV_NORULES",
            "WIND_rules": "WIND_rules",
            "WIND_no_rules": "WIND_NORULES"
        }

        # Check if the category exists in the mapping
        if category not in category_mapping:
            raise ValueError(f"Unknown category: {category}")

        # Get the folder name for the given category
        folder_name = category_mapping[category]

        # Return the paths based on the folder name
        return {
            "CLASS_MODEL_PATH": f"{base_path}/{folder_name}/classification_best_model_{folder_name}.joblib",
            "CLASS_META_PATH": f"{base_path}/{folder_name}/classification_xgboost_curtailment_model_{folder_name}.joblib",
            "CLASS_THRESH_PATH": f"{base_path}/{folder_name}/classification_best_params_{folder_name}.json",
            "REG_MODEL_PATH": f"{base_path}/{folder_name}/regression_best_model_{folder_name}.joblib",
            "REG_PARAMS_PATH": f"{base_path}/{folder_name}/regression_best_params_{folder_name}.json"
        }


    # =============================================================================
    # MULTI-CATEGORY PIPELINE
    # =============================================================================

    def run_curtailment_forecast_multi_category(
        df_ts: pd.DataFrame,
        plot_class: bool = False,
        plot_reg: bool = False,
        ):
        """
        df_timeseries : time series data for all malos (many rows per malo).
        df_assets     : one row per malo with columns ['malo', 'technology', 'EEG', ...].

        This will:
        - assign a category per malo
        - join category into df_timeseries
        - group by category and run the pipeline WITH THE CORRECT MODEL PATHS
        - return concatenated results.
        """

        if df_ts['category'].isna().any():
            missing_malo = df_ts.loc[df_ts['category'].isna(), 'malo'].unique()
            raise ValueError(f"Some malos in time series do not have a category: {missing_malo}")

        all_combined = []
        all_results_by_category = {}

        # 3) Loop per category, use the correct paths
        for category, df_cat in df_ts.groupby('category'):
            print_header(f"🍫🍫 RUNNING CATEGORY: {category}")
            paths = set_paths_for_category(category)

            res = run_curtailment_forecast(
                df_new_prediction=df_cat,
                cls_model_path=paths["CLASS_MODEL_PATH"],
                cls_meta_path=paths["CLASS_META_PATH"],
                cls_thresh_path=paths["CLASS_THRESH_PATH"],
                reg_model_path=paths["REG_MODEL_PATH"],
                reg_params_path=paths["REG_PARAMS_PATH"],
                plot_class=plot_class,
                plot_reg=plot_reg,
            )

            if res is not None:
                combined_cat = res["combined"].copy()
                combined_cat["category"] = category
                all_combined.append(combined_cat)
                all_results_by_category[category] = res

        if not all_combined:
            return None

        df_all = pd.concat(all_combined, ignore_index=True)

        return {
            "by_category": all_results_by_category,
            "combined": df_all,
        }


    if __name__ == "__main__":
        df_ts = df_ts

        results = run_curtailment_forecast_multi_category(
            df_ts=df_ts,
            plot_class=False,
            plot_reg=False,
        )

        if results is not None:
            df_out = results["combined"]
            print_header("FINAL COMBINED OUTPUT (ALL CATEGORIES)")
            print(f"Rows in output: {len(df_out)}")

            cols_show = [
                c for c in [
                    "malo",
                    "delivery_start_berlin",
                    "category",
                    "predicted_curtailment_probability",
                    "predicted_curtailment_flag",
                    "predicted_curtailment_kWh_per_kw",
                ]
                if c in df_out.columns
            ]

            print("🫕🫕🫕🫕🫕🫕🫕🫕🫕")
            print(df_out[cols_show].round(2))
            print("🫕🫕🫕🫕🫕🫕🫕🫕🫕")
            print(df_out.info())

            excel_file_path = f"{folder}\\df_out.xlsx"
            rows_per_sheet = 1000000
            with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
                for i in range(0, len(df_out), rows_per_sheet):
                    chunk = df_out.iloc[i:i + rows_per_sheet]
                    sheet_name = f'Sheet_{i//rows_per_sheet + 1}'
                    chunk.to_excel(writer, sheet_name=sheet_name, index=False)


            df_out.rename(columns={'dayaheadprice_eur_mwh': 'dayaheadprice'}, inplace=True)

            df_out["curtailment_forecast_kwh"] = df_out["predicted_curtailment_kWh_per_kw"] * df_out["net_power_kw_unit"]
            df_out["power_kwh"] = df_out["power_kwh"] + df_out["curtailment_forecast_kwh"]

            df_out_process = df_out[["malo","time_berlin","power_kwh","dayaheadprice","monthly_reference_market_price_eur_mwh","available_years","available_months",'year', 'month']]


    def process_production_data(merge_prod_rmv_dayahead, folder):
        """
        Processes the production data by:
        - Renaming columns
        - Dropping duplicates
        - Calculating deltas for spot prices
        - Aggregating monthly and yearly data
        - Saving results to Excel

        Args:
        merge_prod_rmv_dayahead (pd.DataFrame): DataFrame with production and price data.
        folder (str): Path to the folder where output Excel files will be saved.
        """
        input_df_name = "merge_prod_rmv_dayahead"
        folder_name = os.path.join(folder, f"{input_df_name}_forecast_output")
        os.makedirs(folder_name, exist_ok=True)

        # Rename 'power_kwh' column to 'production_kwh'
        merge_prod_rmv_dayahead.rename(columns={'power_kwh': 'production_kwh'}, inplace=True)

        # Remove duplicates based on 'malo', 'time_berlin', and 'production_kwh'
        merge_prod_rmv_dayahead_dropdup = merge_prod_rmv_dayahead.drop_duplicates(subset=["malo", "time_berlin", "production_kwh"])

        # Calculate the 'deltaspot_eur' (delta spot price)
        merge_prod_rmv_dayahead_dropdup['deltaspot_eur'] = (
            (merge_prod_rmv_dayahead_dropdup['production_kwh'] * merge_prod_rmv_dayahead_dropdup['dayaheadprice'] / 1000) -
            (merge_prod_rmv_dayahead_dropdup['production_kwh'] * merge_prod_rmv_dayahead_dropdup['monthly_reference_market_price_eur_mwh'] / 1000)
        )

        # Sort the DataFrame by 'time_berlin' and save to Excel
        merge_prod_rmv_dayahead_dropdup.sort_values('time_berlin').to_excel(f"{folder}\\merge_prod_rmv_dayahead_dropdup.xlsx", index=False)

        # Aggregate data monthly
        monthly_agg = merge_prod_rmv_dayahead_dropdup.groupby(['year', 'month', 'malo']).agg(
            deltaspot_eur_monthly=('deltaspot_eur', 'sum'),
            available_months=('available_months', 'first'),
            available_years=('available_years', 'first'),
        ).reset_index()

        print("🥟🥟")
        print(merge_prod_rmv_dayahead_dropdup)

        # Calculate weighted delta per malo
        weighted_delta_permalo = merge_prod_rmv_dayahead_dropdup.groupby(['malo']).agg(
            total_prod_kwh_malo=('production_kwh', 'sum'),
            spot_rmv_eur_malo=('deltaspot_eur', 'sum'),
        ).reset_index()

        weighted_delta_permalo["weighted_delta_permalo"] = (weighted_delta_permalo['spot_rmv_eur_malo'] / (weighted_delta_permalo['total_prod_kwh_malo'] / 1000)).round(2)

        print("🥟🥟")
        print(weighted_delta_permalo.head(60))
        print("🥟🥟")

        # Ensure 'malo' is in string format for all DataFrames
        for df in [merge_prod_rmv_dayahead_dropdup, monthly_agg, weighted_delta_permalo]:
            df['malo'] = df['malo'].astype(str).str.strip()

        # Calculate total production over the years (not limited to 1 year)
        total_prod = merge_prod_rmv_dayahead_dropdup.groupby(['malo'])['production_kwh'].sum()

        # Map total production back to the original monthly_agg rows
        monthly_agg['total_prod_kwh'] = monthly_agg['malo'].map(total_prod)
        monthly_agg['total_prod_mwh'] = monthly_agg['total_prod_kwh'] / 1000

        print("🍌🍌")
        print(monthly_agg)

        # Save monthly aggregation to Excel
        monthly_agg.to_excel(f"{folder_name}/{input_df_name}_monthly_agg.xlsx", index=False)

        # Aggregate yearly data
        year_agg = monthly_agg.groupby(['malo'], dropna=False).agg(
            available_months=('available_months', 'first'),
            available_years=('available_years', 'first'),
            total_prod_mwh=('total_prod_mwh', 'first')
        ).reset_index()

        print("🥥🥥🥥🥥🥥")
        print(year_agg)

        # Return the final results as a dictionary of DataFrames
        return {
            "monthly_agg": monthly_agg,
            "weighted_delta_permalo": weighted_delta_permalo,
            "year_agg": year_agg
        }

    # original customer data
    delta_spot_rmv = process_production_data(merge_prod_rmv_dayahead, folder)

    monthly_agg = delta_spot_rmv["monthly_agg"]
    weighted_delta_permalo = delta_spot_rmv["weighted_delta_permalo"]
    year_agg = delta_spot_rmv["year_agg"]

    # forecast curtailment customer data
    if 'df_out_process' in locals() and not df_out_process.empty:
        forecast_curt_delta_spot_rmv = process_production_data(df_out_process, folder)

        forecast_curt_monthly_agg = forecast_curt_delta_spot_rmv["monthly_agg"]
        forecast_curt_weighted_delta_permalo = forecast_curt_delta_spot_rmv["weighted_delta_permalo"]
        forecast_curt_year_agg = forecast_curt_delta_spot_rmv["year_agg"]


    ## if result available from SEE
    if 'merge_a3' in locals() and not merge_a3.empty:
        for df in [merge_a3, weighted_delta_permalo]:
            df['malo'] = df['malo'].astype(str).str.strip()

        combine_C1_1 = pd.merge(
            merge_a3,
            weighted_delta_permalo[['malo', "weighted_delta_permalo", "total_prod_kwh_malo"]],
            on= 'malo',
            how='left'
        )

        combine_C1 = pd.merge(
            combine_C1_1,
            year_agg,
            on= 'malo',
            how='left'
        )

        print(combine_C1['total_prod_mwh'].isna().sum())  # Count NaNs in total_prod_mwh
        print(year_agg['available_months'].isna().sum())  # Count NaNs in available_months

        combine_C1["denominator"] = (combine_C1['Power in MW'] * combine_C1['available_months'] * 730)
        combine_C1["denominator"] = combine_C1["denominator"].replace(0, float('nan'))
        combine_C1['capacity_factor_percent'] = ((combine_C1['total_prod_mwh']) / combine_C1["denominator"] * 100).round(2)
        print(combine_C1[["malo","Projekt",'total_prod_kwh_malo',"total_prod_mwh",'capacity_factor_percent']].head(60))
        combine_C1 = combine_C1.rename(columns={'tech': 'Technology'})

        # IF THERE IS CURTAILMENT FORECAST
        if 'df_out_process' in locals() and not df_out_process.empty:
            for df in [forecast_curt_weighted_delta_permalo]:
                df['malo'] = df['malo'].astype(str).str.strip()

            forecast_curt_weighted_delta_permalo.rename(columns={
                'weighted_delta_permalo': 'forecast_weighted_delta_permalo',
                'total_prod_kwh_malo': 'forecast_total_prod_kwh_malo',
                }, inplace=True)

            forecast_curt_year_agg.rename(columns={
                'available_months': 'forecast_available_months',
                'available_years': 'forecast_available_years',
                'total_prod_mwh': 'forecast_total_prod_mwh'
                }, inplace=True)

            combine_v1_with_forecast = pd.merge(
                combine_C1,
                forecast_curt_weighted_delta_permalo[['malo', "forecast_weighted_delta_permalo", "forecast_total_prod_kwh_malo"]],
                on= 'malo',
                how='left'
            )

            combine_v2_with_forecast = pd.merge(
                combine_v1_with_forecast,
                forecast_curt_year_agg,
                on= 'malo',
                how='left'
            )

            combine_v2_with_forecast["denominator_1"] = (combine_v2_with_forecast['Power in MW'] * combine_v2_with_forecast['forecast_available_months'] * 730)
            combine_v2_with_forecast["denominator_1"] = combine_v2_with_forecast["denominator_1"].replace(0, float('nan'))
            combine_v2_with_forecast['forecast_capacity_factor_percent'] = ((combine_v2_with_forecast['forecast_total_prod_kwh_malo']) / 1000 / combine_v2_with_forecast["denominator_1"] * 100).round(2)


        sheet1_order = [
                    'malo',
                    'Projekt',
                    'Technology',
                    'Power in MW',
                    'INB',
                    #"Vermarktung",
                    'EEG',
                    'AW in EUR/MWh',
                    'weighted_2021_eur_mwh_blindleister',
                    'weighted_2023_eur_mwh_blindleister',
                    'weighted_2024_eur_mwh_blindleister',
                    'average_weighted_eur_mwh_blindleister',
                    "2021",
                    "2023",
                    "2024",
                    "avg_enervis",
                    "weighted_delta_permalo",
                    "forecast_weighted_delta_permalo" if 'combine_v2_with_forecast' in locals() and not combine_v2_with_forecast.empty else None,
                    "Curtailment & redispatch included",
                    "available_months",
                    "available_years",
                    "capacity_factor_percent",
                    "forecast_capacity_factor_percent" if 'combine_v2_with_forecast' in locals() and not combine_v2_with_forecast.empty else None,
                    "Balancing Cost",
                    "Curtailment_value_weighted",
                    "AMW/RWM Delta",
                    "Fair Value",
                    "Trading Convenience"
                ]

        sheet2_order = [
                    'malo',
                    'Projekt',
                    'Technology',
                    'Power in MW',
                    'INB',
                    #"Vermarktung",
                    'EEG',
                    'AW in EUR/MWh',
                    'average_weighted_eur_mwh_blindleister',
                    "avg_enervis",
                    "weighted_delta_permalo",
                    "forecast_weighted_delta_permalo" if 'combine_v2_with_forecast' in locals() and not combine_v2_with_forecast.empty else None,

                    "Curtailment & redispatch included",
                    "available_months",
                    "available_years",
                    "capacity_factor_percent",
                    "Balancing Cost",
                    "Curtailment_value_weighted",
                    "AMW/RWM Delta",
                    "Fair Value",
                    "Trading Convenience",
                    "Fee EUR/MWh",
                ]

        sheet3_order = [
                    'malo',
                    'Projekt',
                    "Technology",
                    'Power in MW',
                    'AW in EUR/MWh',
                    "Fee EUR/MWh"
                ]

        if 'combine_v2_with_forecast' in locals():
            output_final_print = combine_v2_with_forecast
        else:
            output_final_print = combine_C1


        # Prepare each sheet
        sheet1_df = ensure_and_reorder(output_final_print.copy(), sheet1_order)
        sheet2_df = ensure_and_reorder(output_final_print.copy(), sheet2_order)
        sheet3_df = ensure_and_reorder(output_final_print.copy(), sheet3_order)

        #out_path = r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\GSW Gold SolarWind Management\GSW_customerpricing.xlsx"

        customer_name = os.path.basename(os.path.dirname(out_path))

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            sheet1_df.to_excel(writer, sheet_name=f"{customer_name}_1", index=False)
            sheet2_df.to_excel(writer, sheet_name=f"{customer_name}_2", index=False)
            sheet3_df.to_excel(writer, sheet_name=f"{customer_name}_3", index=False)

        print("🥥🥥🥥")
        print(sheet2_df)
        print("🥥🥥")

    # IF ONLY PRODUCTION, no result from SEE
    else:
        df_stamm["Power in MW"] = df_stamm["net_power_kw_unit"] / 1000

        merge_a1_agg = df_stamm.groupby(['malo'], dropna=False).agg({
                    'unit_mastr_id': 'first',
                    'Projekt': 'first',
                    'tech': 'first',
                    'Power in MW': 'sum',
                    "INB": lambda x: [convert_date_or_keep_string(date) for date in x],
                    "EEG": lambda x: list(x.unique()),  # Get distinct EEG values
                    #'Vermarktung': lambda x: list(x.unique()),
                    'AW in EUR/MWh': lambda x: [round(float(val), 2) for val in x if is_number(val)],
                    "Curtailment & redispatch included": 'first',
                    "Balancing Cost": 'first',
                    "Curtailment_value_weighted" : 'first',
                    "Trading Convenience": 'first',
                }).reset_index()

        combine_1 = pd.merge(
            merge_a1_agg,
            year_agg,
            on= 'malo',
            how='left'
        )

        combine_C1_1 = pd.merge(
            combine_1,
            weighted_delta_permalo[['malo', "weighted_delta_permalo"]],
            on= 'malo',
            how='left'
        )

        combine_C1_1["denominator"] = (combine_C1_1['Power in MW'] * combine_C1_1['available_months'] * 730)
        combine_C1_1["denominator"] = combine_C1_1["denominator"].replace(0, float('nan'))

        combine_C1_1['capacity_factor_percent'] = (combine_C1_1['total_prod_mwh']) / combine_C1_1["denominator"] * 100
        combine_C1_1 = combine_C1_1.drop(columns=["denominator"])

        combine_C1_1 = combine_C1_1.rename(columns={'tech': 'Technology'})

        # IF THERE IS CURTAILMENT FORECAST
        if 'df_out_process' in locals() and not df_out_process.empty:
            for df in [forecast_curt_weighted_delta_permalo]:
                df['malo'] = df['malo'].astype(str).str.strip()

            forecast_curt_weighted_delta_permalo.rename(columns={
                'weighted_delta_permalo': 'forecast_weighted_delta_permalo',
                'total_prod_kwh_malo': 'forecast_total_prod_kwh_malo',
                }, inplace=True)

            forecast_curt_year_agg.rename(columns={
                'available_months': 'forecast_available_months',
                'available_years': 'forecast_available_years',
                'total_prod_mwh': 'forecast_total_prod_mwh'
                }, inplace=True)

            combine_v1_with_forecast = pd.merge(
                combine_C1_1,
                forecast_curt_weighted_delta_permalo[['malo', "forecast_weighted_delta_permalo", "forecast_total_prod_kwh_malo"]],
                on= 'malo',
                how='left'
            )

            combine_v2_with_forecast = pd.merge(
                combine_v1_with_forecast,
                forecast_curt_year_agg,
                on= 'malo',
                how='left'
            )

            combine_v2_with_forecast["denominator_1"] = (combine_v2_with_forecast['Power in MW'] * combine_v2_with_forecast['forecast_available_months'] * 730)
            combine_v2_with_forecast["denominator_1"] = combine_v2_with_forecast["denominator_1"].replace(0, float('nan'))
            combine_v2_with_forecast['forecast_capacity_factor_percent'] = (combine_v2_with_forecast['forecast_total_prod_kwh_malo']) / combine_v2_with_forecast["denominator_1"] * 100

        sheet1_order = [
            'malo',
            'Projekt',
            'Technology',
            'Power in MW',
            'INB',
            #"Vermarktung",
            'EEG',
            'AW in EUR/MWh',
            'weighted_2021_eur_mwh_blindleister',
            'weighted_2023_eur_mwh_blindleister',
            'weighted_2024_eur_mwh_blindleister',
            'average_weighted_eur_mwh_blindleister',
            "2021",
            "2023",
            "2024",
            "avg_enervis",
            "weighted_delta_permalo",
            "forecast_weighted_delta_permalo" if 'combine_v2_with_forecast' in locals() and not combine_v2_with_forecast.empty else None,
            "Curtailment & redispatch included",
            "available_months",
            "available_years",
            "capacity_factor_percent",
            "forecast_capacity_factor_percent" if 'combine_v2_with_forecast' in locals() and not combine_v2_with_forecast.empty else None,
            "Balancing Cost",
            "Curtailment_value_weighted",
            "AMW/RWM Delta",
            "Fair Value",
            "Trading Convenience"
            ]

        sheet2_order = [
                    'malo',
                    'Projekt',
                    'Technology',
                    'Power in MW',
                    'INB',
                    #"Vermarktung",
                    'EEG',
                    'AW in EUR/MWh',
                    'average_weighted_eur_mwh_blindleister',
                    "avg_enervis",
                    "weighted_delta_permalo",
                    "forecast_weighted_delta_permalo" if 'combine_v2_with_forecast' in locals() and not combine_v2_with_forecast.empty else None,

                    "Curtailment & redispatch included",
                    "available_months",
                    "available_years",
                    "capacity_factor_percent",
                    "Balancing Cost",
                    "Curtailment_value_weighted",
                    "AMW/RWM Delta",
                    "Fair Value",
                    "Trading Convenience",
                    "Fee EUR/MWh",
                ]

        sheet3_order = [
                    'malo',
                    'Projekt',
                    "Technology",
                    'Power in MW',
                    'AW in EUR/MWh',
                    "Fee EUR/MWh"
                ]

        if 'combine_v2_with_forecast' in locals():
            output_final_print = combine_v2_with_forecast
        else:
            output_final_print = combine_C1_1

        # Prepare each sheet
        sheet1_df = ensure_and_reorder(output_final_print.copy(), sheet1_order)
        sheet2_df = ensure_and_reorder(output_final_print.copy(), sheet2_order)
        sheet3_df = ensure_and_reorder(output_final_print.copy(), sheet3_order)

        customer_name = os.path.basename(os.path.dirname(out_path))

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            sheet1_df.to_excel(writer, sheet_name=f"{customer_name}_1", index=False)
            sheet2_df.to_excel(writer, sheet_name=f"{customer_name}_2", index=False)
            sheet3_df.to_excel(writer, sheet_name=f"{customer_name}_3", index=False)

        print("🔑🔑🔑")
        print("ONLY PRODUCTION")
        print(sheet2_df)
        print("🔑🔑")

from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
highlight_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

def process_excel_with_highlight_and_formula(out_path):
    """
    Processes the Excel file by:
    - Highlighting the first row (header) in each sheet.
    - Applying a formula to the 'Fair Value' column if it exists.
    - Linking the 'Fee EUR/MWh' column if it exists.

    Args:
    file_path (str): Path to the Excel file.
    """
    # Load the workbook and sheets
    wb = load_workbook(out_path)

    # Define the highlight color (RGB: #003366)
    highlight_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    white_font = Font(color="FFFFFF")

    # Loop through each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Apply the highlight to the first row (header row)
        for cell in sheet[1]:  # sheet[1] refers to the first row (header)
            cell.fill = highlight_fill
            cell.font = white_font

        for col in sheet.columns:

            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            adjusted_width = (max_length + 1)
            sheet.column_dimensions[column].width = adjusted_width


        # # Check if the 'Fair Value' column exists
        # if 'Fair Value' in sheet.columns:
        #     # Apply the formula to 'Fair Value' for each row
        #     for row in range(2, sheet.max_row + 1):  # Start from row 2 to avoid header
        #         balancing_cost = sheet[f'Balancing Cost{row}'].value
        #         curtailment_value_weighted = sheet[f'Curtailment_value_weighted{row}'].value
        #         amw_rwm_delta = sheet[f'AMW/RWM Delta{row}'].value

        #         if balancing_cost and curtailment_value_weighted and amw_rwm_delta:
        #             fair_value = balancing_cost - curtailment_value_weighted - amw_rwm_delta
        #             sheet[f'Fair Value{row}'].value = fair_value

    # Save the modified workbook with updates
    wb.save(out_path_highlighted)
    print("Excel file processed successfully!")

process_excel_with_highlight_and_formula(out_path)
print("🖌️🖌️")