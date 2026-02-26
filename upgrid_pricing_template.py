from typing import Any, Dict, List, Optional, Tuple
from dotenv import load_dotenv
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.metrics import (
    accuracy_score,
    precision_score,
    recall_score,
    f1_score,
    roc_auc_score,
    average_precision_score,
    confusion_matrix,
    mean_squared_error,
    mean_absolute_error,
    mean_absolute_percentage_error,
    r2_score,
)
import joblib
import pandas_gbq
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
import openpyxl
from thefuzz import fuzz, process
import psutil
import requests
import pytz
import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional, Any
from io import StringIO
from pathlib import Path
from datetime import datetime
import time
import json
import ast
import re
import gc
import os
import warnings

# =============================================================================
# CONFIGURATION
# =============================================================================

load_dotenv(r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\.env")

# Get the directory where this script is located
SCRIPT_DIR = Path(__file__).parent.resolve()

# API Credentials
BLINDLEISTER_EMAIL = os.getenv("BLINDLEISTER_EMAIL", "").strip()
BLINDLEISTER_PASSWORD = os.getenv("BLINDLEISTER_PASSWORD", "").strip()
BLINDLEISTER_HARDCODED_TOKEN = os.getenv("BLINDLEISTER_HARDCODED_TOKEN", "").strip()
ANEMOS_EMAIL = os.getenv("ANEMOS_EMAIL", "").strip()
ANEMOS_PASSWORD = os.getenv("ANEMOS_PASSWORD", "").strip()

# File Paths (relative to script location)
TURBINE_REFERENCE_PATH = r"C:\Users\JerrySetiawan\OneDrive - CFP FlexPower\Work\Data pricing\Enervis template\turbine_types_id_enervis_eraseMW.xlsx"
STAMMDATEN_PATH = SCRIPT_DIR / "UKA_0226_stammdaten.xlsx"
OUTPUT_FOLDER = SCRIPT_DIR / "combined_results"

# Processing Parameters
BATCH_SIZE = 500
TARGET_YEARS = [2021, 2023, 2024, 2025]

# =============================================================================
# PRODUCTION / PRICING CONFIGURATION (Step 5+)
# =============================================================================

PROJECT_ID = "flex-power"


BERLIN_TZ = None
if pytz is not None:
    BERLIN_TZ = pytz.timezone(os.getenv("BERLIN_TZ", "Europe/Berlin"))

DAY_AHEAD_PRICE_PATH = os.getenv("DAY_AHEAD_PRICE_PATH", "").strip()
RMV_PRICE_PATH = os.getenv("RMV_PRICE_PATH", "").strip()
MODEL_BASE_PATH = os.getenv("MODEL_BASE_PATH", "").strip()

ENABLE_FORECASTING = os.getenv("ENABLE_FORECASTING", "1").strip().lower() not in {
    "0",
    "false",
    "no",
    "off",
}

CURTAILMENT_FORECAST_TABLE = os.getenv(
    "CURTAILMENT_FORECAST_TABLE",
    "flex-power.sales.price_volume_data_for_curtailment_forecast_table",
).strip()

ROWS_PER_FULL_YEAR = int(os.getenv("ROWS_PER_FULL_YEAR", "35000"))
CUTOFF_DATE_HOURLY = os.getenv("CUTOFF_DATE_HOURLY", "2025-10-01")
EXCEL_MAX_ROWS = int(os.getenv("EXCEL_MAX_ROWS", "1000000"))

# =============================================================================
# HARDCODED TURBINE MAPPINGS (for Enervis)
# =============================================================================

TURBINE_HARDCODED_MAP = {
    "SG170-7.0 MW": "SG 7.0-170",
    "N 149-4.5 MW": "N-149/4500",
    "Nordex N 149-4.5 MW": "N-149/4500",
    "V90 MK8 Gridstreamer": "V-90 2.0MW Gridstreamer",
    "V126-3.45MW": "V-126 3.45MW",
    "V-90": "V-90 2.0MW Gridstreamer",
    "Vestas V90 2,00 MW 2.0MW": "V-90 2.0MW Gridstreamer",
    "Vestas V 90 NH 95m 2,00 MW 2.0MW": "V-90 2.0MW Gridstreamer",
    "Vestas V90 2,00 MW": "V-90 2.0MW Gridstreamer",
    "V-112 2.0MW": "V-112 3.3MW",
    "V136-3.6MW": "V-136 3.6MW",
    "V112-3,45": "V-112 3.45MW",
    "V162-5.6 MW": "V-162 5.6MW",
    "V162-6.2 MW": "V-162 6.2MW",
    "Vestas V162": "N-163/6800",
    "V 150-4.2 MW": "V-150 4.2MW (PO)",
    "Vestas V112-3.3 MW MK2A": "V-112 3.3MW",
    "V 80 - 2.0MW / Mode 105.1 dB": "V-80 2.0MW GridStreamer",
    "NTK600/43": "V-44 0.6MW",
    "NTK 600 - 180": "V-44 0.6MW",
    "Nordex N149-5.7 MW": "N-149/5700",
    "Nordex N149-5.X": "N-149/5700",
    "N149-5.7 MW": "N-149/5700",
    "N175-6.8 MW": "N-175/6800",
    "N163-6.8 MW": "N-163/6800",
    "N163-5.7 MW": "N-163/5700",
    "N163-7.0 MW": "N-163/7000",
    "N131/3000 PH134": "N-131/3000",
    "N149/4.0-4.5 NH164": "N-149/4500",
    "N117/2400R91 2400.0": "N-117/2400",
    "N 90-2.5": "N-90/2500",
    "Nordex N149 4500": "N-149/4500",
    "Nordex N131 3600": "N-131/3600",
    "N 117-2.4": "N-117/2400",
    "NordexN131 3300": "N-131/3300",
    "Vestas V90 2.0MW": "V-90 2.0MW Gridstreamer",
    "V-90 MK1-6": "V-90 2.0MW Gridstreamer",
    "V150-4.2 4.2MW": "V-150 4.2MW (PO)",
    "V150-4.2": "V-150 4.2MW (PO)",
    "V117-3.3/3.45MWBWC": "V-117 3.45MW",
    "E-66 1.8MW": "E-66/18.70",
    "E-66": "E-66/18.70",
    "E 53-0.81": "E-53 0.8MW",
    "E-82": "E-82 E2 2.0MW",
    "E-58 1.0MW": "E-58/10.58",
    "E-58": "E-58/10.58",
    "V-80 2.0MW": "V-80 2.0MW GridStreamer",
    "V-80": "V-80 2.0MW GridStreamer",
    "v80 2.0MW": "V-80 2.0MW GridStreamer",
    "V150 4.2": "V-150 4.2MW (PO)",
    "E-138 EP3 E2-HAT-160-ES-C-01": "E-138 EP3 4.2MW",
    "N 131-3300": "N-131/3300",
    "N163/5.X": "N-163/5700",
    "Nordex N117/3600": "N-117/3600",
    "N117/3.6": "N-117/3600",
    "N-117 3150": "N-117/3000",
    "N133 / 4.8 TS110": "N-133/4800",
    "N149/5.7": "N-149/5700",
    "Vensys 77": "77/1500",
    "Senvion 3.4M104": "3.4M104",
    "Senvion 3.2M": "3.2M114",
    "Senvion 3.0M114": "3.2M114",
    "3.2M123": "3.2M122",
    "REpower 3.4M 104 3.37MW": "3.4M122",
    "REpower 3.4M 104": "3.4M122",
    "Senvion 3.2M": "3.4M122",
    "E-141 EP4 4,2 MW": "E-141 EP4 4.2MW",
    "E-70 E4-2/CS 82 a 2.3MW": "E-70 E4 2.3MW",
    "E115 EP3  E3 4.2MW": "E-115 EP3 4.2MW",
    "E115 EP3  E3": "E-115 EP3 4.2MW",
    "E115 EP3 E3": "E-115 EP3 4.2MW",
    "E-53/S/72/3K/02": "E-53 0.8MW",
    "E82 E 2 2.3MW": "E-82 E2 2.3MW",
    "E-40 0.5MW": "E-40/5.40",
    "E 53-0.81 0.8MW": "E-53 0.8MW",
    "NM48/600": "NM 48/600",
    "NEG MICON NM 600/48": "NM 48/600",
    "NM600/48": "NM 48/600",
    "E-70 E4 2300": "E-70 E4 2.3MW",
    "E 82 Serrations": "E-82 E2 2.3MW",
    "E40/540/E1": "E-40/5.40",
    "TW600": "TW 600-43",
    "MM-92": "MM 92 2.05MW",
    "MM92 2.05MW": "MM 92 2.05MW",
    "MM-100": "MM 100 2.0MW",
    "MM-82": "MM 82 2.05MW",
    "MD-77": "MD 77 1.5MW",
    "SWT-3.2": "SWT-3.2-113",
    "GE-5.5": "GE 5.5-158",
    "GE-3.6": "GE 3.6-137",
    "AN62 1,3h": "1300/62",
    "N-117 Gamma": "N-117/2400",
    "AN600": "600/44",
    "AN76 2,0": "2000/76",
    "GE B1500": "GE 1.5sl",
    "Enercon-40": "E-40/5.40",
    "M 1500-600": "M1500-600",
    "GE Wind Energy 1.5 SL": "GE 1.5sl",
    "N-149/4500": "Nordex N149 4500.0",
    "Nordex N131 3600.0": "N-131/3600",
    "NordexN131 3300.0": "N-131/3000",
    "Vestas V150 4.2MW": "V-150 4.2MW (PO)",
    "V112 Mk2": "V-112 3.3MW",
    "MM92": "MM92/2050",
    "N149/5.7 TCS164": "N149/5.7 TCS164",
    "N149/4.5 TCS164": "N-149/4500",
    "N149-4,5MW": "N-149/4500",
    "V117-3,6MW": "V-117 3.45MW",
    "V117- 3,45 MW": "V-117 3.45MW",
    "Vestas V90 - 2,0 MW": "V-90 2.0MW Gridstreamer",
    "3.4M 98.0": "3.4M104",
    "eno92": "eno 92",
    "E53 800kW": "E-53 0.8MW",
    "Vestas V126 - 3,3 MW": "V-126 3.3MW",
    "Vestas V162-6.2": "V-162 6.2MW",
    "N163/6.X TCS164": "N-163/6800",
    "Nordex N100": "N-100/2500",
    "N149/4.0-4.5 Delta 4000": "N-149/4500",
    "N149/4.5 TS 125": "N-149/4500",
    "N149/5.x TS125": "N-149/5700",
    "N117/120 2400.0": "N-117/2400",
    "N117/120 2400": "N-117/2400",
    "SE 3.2M-114": "3.2M114",
    "N163/5.X,": "N-163/5700",
    "N149/5.X TS125-04 5700": "N-149/5700",
    "Delta 4000 N149/5.X": "N-149/5700",
    "N149-5.7": "N-149/5700",
    "V150-4,2": "V-150 4.2MW (PO)",
    "N149/4.0-4.5": "N149/4500",
    "NM48/750": "NM 48/600",
    "V162/6.2MW": "V162/6200",
    "N131": "N-131/3900",
    "V 150 En Ventus 5.6MW": "V-150 5.6MW",
    "V172-7.2 7.2MW": "V-172 7.2MW",
    "V90/2,0 MW 2.0MW": "V-90 2.0MW Gridstreamer",
    "V80": "V-150 5.6MW",
    "E-115 E2": "E-115 3.2MW",
    "E-70 E4": "E-70 E4 2.3MW",
    "Repower MD 77": "MD 77 1.5MW",
    "V90-2.0MW": "V-90 2.0MW Gridstreamer",
    "E-101": "E-101 3.05MW",
    "E-115 EP3": "E-115 3.0MW",
    "V90": "V-90 2.0MW Gridstreamer",
    "E-70": "E-70 E4 2.3MW",
    "MD 77": "MD 77 1.5MW",
    "V90/2MW 2.0MW": "V-90 2.0MW Gridstreamer",
    "N163/5.X TS118": "N-163/5700",
    "E-92": "E-92 2.35MW",
    "V126": "V-126 3.45MW",
    "N149": "N-149/4500",
    "E 82": "E-82 E2 2.3MW",
    "E-115": "E-115 3.0MW",
    "V 150": "V-150 4.2MW (PO)",
    "V136-4.2": "V-136 4.2MW",
    "V117": "V-117 3.3MW",
    "V112-3.45": "V-112 3.45MW",
    "N163/6.X": "N-163/6800",
    "E66": "E-66/20.70",
    "E160-5.56": "E-160 EP5 E3 5.56MW",
    "E138-4.2": "E-138 EP3 4.2MW",
}

# =============================================================================
# BLINDLEISTER API CLIENT
# =============================================================================


class BlindleisterAPI:
    """Client for Blindleister API - Using working authentication from ori_old_upgrid.py"""

    BASE_URL = "https://api.blindleister.de"

    def __init__(
        self, email: str, password: str, hardcoded_token: Optional[str] = None
    ):
        self.email = email
        self.password = password
        self.token = None
        self.hardcoded_token = hardcoded_token  # Use hardcoded token for API calls

        # DEBUG: Print initialization parameters
        # print("\n" + "="*80)
        # print("BLINDLEISTER API INITIALIZATION")
        # print("="*80)
        # print(f"Base URL: {self.BASE_URL}")
        # print(f"Email: {self.email}")
        # print(f"Password: {'*' * len(self.password) if self.password else 'NOT SET'}")
        # print(f"Hardcoded Token Provided: {bool(self.hardcoded_token)}")
        # if self.hardcoded_token:
        #     print(f"Hardcoded Token (first 20 chars): {self.hardcoded_token[:20]}...")
        #     print(f"Hardcoded Token (last 20 chars): ...{self.hardcoded_token[-20:]}")
        # print("="*80 + "\n")

    def get_token(self) -> str:
        """Get access token from Blindleister API"""
        headers = {
            "accept": "text/plain",
            "Content-Type": "application/json",
        }

        json_data = {
            "email": self.email,
            "password": self.password,
        }

        # DEBUG: Print token fetch details
        print("\n" + "=" * 80)
        print("BLINDLEISTER TOKEN FETCH")
        print("=" * 80)
        print(f"URL: {self.BASE_URL}/api/v1/authentication/get-access-token")
        print(f"Headers: {headers}")
        print(f"Payload: {{email: {self.email}, password: ***}}")
        print("=" * 80)

        response = requests.post(
            f"{self.BASE_URL}/api/v1/authentication/get-access-token",
            headers=headers,
            json=json_data,
        )

        # DEBUG: Print token response
        # print(f"\nResponse Status: {response.status_code}")
        # print(f"Response Headers: {dict(response.headers)}")
        # print(f"Response Text (raw): {response.text}")
        # print(f"Response Text (stripped): {response.text.strip('\"')}")

        if response.status_code != 200:
            raise Exception(
                f"Failed to get Blindleister token: {response.status_code} - {response.text}"
            )

        self.token = response.text.strip('"')
        # print(f"\n✅ Blindleister token obtained")
        # print(f"Token (first 20 chars): {self.token[:20]}...")
        # print(f"Token (last 20 chars): ...{self.token[-20:]}")
        # print(f"Token length: {len(self.token)}")
        # print("="*80 + "\n")
        return self.token

    def get_market_prices(self, site_ids: List[str], years: List[int]) -> pd.DataFrame:
        """Fetch market prices for multiple sites and years

        Strategy:
        1. If hardcoded token is provided, try it first
        2. If hardcoded token fails (401), fetch fresh token and retry
        3. If no hardcoded token, use fresh token from the start
        """
        # Determine which token to use initially
        if self.hardcoded_token:
            auth_token = self.hardcoded_token
            token_source = "hardcoded"
            # print("🔑 Using hardcoded authorization token")
        else:
            # No hardcoded token, fetch fresh one
            if not self.token:
                self.get_token()
            auth_token = self.token
            token_source = "fresh"
            print("🔄 Using freshly fetched authorization token")

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {auth_token}",
        }

        # DEBUG: Print initial request configuration
        # print("\n" + "="*80)
        # print("BLINDLEISTER MARKET PRICE FETCH")
        # print("="*80)
        # print(f"Token source: {token_source}")
        # print(f"Auth token (first 20 chars): {auth_token[:20]}...")
        # print(f"Auth token (last 20 chars): ...{auth_token[-20:]}")
        # print(f"Auth token length: {len(auth_token)}")
        # print(f"Headers: {headers}")
        # print(f"Number of site IDs: {len(site_ids)}")
        # print(f"Site IDs: {site_ids[:5]}..." if len(site_ids) > 5 else f"Site IDs: {site_ids}")
        # print(f"Years: {years}")
        # print("="*80 + "\n")

        records = []
        token_refreshed = False

        # Loop through each ID and fetch data for each year (matching ori_old_upgrid.py)
        for idx, site_id in enumerate(site_ids):
            print(f"[{idx+1}/{len(site_ids)}] Processing: {site_id}")

            for year in years:
                payload = {"ids": [site_id], "year": year}

                # DEBUG: Print request details for first site only
                if idx == 0:
                    print(f"\n  --- Request Details (Year {year}) ---")
                    print(
                        f"  URL: {self.BASE_URL}/api/v1/market-price-atlas-api/get-market-price"
                    )
                    print(f"  Headers: {headers}")
                    print(f"  Payload: {payload}")
                    print(f"  --- End Request Details ---\n")

                response = requests.post(
                    f"{self.BASE_URL}/api/v1/market-price-atlas-api/get-market-price",
                    headers=headers,
                    json=payload,
                )

                # DEBUG: Print response for first site
                if idx == 0:
                    print(f"  Response Status: {response.status_code}")
                    print(f"  Response Headers: {dict(response.headers)}")
                    print(
                        f"  Response Text: {response.text[:500]}..."
                        if len(response.text) > 500
                        else f"  Response Text: {response.text}"
                    )

                # If token fails with 401 and we haven't refreshed yet, get fresh token and retry
                if response.status_code == 401 and not token_refreshed:
                    if token_source == "hardcoded":
                        print("\n⚠️ Hardcoded token returned 401 (Invalid token)")
                    print("🔄 Fetching fresh token via login...")
                    auth_token = self.get_token()
                    headers["Authorization"] = f"Bearer {auth_token}"
                    token_source = "fresh"
                    token_refreshed = True
                    print(f"✅ Retrying request with fresh token...")

                    # Retry the request with new token
                    response = requests.post(
                        f"{self.BASE_URL}/api/v1/market-price-atlas-api/get-market-price",
                        headers=headers,
                        json=payload,
                    )
                    print(f"  Retry Response Status: {response.status_code}")
                    print(
                        f"  Retry Response Text: {response.text[:200]}..."
                        if len(response.text) > 200
                        else f"  Retry Response Text: {response.text}"
                    )

                if response.status_code != 200:
                    print(
                        f"  Year {year}: Failed ({response.status_code}) - {response.text}"
                    )
                    continue

                try:
                    result = response.json()
                    for entry in result:
                        entry["year"] = year
                        records.append(entry)
                    if idx == 0:
                        print(f"  Year {year}: ✅ Success - {len(result)} entries")
                except Exception as e:
                    print(f"  Year {year}: Error parsing response - {e}")
                    continue

        if not records:
            print("❌ No Blindleister records fetched")
            return pd.DataFrame()

        # Flatten JSON structure
        df_flat = pd.json_normalize(
            records,
            record_path="months",
            meta=[
                "year",
                "unit_mastr_id",
                "gross_power_kw",
                "energy_source",
                "annual_generated_energy_mwh",
                "benchmark_market_price_eur_mwh",
            ],
            errors="ignore",
        )

        print(f"✅ Fetched {len(df_flat)} Blindleister market price records")
        return df_flat

    def get_generator_details(
        self, site_ids: List[str], year: int = 2025
    ) -> pd.DataFrame:
        """Fetch generator/asset details (MaStR) for SEE ids.

        Mirrors the behavior in `ori_old_upgrid.py`:
        - POST to `/api/v1/mastr-api/get-generator-details`
        - payload: {"ids": [SEE...], "year": <year>}
        - Uses hardcoded Bearer token if provided, otherwise freshly fetched token
        """
        if not site_ids:
            return pd.DataFrame()

        # Determine which token to use initially
        if self.hardcoded_token:
            auth_token = self.hardcoded_token
            token_source = "hardcoded"
            # print("🔑 Using hardcoded authorization token (generator-details)")
        else:
            if not self.token:
                self.get_token()
            auth_token = self.token
            token_source = "fresh"
            print("🔄 Using freshly fetched authorization token (generator-details)")

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {auth_token}",
        }

        records: List[dict] = []
        token_refreshed = False

        endpoint = f"{self.BASE_URL}//api/v1/mastr-api/get-generator-details"

        for idx, site_id in enumerate(site_ids):
            payload = {"ids": [site_id], "year": year}

            response = requests.post(endpoint, headers=headers, json=payload)

            # If token fails with 401 and we haven't refreshed yet, get fresh token and retry
            if response.status_code == 401 and not token_refreshed:
                if token_source == "hardcoded":
                    print("⚠️ Hardcoded token returned 401 (generator-details)")
                print("🔄 Fetching fresh token via login (generator-details)...")
                auth_token = self.get_token()
                headers["Authorization"] = f"Bearer {auth_token}"
                token_source = "fresh"
                token_refreshed = True
                response = requests.post(endpoint, headers=headers, json=payload)

            if response.status_code != 200:
                print(
                    f"[{idx+1}/{len(site_ids)}] {site_id}: Failed ({response.status_code}) - {response.text}"
                )
                continue

            try:
                result = response.json()
                for entry in result:
                    entry["year"] = year
                    records.append(entry)
            except Exception as e:
                print(
                    f"[{idx+1}/{len(site_ids)}] {site_id}: Error parsing response - {e}"
                )
                continue

        if not records:
            print("⚠️ No Blindleister generator-details records fetched")
            return pd.DataFrame()

        df = pd.DataFrame(records)
        print(f"✅ Fetched {len(df)} Blindleister generator-details records")
        return df


# =============================================================================
# ENERVIS (ANEMOS) API CLIENT
# =============================================================================


class AnemosAPI:
    """Client for Anemos (Enervis) API"""

    BASE_URL = "https://api.anemosgmbh.com"
    AUTH_URL = (
        "https://keycloak.anemosgmbh.com/auth/realms/awis/protocol/openid-connect/token"
    )

    def __init__(self, email: str, password: str):
        self.email = email
        self.password = password
        self.token = None

    def get_token(self) -> str:
        """Get access token"""
        data = {
            "client_id": "webtool_vue",
            "grant_type": "password",
            "username": self.email,
            "password": self.password,
        }
        response = requests.post(self.AUTH_URL, data=data)
        response.raise_for_status()
        self.token = response.json()["access_token"]
        return self.token

    def get_historical_product_id(self) -> int:
        """Get hist-ondemand product ID"""
        if not self.token:
            self.get_token()
        headers = {"Authorization": f"Bearer {self.token}"}
        response = requests.get(f"{self.BASE_URL}/products_mva", headers=headers)
        response.raise_for_status()
        products = response.json()
        for p in products:
            if "hist-ondemand" in p["mva_product_type"]["name"].lower():
                return p["id"]
        raise Exception("hist-ondemand product not found")

    def start_job(self, product_id: int, parkinfo: List[Dict]) -> Optional[str]:
        """Start historical job with automatic token refresh on 401"""
        if not parkinfo:
            print("⚠️ No parkinfo provided, skipping job")
            return None

        headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json",
        }
        payload = {"mva_product_id": product_id, "parameters": {"parkinfo": parkinfo}}

        # Retry logic for token expiration
        for attempt in range(2):
            response = requests.post(
                f"{self.BASE_URL}/jobs", headers=headers, json=payload
            )

            if response.status_code == 401 and attempt == 0:
                print("🔄 Token expired during job start, refreshing...")
                self.get_token()
                headers["Authorization"] = f"Bearer {self.token}"
                continue

            if response.status_code != 200:
                print(f"❌ Job start failed: {response.text}")
                response.raise_for_status()

            job_uuid = response.json()["uuid"]
            print(f"✅ Job started: {job_uuid}")
            return job_uuid

        return None

    def wait_for_job(self, job_uuid: str, poll_interval: int = 10) -> Dict:
        """Poll job status until complete"""
        url = f"{self.BASE_URL}/jobs/{job_uuid}"
        while True:
            headers = {"Authorization": f"Bearer {self.token}"}
            response = requests.get(url, headers=headers)

            if response.status_code == 401:
                print("🔄 Token expired, refreshing...")
                self.get_token()
                headers = {"Authorization": f"Bearer {self.token}"}
                response = requests.get(url, headers=headers)

            response.raise_for_status()
            job_info = response.json()
            if isinstance(job_info, list):
                job_info = job_info[0]

            status = job_info.get("status")
            print(f"⏳ Job status: {status}")

            if status in ["DONE", "COMPLETED"]:
                return job_info
            elif status in ["FAILED", "CANCELED"]:
                raise Exception(f"Job ended with status: {status}")

            time.sleep(poll_interval)

    def extract_results(self, job_info: Dict) -> List[pd.DataFrame]:
        """Extract results from job info"""
        results = job_info.get("info", {}).get("results", [])
        if not results:
            print("❌ No results found in job")
            return []

        dfs = []
        for result in results:
            turbine_id = result.get("id")
            year_data = result.get("Marktwertdifferenzen")
            if year_data:
                df = pd.DataFrame.from_dict(
                    year_data, orient="index", columns=["Marktwertdifferenz"]
                )
                df.index.name = "Year"
                df = df.reset_index()
                df["id"] = turbine_id
                dfs.append(df)

        print(f"✅ Extracted {len(dfs)} Enervis result DataFrames")
        return dfs


# =============================================================================
# TURBINE MATCHING FUNCTIONS (for Enervis)
# =============================================================================


def clean_manufacturer_name(name: str) -> str:
    """Clean manufacturer name by removing common suffixes"""
    if not name or pd.isna(name):
        return ""
    name = str(name).strip()

    remove_phrases = [
        "gmbh & co. kg",
        "central europe",
        "deutschland gmbh",
        "energy gmbh",
        "deutschland",
        "gmbh",
        "se",
        "energy",
        "ag",
    ]
    for phrase in remove_phrases:
        pattern = r"\b" + re.escape(phrase) + r"\b"
        name = re.sub(pattern, "", name, flags=re.IGNORECASE)

    name = re.sub(r"\s+", " ", name)
    name_lower = name.lower().strip()

    if "ge wind" in name_lower or "ge energy" in name_lower:
        name = "ge"
    elif "neg micon" in name_lower:
        name = "nm"
    elif "repower" in name_lower:
        name = "repower"

    return name.strip().lower()


def clean_turbine_model(name: str) -> str:
    """Clean turbine model name"""
    name = str(name)
    if not name or pd.isna(name):
        return ""

    name = name.strip()

    remove_prefixes = [
        "mit serrations",
        "delta 4000",
        "delta4000",
        "neg micon",
        "senvion",
        "enercon",
        "nercon",
        "vensys",
        "vestas",
        "nordex",
        "repower",
        "siemens",
    ]
    for prefix in remove_prefixes:
        pattern = r"\b" + re.escape(prefix) + r"\b"
        name = re.sub(pattern, "", name, flags=re.IGNORECASE)

    name = re.sub(r"(\d),(\d)", r"\1.\2", name)

    suffixes_to_remove = ["turbine", "wind"]
    for suffix in suffixes_to_remove:
        pattern = r"\b" + re.escape(suffix) + r"\b"
        name = re.sub(pattern, "", name, flags=re.IGNORECASE)

    name = name.strip("- ").strip()
    return name


def prepare_turbine_matching_dataframe(
    df_turbines: pd.DataFrame,
    df_ref: pd.DataFrame,
    nan_path: Path,
    threshold: int = 76,
) -> pd.DataFrame:
    """Prepare turbine DataFrame with fuzzy matching to reference database"""
    df = df_turbines.copy()
    nan_path.parent.mkdir(parents=True, exist_ok=True)

    df["clean_manufacturer"] = df["manufacturer"].apply(clean_manufacturer_name)
    df["turbine_model_clean"] = df["turbine_model"].apply(clean_turbine_model)
    df["add_turbine"] = df["turbine_model"]
    df["net_power_mw"] = df["net_power_kw"] / 1000

    vestas_senvion_enercon = [
        "Vestas Deutschland GmbH",
        "Senvion Deutschland GmbH",
        "ENERCON GmbH",
        "VENSYS Energy AG",
        "Enron Wind GmbH",
        "NEG Micon Deutschland GmbH",
    ]
    nordex_repower = [
        "Nordex Energy GmbH",
        "REpower Systems SE",
        "Nordex Germany GmbH",
        "eno energy GmbH",
    ]

    df.loc[df["manufacturer"].isin(vestas_senvion_enercon), "add_turbine"] = (
        df["turbine_model_clean"].astype(str).str.strip()
        + " "
        + df["net_power_mw"].round(3).astype(str)
        + "MW"
    )
    df.loc[df["manufacturer"].isin(nordex_repower), "add_turbine"] = (
        df["turbine_model_clean"].astype(str).str.strip()
        + " "
        + df["net_power_kw"].astype(str)
    )
    df.loc[df["manufacturer"].isin(["REpower Systems SE"]), "add_turbine"] = (
        df["turbine_model_clean"].astype(str).str.strip()
        + " "
        + df["hub_height_m"].astype(str)
    )

    name_to_power = {}
    if "rated_power" in df_ref.columns:
        name_to_power = df_ref.set_index("name")["rated_power"].to_dict()

    hardcoded_keys = list(TURBINE_HARDCODED_MAP.keys())

    def match_with_hardcoded(row):
        if row["turbine_model"] in TURBINE_HARDCODED_MAP:
            return TURBINE_HARDCODED_MAP[row["turbine_model"]]
        if row["add_turbine"] in TURBINE_HARDCODED_MAP:
            return TURBINE_HARDCODED_MAP[row["add_turbine"]]
        if row["turbine_model_clean"] in TURBINE_HARDCODED_MAP:
            return TURBINE_HARDCODED_MAP[row["turbine_model_clean"]]

        if isinstance(row["add_turbine"], str) and row["add_turbine"].strip():
            match_result = process.extractOne(
                row["add_turbine"], hardcoded_keys, scorer=fuzz.token_set_ratio
            )
            if match_result and match_result[1] >= threshold:
                hardcoded_match = match_result[0]
                matched_name = TURBINE_HARDCODED_MAP[hardcoded_match]
                if matched_name in name_to_power:
                    ref_power = name_to_power[matched_name]
                    user_power = row.get("net_power_kw", 0)
                    power_diff = (
                        abs(ref_power - user_power) / ref_power if ref_power else 1
                    )
                    if power_diff < 0.1:
                        return matched_name

        if isinstance(row["add_turbine"], str) and row["add_turbine"].strip():
            match_result = process.extractOne(
                row["add_turbine"],
                df_ref["name"].dropna().unique(),
                scorer=fuzz.token_set_ratio,
            )
            if match_result and match_result[1] >= threshold:
                return match_result[0]

        return None

    df["Matched_Turbine_Name"] = df.apply(match_with_hardcoded, axis=1)
    name_to_id = df_ref.set_index("name")["id"].to_dict()
    df["Matched_Turbine_ID"] = df["Matched_Turbine_Name"].map(name_to_id)

    df.to_excel(nan_path, index=False)
    print(f"💾 Turbine matching saved to {nan_path}")

    matched_count = df["Matched_Turbine_ID"].notna().sum()
    print(f"✅ Matched {matched_count}/{len(df)} turbines to reference database")

    return df


# =============================================================================
# BLINDLEISTER DATA PROCESSING
# =============================================================================


def process_blindleister_data(
    df_flat: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Process Blindleister market price data
    Returns: (yearly_weighted_df, overall_weighted_df)
    """
    if df_flat.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Calculate monthly spot-rmv difference
    df_flat["spot_rmv_EUR_monthly_ytd"] = (
        df_flat["monthly_generated_energy_mwh"]
        * df_flat["monthly_market_price_eur_mwh"]
    ) - (
        df_flat["monthly_generated_energy_mwh"]
        * df_flat["monthly_reference_market_price_eur_mwh"]
    )

    # Aggregate by year and unit
    permalo_yearly_blind = (
        df_flat.groupby(["year", "unit_mastr_id"], dropna=False)
        .agg(
            spot_rmv_EUR_yearly=("spot_rmv_EUR_monthly_ytd", "sum"),
            sum_prod_yearly=("monthly_generated_energy_mwh", "sum"),
        )
        .assign(blind_yearly=lambda x: x["spot_rmv_EUR_yearly"] / x["sum_prod_yearly"])
        .reset_index()
    )

    # Aggregate over the years
    permalo_blind = (
        df_flat.groupby("unit_mastr_id", dropna=False)
        .agg(
            spot_rmv_EUR_ytd=("spot_rmv_EUR_monthly_ytd", "sum"),
            sum_prod_ytd=("monthly_generated_energy_mwh", "sum"),
        )
        .assign(
            average_weighted_eur_mwh_blindleister=lambda x: x["spot_rmv_EUR_ytd"]
            / x["sum_prod_ytd"]
        )
        .reset_index()
    )

    # Pivot by year
    weighted_years_pivot = permalo_yearly_blind.pivot(
        index="unit_mastr_id", columns="year", values="blind_yearly"
    ).reset_index()

    weighted_years_pivot.columns.name = None
    weighted_years_pivot = weighted_years_pivot.rename(
        columns={
            2021: "weighted_2021_eur_mwh_blindleister",
            2023: "weighted_2023_eur_mwh_blindleister",
            2024: "weighted_2024_eur_mwh_blindleister",
            2025: "weighted_2025_eur_mwh_blindleister",
        }
    )

    # Merge yearly and overall
    final_weighted_blindleister = pd.merge(
        weighted_years_pivot,
        permalo_blind[["unit_mastr_id", "average_weighted_eur_mwh_blindleister"]],
        on="unit_mastr_id",
        how="left",
    )

    # Round values (convert column names to string first to handle integer year columns)
    cols_to_round = [
        col
        for col in final_weighted_blindleister.columns
        if "eur_mwh" in str(col).lower()
    ]
    final_weighted_blindleister[cols_to_round] = final_weighted_blindleister[
        cols_to_round
    ].round(2)

    print(
        f"✅ Processed Blindleister data for {len(final_weighted_blindleister)} units"
    )

    return final_weighted_blindleister


def merging_monthly_production_blindleister_real_production_stamm(
    df_stamm: pd.DataFrame, df_blindleister: pd.DataFrame, df_production: pd.DataFrame
) -> pd.DataFrame:
    """Merge Blindleister monthly production with real production data"""
    if df_blindleister.empty or df_production.empty:
        print("⚠️ One or both DataFrames are empty, cannot merge")
        return pd.DataFrame()

    # truncate "month" from "time_berlin" which has format "YYYY-MM-DD HH:MM:SS" to get "date_month" in format "YYYY-MM"
    df_production["date_month"] = (
        pd.to_datetime(df_production["time_berlin"]).dt.to_period("M").dt.to_timestamp()
    )

    # Build monthly timestamp key from numeric year/month (avoids int + str errors)
    df_blindleister["date_month"] = pd.to_datetime(
        {
            "year": pd.to_numeric(df_blindleister["year"], errors="coerce"),
            "month": pd.to_numeric(df_blindleister["month"], errors="coerce"),
            "day": 1,
        },
        errors="coerce",
    )

    if "Marktstammdatenregister-ID" in df_stamm.columns:
        df_stamm.rename(
            columns={"Marktstammdatenregister-ID": "unit_mastr_id"}, inplace=True
        )

    df_blindleister["unit_mastr_id"] = (
        df_blindleister["unit_mastr_id"].astype(str).str.strip()
    )
    df_stamm["unit_mastr_id"] = df_stamm["unit_mastr_id"].astype(str).str.strip()

    # checking if all unit_mastr_id in Blindleister are present in Stamm
    blind_units_set = set(df_blindleister["unit_mastr_id"].unique())
    stamm_units_set = set(df_stamm["unit_mastr_id"].unique())
    missing_in_blind = stamm_units_set - blind_units_set
    if missing_in_blind:
        print(
            f"⚠️ {len(missing_in_blind)} unit_mastr_id from stamm are missing in blind. {list(missing_in_blind)}"
        )
    else:
        print("✅ All unit_mastr_id from Blindleister are present in Stamm")

    blind_stamm_unit = pd.merge(
        df_blindleister[
            [
                "unit_mastr_id",
                "year",
                "month",
                "date_month",
                "monthly_generated_energy_mwh",
            ]
        ],
        df_stamm[["unit_mastr_id", "malo", "Projekt"]],
        on="unit_mastr_id",
        how="left",
    )

    # skip malo which has missing missing_in_blind
    blind_stamm_unit = blind_stamm_unit[
        ~blind_stamm_unit["malo"].isin(missing_in_blind)
    ]

    # aggregating by malo + month, and summing monthly production
    blind_stamm_malo = blind_stamm_unit.groupby(
        ["malo", "date_month"], as_index=False
    ).agg(
        Project=("Projekt", lambda x: x.dropna().unique().tolist()),
        monthly_blindleister_mwh=("monthly_generated_energy_mwh", "sum"),
    )

    agg_dict = {}
    # aggregate df_production by malo and date_month, summing each different "production" columns, skip aggregating the column is not available
    if "total_produce_kWh" in df_production.columns:
        agg_dict["total_produce_kWh"] = "sum"
    if "best_power_curve_output_kwh" in df_production.columns:
        agg_dict["best_power_curve_output_kwh"] = "sum"
    if "DA_forecast_prod_kwh_sum_hourly" in df_production.columns:
        agg_dict["DA_forecast_prod_kwh_sum_hourly"] = "sum"

    df_production_monthly = (
        df_production.groupby(["malo", "date_month"]).agg(agg_dict).reset_index()
    )

    blind_stamm_malo["malo"] = blind_stamm_malo["malo"].astype(str).str.strip()
    df_production_monthly["malo"] = (
        df_production_monthly["malo"].astype(str).str.strip()
    )

    merged_monthly_malo = pd.merge(
        blind_stamm_malo,
        df_production_monthly,
        on=["malo", "date_month"],
        how="right",
    )

    merged_monthly_malo["monthly_real_produce_MWh"] = (
        merged_monthly_malo["total_produce_kWh"] / 1000
    )
    merged_monthly_malo["monthly_best_power_curve_output_MWh"] = (
        merged_monthly_malo["best_power_curve_output_kwh"] / 1000
    )
    merged_monthly_malo["monthly_DA_forecast_prod_MWh"] = (
        merged_monthly_malo["DA_forecast_prod_kwh_sum_hourly"] / 1000
    )

    merged_monthly_malo.drop(
        columns=[
            "total_produce_kWh",
            "best_power_curve_output_kwh",
            "DA_forecast_prod_kwh_sum_hourly",
        ],
        inplace=True,
    )

    print(
        f"✅ Merged Blindleister with production data: {len(merged_monthly_malo)} records"
    )

    return merged_monthly_malo


def merging_monthly_production_blindleister_real_production(
    df_blindleister: pd.DataFrame, df_production: pd.DataFrame
) -> pd.DataFrame:
    """Merge Blindleister monthly production with real production data"""
    if df_blindleister.empty or df_production.empty:
        print("⚠️ One or both DataFrames are empty, cannot merge")
        return pd.DataFrame()

    # truncate "month" from "time_berlin" which has format "YYYY-MM-DD HH:MM:SS" to get "date_month" in format "YYYY-MM"
    df_production["date_month"] = (
        pd.to_datetime(df_production["time_berlin"]).dt.to_period("M").dt.to_timestamp()
    )

    # Build monthly timestamp key from numeric year/month (avoids int + str errors)
    df_blindleister["date_month"] = pd.to_datetime(
        {
            "year": pd.to_numeric(df_blindleister["year"], errors="coerce"),
            "month": pd.to_numeric(df_blindleister["month"], errors="coerce"),
            "day": 1,
        },
        errors="coerce",
    )

    df_blindleister["unit_mastr_id"] = (
        df_blindleister["unit_mastr_id"].astype(str).str.strip()
    )
    df_production["unit_mastr_id"] = df_production["malo"].astype(str).str.strip()

    agg_dict = {}
    # aggregate df_production by malo and date_month, summing each different "production" columns, skip aggregating the column is not available
    if "production_mwh" in df_production.columns:
        agg_dict["total_produce_kWh"] = "sum"
    if "best_power_curve_output_kwh" in df_production.columns:
        agg_dict["best_power_curve_output_kwh"] = "sum"
    if "DA_forecast_prod_kwh_sum_hourly" in df_production.columns:
        agg_dict["DA_forecast_prod_kwh_sum_hourly"] = "sum"

    df_production_monthly = (
        df_production.groupby(["malo", "date_month"]).agg(agg_dict).reset_index()
    )

    merged_monthly_malo = pd.merge(
        df_blindleister,
        df_production_monthly,
        on=["malo", "date_month"],
        how="right",
    )

    merged_monthly_malo["monthly_real_produce_MWh"] = (
        merged_monthly_malo["total_produce_kWh"] / 1000
    )
    merged_monthly_malo["monthly_best_power_curve_output_MWh"] = (
        merged_monthly_malo["best_power_curve_output_kwh"] / 1000
    )
    merged_monthly_malo["monthly_DA_forecast_prod_MWh"] = (
        merged_monthly_malo["DA_forecast_prod_kwh_sum_hourly"] / 1000
    )

    print(
        f"✅ Merged Blindleister with production data: {len(merged_monthly_malo)} records"
    )

    return merged_monthly_malo


# =============================================================================
# ENERVIS DATA PROCESSING
# =============================================================================


def process_enervis_results(
    dfs: List[pd.DataFrame], target_years: List[int] = TARGET_YEARS
) -> pd.DataFrame:
    """Process Enervis API results to create pivot table with yearly averages"""
    if not dfs:
        return pd.DataFrame()

    all_df = pd.concat(dfs, ignore_index=True)
    all_df["Year"] = all_df["Year"].astype(str)

    # Convert target_years to strings for comparison
    target_years_str = [str(y) for y in target_years]

    existing_years = all_df["Year"].unique().tolist()
    valid_years = [y for y in target_years_str if y in existing_years]

    if not valid_years:
        print("⚠️ No target years found in Enervis data")
        return pd.DataFrame()

    all_df = all_df[all_df["Year"].isin(valid_years)].copy()

    # choosing the minimum Marktwertdifferenz per id/malo and year
    df_filtered = all_df.loc[
        all_df.groupby(["id", "Year"])["Marktwertdifferenz"].idxmin()
    ].copy()
    df_filtered["Marktwertdifferenz"] = df_filtered["Marktwertdifferenz"].round(2)

    df_pivot = (
        df_filtered.pivot(index="id", columns="Year", values="Marktwertdifferenz")
        .rename_axis(None, axis=1)
        .reset_index()
    )

    for year_str in target_years_str:
        if year_str not in df_pivot.columns:
            df_pivot[year_str] = np.nan

    # after filtering the minimum value each year, calculate the average over the target years
    df_pivot["avg_enervis"] = (
        df_pivot[target_years_str].mean(axis=1, skipna=True).round(2)
    )
    columns_to_keep = ["id"] + target_years_str + ["avg_enervis"]
    df_result = df_pivot[columns_to_keep]

    # Rename columns to match Blindleister naming
    rename_dict = {"id": "malo"}
    for year in target_years:
        rename_dict[str(year)] = f"enervis_{year}"

    df_result = df_result.rename(columns=rename_dict)

    print(f"✅ Processed Enervis data for {len(df_result)} turbines")
    return df_result


def run_enervis_calculation(
    df_wind_units: pd.DataFrame,
    anemos_api: AnemosAPI,
    turbine_ref_df: pd.DataFrame,
    nan_path: Path,
    batch_info: str = "",
) -> Tuple[Optional[pd.DataFrame], Optional[pd.DataFrame]]:
    """
    Runs the Enervis calculation for a given set of wind units.
    Returns: (results_df, matching_df)
    """
    if df_wind_units.empty:
        return None, None

    df_turbines_with_match = prepare_turbine_matching_dataframe(
        df_wind_units, turbine_ref_df, threshold=76, nan_path=nan_path
    )

    df_matched_turbines = df_turbines_with_match.dropna(
        subset=["Matched_Turbine_ID"]
    ).copy()

    if df_matched_turbines.empty:
        print("⚠️ No turbines were matched for Enervis API call.")
        return None, df_turbines_with_match

    df_matched_turbines["hub_height_m"] = (
        df_matched_turbines["hub_height_m"].fillna(104).replace(0, 104)
    )
    df_matched_turbines["hub_height_m"] = df_matched_turbines["hub_height_m"].astype(
        int
    )

    # Validate and fix hub heights (API typically accepts 50-250m range)
    MIN_HUB_HEIGHT = 50
    MAX_HUB_HEIGHT = 250

    invalid_heights = df_matched_turbines[
        (df_matched_turbines["hub_height_m"] < MIN_HUB_HEIGHT)
        | (df_matched_turbines["hub_height_m"] > MAX_HUB_HEIGHT)
    ]

    if not invalid_heights.empty:
        print(f"⚠️ Found {len(invalid_heights)} turbines with invalid hub heights:")
        for idx, row in invalid_heights.iterrows():
            old_height = row["hub_height_m"]
            print(
                f"   malo={row['malo']}, hub_height={old_height}m (outside {MIN_HUB_HEIGHT}-{MAX_HUB_HEIGHT}m range)"
            )

        # Fix invalid heights: clip to valid range
        df_matched_turbines["hub_height_m"] = df_matched_turbines["hub_height_m"].clip(
            MIN_HUB_HEIGHT, MAX_HUB_HEIGHT
        )
        print(
            f"✅ Hub heights clipped to valid range ({MIN_HUB_HEIGHT}-{MAX_HUB_HEIGHT}m)"
        )

    df_matched_turbines["Matched_Turbine_ID"] = df_matched_turbines[
        "Matched_Turbine_ID"
    ].astype(int)
    df_matched_turbines["malo"] = df_matched_turbines["malo"].astype(str).str.strip()

    # Build parkinfo with detailed logging and validation
    parkinfo = []
    error_log_path = nan_path.parent / f"{nan_path.stem}_error_log.txt"
    skipped_rows = []

    def _ensure_all_submitted_ids(df_results: Optional[pd.DataFrame]) -> pd.DataFrame:
        """Ensure returned results include all submitted ids."""
        submitted_ids = [str(p["id"]) for p in parkinfo]
        df_submitted = pd.DataFrame({"malo": submitted_ids})

        # Use enervis_ prefix for year columns
        enervis_year_cols = [f"enervis_{year}" for year in TARGET_YEARS]

        if df_results is None or df_results.empty:
            df_results_norm = pd.DataFrame(
                columns=["malo"] + enervis_year_cols + ["avg_enervis"]
            )
        else:
            df_results_norm = df_results.copy()
            df_results_norm["malo"] = df_results_norm["malo"].astype(str)
            for year_col in enervis_year_cols:
                if year_col not in df_results_norm.columns:
                    df_results_norm[year_col] = np.nan
            if "avg_enervis" not in df_results_norm.columns:
                df_results_norm["avg_enervis"] = (
                    df_results_norm[enervis_year_cols]
                    .mean(axis=1, skipna=True)
                    .round(2)
                )

        merged = df_submitted.merge(df_results_norm, on="malo", how="left")
        for year_col in enervis_year_cols:
            if year_col not in merged.columns:
                merged[year_col] = np.nan
        if "avg_enervis" not in merged.columns:
            merged["avg_enervis"] = (
                merged[enervis_year_cols].mean(axis=1, skipna=True).round(2)
            )
        return merged

    for idx, row in df_matched_turbines.iterrows():
        try:
            hub_height = int(row["hub_height_m"])

            if hub_height < MIN_HUB_HEIGHT or hub_height > MAX_HUB_HEIGHT:
                skipped_rows.append(
                    {
                        "malo": row["malo"],
                        "reason": f"Invalid hub height: {hub_height}m",
                        "lat": row.get("latitude"),
                        "lon": row.get("longitude"),
                    }
                )
                continue

            park_entry = {
                "id": int(row["malo"]),
                "lat": str(row["latitude"]),
                "lon": str(row["longitude"]),
                "turbine_type_id": int(row["Matched_Turbine_ID"]),
                "hub_height": hub_height,
            }
            parkinfo.append(park_entry)

        except Exception as e:
            error_msg = f"Error building parkinfo for malo {row.get('malo', 'UNKNOWN')}: {str(e)}\n"
            error_msg += f"  Row data: lat={row.get('latitude')}, lon={row.get('longitude')}, turbine_id={row.get('Matched_Turbine_ID')}, hub_height={row.get('hub_height_m')}\n"
            print(f"⚠️ {error_msg}")

            with open(error_log_path, "a") as f:
                f.write(error_msg)
            skipped_rows.append(
                {
                    "malo": row.get("malo", "UNKNOWN"),
                    "reason": str(e),
                    "lat": row.get("latitude"),
                    "lon": row.get("longitude"),
                }
            )

    if skipped_rows:
        print(f"⚠️ Skipped {len(skipped_rows)} rows due to validation errors:")
        for skip in skipped_rows[:10]:  # Show first 10
            print(f"   {skip}")
        with open(error_log_path, "a") as f:
            f.write(f"\nSkipped rows:\n")
            for skip in skipped_rows:
                f.write(f"  {skip}\n")

    print(f"\n📍 Submitting {len(parkinfo)} locations to Enervis API")
    if len(parkinfo) == 0:
        print("⚠️ No valid locations to submit after validation")
        return None, df_turbines_with_match

    print(
        f"   Hub height range: {min(p['hub_height'] for p in parkinfo)}-{max(p['hub_height'] for p in parkinfo)}m"
    )
    print(f"   Sample locations (first 3):")
    for i, entry in enumerate(parkinfo[:3]):
        print(
            f"     {i+1}. malo={entry['id']}, lat={entry['lat']}, lon={entry['lon']}, turbine_type={entry['turbine_type_id']}, height={entry['hub_height']}m"
        )

    product_id = anemos_api.get_historical_product_id()

    # Try submitting all locations first
    try:
        job_uuid = anemos_api.start_job(product_id, parkinfo)
        if job_uuid:
            job_info = anemos_api.wait_for_job(job_uuid)
            dfs = anemos_api.extract_results(job_info)
            df_results = process_enervis_results(dfs)
            return _ensure_all_submitted_ids(df_results), df_turbines_with_match

    except Exception as e:
        # Batch submission failed - retry individually to find the problematic row(s)
        print(f"⚠️ Batch submission failed: {str(e)}")
        print(
            f" 🍌🍌🍌 Retrying with individual location submissions to identify problematic row(s)..."
        )

        # Try submitting locations one by one
        successful_results = []
        failed_locations = []

        for i, entry in enumerate(parkinfo):
            try:
                print(
                    f"  Processing {i+1}/{len(parkinfo)}: malo={entry['id']}...",
                    end=" ",
                )
                job_uuid = anemos_api.start_job(product_id, [entry])
                if job_uuid:
                    job_info = anemos_api.wait_for_job(job_uuid)
                    dfs = anemos_api.extract_results(job_info)
                    if dfs:
                        successful_results.extend(dfs)
                        print("✅")
                    else:
                        print("⚠️ No results")
            except Exception as location_error:
                print(
                    f"🍌🍌🍌  Error processing malo={entry['id']}: {str(location_error)[:50]}"
                )
                error_detail = {
                    "malo": entry["id"],
                    "lat": entry["lat"],
                    "lon": entry["lon"],
                    "turbine_type_id": entry["turbine_type_id"],
                    "hub_height": entry["hub_height"],
                    "error": str(location_error),
                }
                failed_locations.append(error_detail)
                print(f"❌ {str(location_error)[:50]}")

        # Summary
        print(f"\n📊 Individual submission results:")
        print(f"   ✅ Successful: {len(successful_results)}/{len(parkinfo)}")
        print(f"   ❌ Failed: {len(failed_locations)}/{len(parkinfo)}")

        if failed_locations:
            # Log ONLY the failed rows with full details
            print(
                f"\n🍌🍌🍌  Found {len(failed_locations)} problematic row(s) - logging details to: {error_log_path}"
            )

            with open(error_log_path, "a") as f:
                f.write("=" * 80 + "\n")
                f.write(f"❌ API ERROR{' for ' + batch_info if batch_info else ''}:\n")
                f.write(f"   Initial error: {str(e)}\n")
                f.write(f"   Batch size: {len(parkinfo)} locations\n")
                f.write(f"   Failed rows: {len(failed_locations)}\n")
                f.write("=" * 80 + "\n\n")

                # Get original dataframe data for ONLY the failed rows
                failed_malo_ids = [str(fail["malo"]) for fail in failed_locations]
                df_failed_rows = df_matched_turbines[
                    df_matched_turbines["malo"].isin(failed_malo_ids)
                ].copy()

                f.write(
                    f"PROBLEMATIC ROW(S) - {len(failed_locations)} row(s) that caused the error:\n"
                )
                f.write("=" * 80 + "\n\n")

                for i, fail in enumerate(failed_locations, 1):
                    f.write(f"--- Failed Row {i}/{len(failed_locations)} ---\n")
                    f.write(f"Error: {fail['error']}\n\n")
                    f.write(f"Submitted to API:\n")
                    f.write(f"  malo: {fail['malo']}\n")
                    f.write(f"  latitude: {fail['lat']}\n")
                    f.write(f"  longitude: {fail['lon']}\n")
                    f.write(f"  turbine_type_id: {fail['turbine_type_id']}\n")
                    f.write(f"  hub_height: {fail['hub_height']}m\n\n")

                    # Find and write full row details
                    malo_row = df_failed_rows[
                        df_failed_rows["malo"] == str(fail["malo"])
                    ]
                    if not malo_row.empty:
                        f.write(f"Full row details from original data:\n")
                        for col in malo_row.columns:
                            f.write(f"  {col}: {malo_row.iloc[0][col]}\n")
                    f.write("\n" + "🍌" * 10 + "\n\n")

                f.write("=" * 80 + "\n\n")

            print(
                f"📄 See {error_log_path.name} for details on the {len(failed_locations)} problematic row(s)"
            )

        df_partial = (
            process_enervis_results(successful_results) if successful_results else None
        )
        df_results_all = _ensure_all_submitted_ids(df_partial)

        if not successful_results:
            print("❌ No successful results after individual submissions")
        return df_results_all, df_turbines_with_match


# =============================================================================
# MAIN EXECUTION
# =============================================================================


def load_stammdaten(path: Path) -> pd.DataFrame:
    """Load and validate master data (stammdaten) from Excel"""
    df = pd.read_excel(path, sheet_name="stammdaten", engine="openpyxl")
    df.columns = df.columns.str.strip()

    # Convert malo to string
    df["malo"] = (
        df["malo"]
        .apply(
            lambda x: (
                str(int(x)) if isinstance(x, (float, int)) and pd.notna(x) else str(x)
            )
        )
        .str.strip()
    )

    # Convert MaStR ID to string
    if "Marktstammdatenregister-ID" in df.columns:
        df["Marktstammdatenregister-ID"] = (
            df["Marktstammdatenregister-ID"].astype(str).str.strip()
        )
        df.rename(columns={"Marktstammdatenregister-ID": "unit_mastr_id"}, inplace=True)

    # Normalize common capacity column names for Enervis matching
    # `prepare_turbine_matching_dataframe()` expects `net_power_kw`.
    if "net_power_kw" not in df.columns:
        if "net_power_kw_unit" in df.columns:
            df.rename(columns={"net_power_kw_unit": "net_power_kw"}, inplace=True)
        elif "net_power_mw" in df.columns:
            df["net_power_kw"] = (
                pd.to_numeric(df["net_power_mw"], errors="coerce") * 1000
            )

    # Remove rows without malo
    df.dropna(subset=("malo",), axis=0, inplace=True)

    print(f"✅ Loaded {len(df)} units from stammdaten")
    return df


# =============================================================================
# PRODUCTION / PRICING UTILITIES
# =============================================================================


def convert_date_or_keep_string(date: Any) -> str:
    """Convert a date-like value to string; keep original if conversion fails."""
    try:
        dt = pd.to_datetime(date, errors="raise", dayfirst=True)
        if pd.isna(dt):
            return ""
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return (
            ""
            if date is None or (isinstance(date, float) and np.isnan(date))
            else str(date)
        )


def is_number(val: Any) -> bool:
    """Check if a value can be converted to float."""
    try:
        float(val)
        return True
    except Exception:
        return False


def ensure_and_reorder(df: pd.DataFrame, order: List[str]) -> pd.DataFrame:
    """Ensure columns exist in DataFrame and reorder them."""
    missing_cols = [col for col in order if col and col not in df.columns]
    for col in missing_cols:
        df[col] = np.nan
    valid_order = [col for col in order if col]
    return df[valid_order]


def normalize_nan_strings(df: pd.DataFrame) -> pd.DataFrame:
    """Convert literal 'nan' strings (any casing/whitespace) to actual nulls."""
    if df is None or df.empty:
        return df
    return df.replace(to_replace=r"(?i)^\s*nan\s*$", value=np.nan, regex=True)


def sanitize_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Sanitize DataFrame for Excel export.

    Excel files can become "repairable" if cells contain:
    - illegal XML control characters (e.g. \x00)
    - extremely long strings (> 32,767 characters)
    - non-scalar Python objects (lists/dicts/arrays/Period)
    - inf values
    """
    if df is None or df.empty:
        return df

    df = df.copy()

    # Excel's max string length per cell
    EXCEL_MAX_CELL_CHARS = 32767
    # Illegal XML characters for XLSX (control chars except tab/newline/CR)
    illegal_xml_re = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

    def _to_safe_string(val: Any) -> str:
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return ""
        if isinstance(val, (dict, list, np.ndarray)):
            try:
                val = json.dumps(val, ensure_ascii=False, default=str)
            except Exception:
                val = str(val)
        elif isinstance(val, pd.Period):
            val = str(val)
        else:
            val = str(val)

        # Remove illegal control characters
        val = illegal_xml_re.sub("", val)
        # Truncate to Excel limit
        if len(val) > EXCEL_MAX_CELL_CHARS:
            val = val[: EXCEL_MAX_CELL_CHARS - 15] + "...(truncated)"
        return val

    # Sanitize column headers too (they become Excel cell values)
    df.columns = [illegal_xml_re.sub("", str(c))[:255] for c in df.columns]

    for col in df.columns:
        # Replace inf values with empty string
        df[col] = df[col].replace([np.inf, -np.inf], np.nan)

        # Convert non-scalar / problematic objects to safe strings
        if df[col].dtype == object:
            df[col] = df[col].apply(
                lambda x: (
                    _to_safe_string(x)
                    if isinstance(x, (list, dict, np.ndarray, pd.Period, str))
                    else ("" if pd.isna(x) else x)
                )
            )
        elif isinstance(df[col].dtype, pd.PeriodDtype):
            df[col] = df[col].astype(str).apply(_to_safe_string)
        elif pd.api.types.is_string_dtype(df[col].dtype):
            df[col] = df[col].astype(str).apply(_to_safe_string)

        # Convert numpy types to native Python types
        if df[col].dtype.name.startswith("int"):
            df[col] = df[col].astype("Int64")  # Nullable integer
        elif df[col].dtype.name.startswith("float"):
            df[col] = df[col].astype("float64")

    return df


def sanitize_sheet_name(name: str) -> str:
    """Create an Excel-safe sheet name (<=31 chars, no special chars)."""
    if name is None:
        name = "Sheet"
    name = str(name)
    # Invalid in Excel sheet names: : \ / ? * [ ]
    name = re.sub(r"[:\\/\?\*\[\]]", "_", name)
    name = name.strip() or "Sheet"
    return name[:31]


# b
def build_capacity_factor_comparison(
    df_base: pd.DataFrame,
    df_forecast: pd.DataFrame,
) -> pd.DataFrame:
    """Build per-malo capacity factor comparison table.

    Expects both inputs to have:
    - malo
    - Projekt
    - available_months, total_prod_mwh, capacity_factor_percent
    - forecast_available_months, forecast_total_prod_mwh, forecast_capacity_factor_percent
    """
    if df_base is None or df_base.empty:
        return pd.DataFrame()

    cols_base = [
        "malo",
        "Projekt" "available_months",
        "available_years",
        "total_prod_mwh",
        "capacity_factor_percent",
    ]
    cols_forecast = [
        "malo",
        "forecast_available_months",
        "forecast_available_years",
        "forecast_total_prod_mwh",
        "forecast_capacity_factor_percent",
    ]

    base = df_base.copy()
    for c in cols_base:
        if c not in base.columns:
            base[c] = np.nan
    base = base[cols_base]

    fc = (
        df_forecast.copy()
        if df_forecast is not None
        else pd.DataFrame(columns=cols_forecast)
    )
    for c in cols_forecast:
        if c not in fc.columns:
            fc[c] = np.nan
    fc = fc[cols_forecast]

    out = base.merge(fc, on="malo", how="left")

    for c in [
        "available_months",
        "forecast_available_months",
        "total_prod_mwh",
        "forecast_total_prod_mwh",
        "capacity_factor_percent",
        "forecast_capacity_factor_percent",
    ]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    out["delta_cf_pct_points"] = (
        out["forecast_capacity_factor_percent"] - out["capacity_factor_percent"]
    )
    out["delta_months"] = out["forecast_available_months"] - out["available_months"]
    out["delta_total_prod_mwh"] = out["forecast_total_prod_mwh"] - out["total_prod_mwh"]

    def _reason(row: pd.Series) -> str:
        reasons: List[str] = []
        if pd.isna(row.get("forecast_capacity_factor_percent")):
            reasons.append("no_forecast_metrics")
        if pd.notna(row.get("delta_months")) and row.get("delta_months") < 0:
            reasons.append("fewer_months_in_forecast")
        if (
            pd.notna(row.get("delta_total_prod_mwh"))
            and row.get("delta_total_prod_mwh") < 0
        ):
            reasons.append("lower_forecast_total_prod")
        if not reasons:
            reasons.append("ok")
        return ";".join(reasons)

    out["likely_reason"] = out.apply(_reason, axis=1)
    return out


def print_capacity_factor_debug(
    df_compare: pd.DataFrame,
    top_n: int = 25,
    only_drops: bool = True,
):
    """Print a compact CF comparison summary (top drops by forecast-original)."""
    if df_compare is None or df_compare.empty:
        print("ℹ️ No capacity-factor comparison data to display")
        return

    df_view = df_compare.copy()
    df_view = df_view[df_view["malo"].notna()].copy()

    if only_drops:
        df_view = df_view[
            df_view["delta_cf_pct_points"].notna()
            & (df_view["delta_cf_pct_points"] < 0)
        ].copy()

    df_view = df_view.sort_values("delta_cf_pct_points", ascending=True)
    df_view = df_view.head(int(top_n))

    print("\n===== CAPACITY FACTOR DEBUG (TOP DROPS) =====")
    if df_view.empty:
        print("✅ No forecast CF drops found (or no forecast CF available)")
        return

    cols = [
        "malo",
        "capacity_factor_percent",
        "forecast_capacity_factor_percent",
        "delta_cf_pct_points",
        "available_months",
        "forecast_available_months",
        "delta_months",
        "total_prod_mwh",
        "forecast_total_prod_mwh",
        "delta_total_prod_mwh",
        "likely_reason",
    ]
    cols = [c for c in cols if c in df_view.columns]
    df_print = df_view[cols].copy()

    # Make it more readable in logs
    for c in [
        "capacity_factor_percent",
        "forecast_capacity_factor_percent",
        "delta_cf_pct_points",
    ]:
        if c in df_print.columns:
            df_print[c] = df_print[c].round(3)
    for c in ["total_prod_mwh", "forecast_total_prod_mwh", "delta_total_prod_mwh"]:
        if c in df_print.columns:
            df_print[c] = df_print[c].round(2)

    print(df_print.to_string(index=False))


def print_header(title: str):
    """Print formatted section header."""
    print("\n" + "=" * 50)
    print(title)
    print("=" * 50)


def load_bigquery_fixings(project_id: str) -> pd.DataFrame:
    """Load fixing prices from BigQuery."""
    if pandas_gbq is None:
        raise ImportError("pandas_gbq is not installed; cannot load BigQuery fixings")

    query = """
    SELECT *
    FROM `flex-power.sales.origination_fixings`
    """

    df = pandas_gbq.read_gbq(query, project_id=project_id)
    if "Tenor" in df.columns:
        df["Tenor"] = df["Tenor"].astype(str)
    print(f"✅ Loaded {len(df)} fixing records from BigQuery")
    return df


def load_day_ahead_prices(path: str) -> pd.DataFrame:
    """Load day-ahead prices and convert to Berlin timezone."""
    if not path:
        raise ValueError("DAY_AHEAD_PRICE_PATH is not set")
    if BERLIN_TZ is None:
        raise RuntimeError("pytz is required for timezone handling (BERLIN_TZ)")

    df = pd.read_csv(path)
    df["delivery_start__utc_"] = pd.to_datetime(df["delivery_start__utc_"], utc=True)
    df["time_berlin"] = df["delivery_start__utc_"].dt.tz_convert(BERLIN_TZ)
    df["naive_time"] = df["time_berlin"].dt.tz_localize(None)
    df_avg = df.groupby("naive_time", as_index=False)["dayaheadprice"].mean()
    df_avg = df_avg.rename(columns={"naive_time": "time_berlin"})
    df_avg = df_avg.drop_duplicates(subset=["time_berlin", "dayaheadprice"])
    print(f"✅ Loaded {len(df_avg)} day-ahead price records")
    return df_avg


def load_rmv_prices(path: str) -> pd.DataFrame:
    """Load RMV prices."""
    if not path:
        raise ValueError("RMV_PRICE_PATH is not set")
    df = pd.read_csv(path)
    if "tech" in df.columns:
        df["tech"] = df["tech"].str.strip().str.upper().astype("category")
    print(f"✅ Loaded {len(df)} RMV price records")
    return df


def set_category_based_on_conditions(df_assets: pd.DataFrame) -> pd.DataFrame:
    """Assign EEG rules and categories based on technology/capacity/commissioning date."""
    df = df_assets.copy()

    if "INB" not in df.columns:
        df["INB"] = np.nan
    if "net_power_kw_unit" not in df.columns:
        if "net_power_kw" in df.columns:
            df["net_power_kw_unit"] = pd.to_numeric(df["net_power_kw"], errors="coerce")
        else:
            df["net_power_kw_unit"] = np.nan
    if "tech" not in df.columns:
        df["tech"] = np.nan

    df["INB_date"] = pd.to_datetime(df["INB"], dayfirst=True, errors="coerce")
    df["INB_year"] = df["INB_date"].dt.year

    df["tech"] = df["tech"].astype(str).str.strip().str.upper()

    conditions = [
        df["INB"].isna() | (df["INB"].astype(str).str.strip() == ""),
        (df["tech"] == "WIND")
        & (df["net_power_kw_unit"] >= 3000)
        & (df["INB_year"] >= 2016)
        & (df["INB_year"] < 2021),
        (df["tech"] == "PV")
        & (df["net_power_kw_unit"] >= 500)
        & (df["INB_year"] >= 2016)
        & (df["INB_year"] < 2021),
        (df["net_power_kw_unit"] >= 500)
        & (df["INB_year"] >= 2021)
        & (df["INB_year"] < 2023),
        (df["net_power_kw_unit"] >= 100) & (df["INB_year"] >= 2023),
    ]
    choices = ["rules", "6h rules", "6h rules", "4h rules", "4_3_2_1 rules"]
    df["EEG"] = np.select(conditions, choices, default="no rules")
    df = df.drop(columns=["INB_date", "INB_year"], errors="ignore")

    df["category"] = df.apply(
        lambda row: (
            "PV_no_rules"
            if row["tech"] == "PV" and row["EEG"] == "no rules"
            else (
                "PV_rules"
                if row["tech"] == "PV"
                else (
                    "WIND_no_rules"
                    if row["tech"] == "WIND" and row["EEG"] == "no rules"
                    else "WIND_rules"
                )
            )
        ),
        axis=1,
    )
    print("✅ EEG categories assigned")
    return df


def get_fix_value(
    df_fixings: pd.DataFrame, tech: str, variable: str, year: str
) -> float:
    """Extract a single fixing value for technology/variable/year."""
    sel = df_fixings[
        (df_fixings["Technology"] == tech)
        & (df_fixings["Variable"] == variable)
        & (df_fixings["Tenor"].astype(str) == str(year))
    ]

    if "new_Fixing" in sel.columns:
        s_new = sel["new_Fixing"].replace(0, np.nan).dropna()
        if not s_new.empty:
            return float(s_new.iloc[0])

    s_eur = (
        sel["EUR_MWh"].dropna() if "EUR_MWh" in sel.columns else pd.Series(dtype=float)
    )
    return float(s_eur.iloc[0]) if not s_eur.empty else np.nan


def extract_all_fixings(
    df_fixings: pd.DataFrame, year: str = "2026"
) -> Dict[str, float]:
    """Extract all required fixing values for a given year."""
    fixings = {
        "bc_pv": get_fix_value(df_fixings, "PV", "Balancing Cost", year),
        "bc_wind": get_fix_value(df_fixings, "Wind", "Balancing Cost", year),
        "tc_pv": get_fix_value(df_fixings, "PV", "Trading Convenience", year),
        "tc_wind": get_fix_value(df_fixings, "Wind", "Trading Convenience", year),
        "cv_pv_no": get_fix_value(
            df_fixings, "PV", "Curtailment Value without Rule", year
        ),
        "cv_pv_yes": get_fix_value(
            df_fixings, "PV", "Curtailment Value with any Rule", year
        ),
        "cv_w_no": get_fix_value(
            df_fixings, "Wind", "Curtailment Value without Rule", year
        ),
        "cv_w_yes": get_fix_value(
            df_fixings, "Wind", "Curtailment Value with any Rule", year
        ),
    }
    return fixings


def apply_fixings_to_stammdaten(
    df_stamm: pd.DataFrame, fixings: Dict[str, float]
) -> pd.DataFrame:
    """Apply fixing values to stammdaten DataFrame."""
    df = df_stamm.copy()

    if "tech" not in df.columns:
        df["tech"] = np.nan
    if "EEG" not in df.columns:
        df["EEG"] = "no rules"
    if "net_power_kw_unit" not in df.columns:
        if "net_power_kw" in df.columns:
            df["net_power_kw_unit"] = pd.to_numeric(df["net_power_kw"], errors="coerce")
        else:
            df["net_power_kw_unit"] = np.nan

    m_pv = df["tech"].astype(str).str.upper().eq("PV")
    m_wind = df["tech"].astype(str).str.upper().eq("WIND")
    m_no = df["EEG"].astype(str).str.contains("no rules", case=False, na=False)

    df.loc[m_pv, "Balancing Cost"] = fixings.get("bc_pv")
    df.loc[m_wind, "Balancing Cost"] = fixings.get("bc_wind")
    df.loc[m_pv, "Trading Convenience"] = fixings.get("tc_pv")
    df.loc[m_wind, "Trading Convenience"] = fixings.get("tc_wind")

    df.loc[m_pv & m_no, "Curtailment Value"] = fixings.get("cv_pv_no")
    df.loc[m_pv & ~m_no, "Curtailment Value"] = fixings.get("cv_pv_yes")
    df.loc[m_wind & m_no, "Curtailment Value"] = fixings.get("cv_w_no")
    df.loc[m_wind & ~m_no, "Curtailment Value"] = fixings.get("cv_w_yes")

    if "malo" in df.columns:
        df_curt = (
            df.groupby(["malo"], dropna=False)
            .agg(
                Curtailment_value_weighted=(
                    "Curtailment Value",
                    lambda x: (
                        np.average(
                            x,
                            weights=pd.to_numeric(
                                df.loc[x.index, "net_power_kw_unit"], errors="coerce"
                            ),
                        )
                        if len(x)
                        and pd.to_numeric(
                            df.loc[x.index, "net_power_kw_unit"], errors="coerce"
                        )
                        .notna()
                        .any()
                        else np.nan
                    ),
                )
            )
            .reset_index()
        )
        df = pd.merge(df, df_curt, on="malo", how="left")

    print("✅ Fixings applied to stammdaten")
    return df


def process_time_column(df: pd.DataFrame) -> pd.DataFrame:
    """Process time column - handle both time_berlin and time_utc."""
    df_result = df.copy()
    if "time_berlin" in df_result.columns:
        df_result["time_berlin"] = pd.to_datetime(
            df_result["time_berlin"], errors="coerce"
        )
    elif "time_utc" in df_result.columns:
        df_result["time_utc"] = pd.to_datetime(
            df_result["time_utc"], errors="coerce", utc=True
        )
        if BERLIN_TZ is None:
            raise RuntimeError("pytz is required for timezone conversion")
        df_result["time_berlin"] = (
            df_result["time_utc"].dt.tz_convert(BERLIN_TZ).dt.tz_localize(None)
        )
        df_result.drop(columns=["time_utc"], inplace=True)
    else:
        raise ValueError("No time_berlin or time_utc column found")
    return df_result


def expand_hourly_to_quarter_hourly(
    df: pd.DataFrame, cutoff_date: str = CUTOFF_DATE_HOURLY
) -> pd.DataFrame:
    """Expand hourly data to quarter-hourly before cutoff date; keep later data unchanged."""
    df_indexed = df.set_index("time_berlin")
    cutoff = pd.to_datetime(cutoff_date)
    hourly_data = df_indexed[df_indexed.index < cutoff]
    quarter_hourly_data = df_indexed[df_indexed.index >= cutoff]

    if not hourly_data.empty:
        hourly_data = hourly_data.resample("15T").ffill()
        result = pd.concat([hourly_data, quarter_hourly_data])
    else:
        result = quarter_hourly_data

    return result.sort_index().reset_index()


def filter_production_data_by_completeness(
    df: pd.DataFrame,
    rows_per_full_year: int = ROWS_PER_FULL_YEAR,
    min_rows_per_month: int = 2592,
    debug: Optional[bool] = None,
) -> pd.DataFrame:
    """Filter production data to keep only complete/continuous periods.

    - Only consider years 2021, 2023, 2024, 2025
    - Keep only "valid" months with at least `min_rows_per_month` rows
    - If a malo has (enough) full-year data, keep only full years
    - Otherwise, prefer the most recent continuous 12/24/36-month window
    - Output includes `available_years` and `available_months`
    """

    if debug is None:
        debug = os.getenv("PROD_FILTER_DEBUG", "0").strip().lower() in {
            "1",
            "true",
            "yes",
            "on",
        }

    if df is None or df.empty:
        return pd.DataFrame()

    required_cols = {"malo", "time_berlin"}
    missing = required_cols - set(df.columns)
    if missing:
        raise ValueError(
            f"Missing required columns for production filtering: {sorted(missing)}"
        )

    df_filtered = df.copy()
    df_filtered["malo"] = df_filtered["malo"].astype(str).str.strip()
    df_filtered["time_berlin"] = pd.to_datetime(
        df_filtered["time_berlin"], errors="coerce"
    )
    df_filtered = df_filtered.dropna(subset=["malo", "time_berlin"]).copy()

    target_years = [2021, 2023, 2024, 2025, 2026]
    df_filtered = df_filtered[
        df_filtered["time_berlin"].dt.year.isin(target_years)
    ].copy()

    if df_filtered.empty:
        return pd.DataFrame()

    def _most_recent_continuous_period(
        months_sorted: List[pd.Period],
        period_lengths: List[int] = [12, 24, 36, 48, 60],
    ) -> Optional[List[pd.Period]]:
        if not months_sorted:
            return None

        month_set = set(months_sorted)
        continuous_periods: List[List[pd.Period]] = []

        for period_length in period_lengths:
            if len(months_sorted) < period_length:
                continue

            for i in range(len(months_sorted) - period_length + 1):
                start_month = months_sorted[i]
                candidate = [start_month + j for j in range(period_length)]
                if all(m in month_set for m in candidate):
                    continuous_periods.append(candidate)

        if not continuous_periods:
            return None

        # Prefer the MOST RECENT continuous period (largest start month)
        return max(continuous_periods, key=lambda period: period[0])

    filtered_groups: List[pd.DataFrame] = []

    for malo, group in df_filtered.groupby("malo"):
        group_filtered = group.sort_values("time_berlin").copy()

        # Legacy: determine full years based on raw row counts (before month-valid filtering)
        rows_per_year = group_filtered.groupby(
            group_filtered["time_berlin"].dt.year
        ).size()
        years_in_data = group_filtered["time_berlin"].dt.year.unique().tolist()

        # Apply month validity filtering (>= min_rows_per_month) like legacy
        group_filtered["month"] = group_filtered["time_berlin"].dt.to_period("M")
        month_counts = group_filtered.groupby("month").size()
        valid_months = month_counts[month_counts >= min_rows_per_month].index
        group_valid_months = group_filtered[
            group_filtered["month"].isin(valid_months)
        ].copy()

        filtered_group: Optional[pd.DataFrame] = None
        category = ""

        if len(years_in_data) == 1:
            # Keep all valid months
            filtered_group = group_valid_months
            category = "single_year_keep_valid_months"

        elif (
            len(years_in_data) >= 2
            and not rows_per_year.empty
            and (rows_per_year >= rows_per_full_year).any()
        ):
            # Keep only the full years, but still restricted to valid months
            full_years_available = rows_per_year[
                rows_per_year >= rows_per_full_year
            ].index.tolist()
            filtered_group = group_valid_months[
                group_valid_months["time_berlin"].dt.year.isin(full_years_available)
            ].copy()
            category = (
                f"full_years_only({','.join(map(str, sorted(full_years_available)))})"
            )

        elif len(years_in_data) >= 2:
            # Choose most recent continuous window (12/24/36 months) among valid months
            months_sorted = sorted(group_valid_months["month"].unique().tolist())
            chosen_period = _most_recent_continuous_period(months_sorted)

            if chosen_period:
                filtered_group = group_valid_months[
                    group_valid_months["month"].isin(chosen_period)
                ].copy()
                category = f"continuous_period({len(chosen_period)}m:{chosen_period[0]}→{chosen_period[-1]})"
            else:
                filtered_group = group_valid_months
                category = "no_continuous_period_keep_valid_months"

        else:
            filtered_group = group_valid_months
            category = "fallback_keep_valid_months"

        if debug:
            rows_per_year_str = (
                ", ".join(
                    f"{int(y)}:{int(c)}" for y, c in sorted(rows_per_year.items())
                )
                if not rows_per_year.empty
                else ""
            )
            valid_months_count = (
                int(group_valid_months["month"].nunique())
                if not group_valid_months.empty
                else 0
            )
            print(
                f"malo: {malo} | category: {category} | rows_per_year: [{rows_per_year_str}] | valid_months: {valid_months_count}"
            )

        if filtered_group is None or filtered_group.empty:
            continue

        # Match legacy formatting: years in chronological order if time-sorted
        available_years = (
            filtered_group.sort_values("time_berlin")["time_berlin"]
            .dt.year.unique()
            .tolist()
        )
        filtered_group = filtered_group.copy()
        filtered_group["available_years"] = ", ".join(map(str, available_years))

        filtered_group.drop(columns=["month"], inplace=True, errors="ignore")
        filtered_groups.append(filtered_group)

    if not filtered_groups:
        return pd.DataFrame()

    df_result = pd.concat(filtered_groups, ignore_index=True)

    # Match legacy: compute available_months per malo across the concatenated output
    df_result["month"] = df_result["time_berlin"].dt.to_period("M")
    month_counts = (
        df_result.groupby("malo")["month"]
        .nunique()
        .reset_index(name="available_months")
    )
    df_result = df_result.merge(month_counts, on="malo", how="left")
    df_result.drop(columns=["month"], inplace=True)

    print(
        f"✅ Filtered production data: {len(df_result)} rows, {df_result['malo'].nunique()} malos"
    )
    return df_result


def save_multisheet_excel(df: pd.DataFrame, path: str, max_rows: int = EXCEL_MAX_ROWS):
    """Save large DataFrame to Excel with multiple sheets if needed."""
    df = sanitize_for_excel(df)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if len(df) <= max_rows:
            df.to_excel(writer, sheet_name="Sheet1", index=False, na_rep="")
        else:
            num_sheets = (len(df) // max_rows) + 1
            for i in range(num_sheets):
                start_row = i * max_rows
                end_row = min((i + 1) * max_rows, len(df))
                df.iloc[start_row:end_row].to_excel(
                    writer,
                    sheet_name=sanitize_sheet_name(f"Sheet{i+1}"),
                    index=False,
                    na_rep="",
                )
    print(f"✅ Saved to {path}")


def format_excel_output(file_path: str):
    """Apply basic formatting to Excel output (header highlight, widths)."""
    if load_workbook is None or PatternFill is None or Font is None:
        print("ℹ️ openpyxl not available; skipping Excel formatting")
        return

    wb = load_workbook(file_path)
    highlight_fill = PatternFill(
        start_color="020227", end_color="020227", fill_type="solid"
    )

    white_font = Font(color="FFFFFF")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = highlight_fill
            cell.font = white_font
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col[:2000]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    wb.save(file_path)
    print(f"✅ Formatted Excel: {file_path}")


def aggregate_stammdaten_by_malo(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate stammdaten-style data by malo (robust to missing columns)."""
    df = df.copy()
    if "net_power_kw_unit" not in df.columns:
        if "net_power_kw" in df.columns:
            df["net_power_kw_unit"] = pd.to_numeric(df["net_power_kw"], errors="coerce")
        else:
            df["net_power_kw_unit"] = np.nan

    df["Power in MW"] = pd.to_numeric(df["net_power_kw_unit"], errors="coerce") / 1000

    agg_dict: Dict[str, Any] = {}
    for col in [
        "unit_mastr_id",
        "Projekt",
        "tech",
        "Curtailment & redispatch included",
        "Balancing Cost",
        "Curtailment_value_weighted",
        "Trading Convenience",
    ]:
        if col in df.columns:
            agg_dict[col] = "first"

    if "Power in MW" in df.columns:
        agg_dict["Power in MW"] = "sum"

    if "INB" in df.columns:
        agg_dict["INB"] = lambda x: [convert_date_or_keep_string(date) for date in x]
    if "EEG" in df.columns:
        agg_dict["EEG"] = lambda x: list(pd.Series(x).dropna().astype(str).unique())
    if "AW in EUR/MWh" in df.columns:
        agg_dict["AW in EUR/MWh"] = lambda x: [
            round(float(v), 2) for v in x if is_number(v)
        ]

    blindleister_cols = [
        "weighted_2021_eur_mwh_blindleister",
        "weighted_2023_eur_mwh_blindleister",
        "weighted_2024_eur_mwh_blindleister",
        "weighted_2025_eur_mwh_blindleister",
        "average_weighted_eur_mwh_blindleister",
    ]
    for col in blindleister_cols:
        if col in df.columns:
            agg_dict[col] = "min"

    enervis_cols = [f"enervis_{y}" for y in TARGET_YEARS] + ["avg_enervis"]
    for col in enervis_cols:
        if col in df.columns:
            agg_dict[col] = "min"

    if not agg_dict:
        return df[["malo"]].drop_duplicates()

    df_agg = df.groupby(["malo"], dropna=False).agg(agg_dict).reset_index()
    print(f"✅ Aggregated {len(df_agg)} malos")
    return df_agg


def process_production_data(
    merge_prod_rmv_dayahead: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    """Compute weighted delta and year-agg capacity inputs from production + prices."""
    df = merge_prod_rmv_dayahead.copy()

    if "production_kwh" in df.columns:
        df["production_kwh"] = pd.to_numeric(df["production_kwh"], errors="coerce")
    elif "power_kwh_forecast" in df.columns:
        df["production_kwh"] = pd.to_numeric(df["power_kwh_forecast"], errors="coerce")
    elif "power_kwh" in df.columns:
        df["production_kwh"] = pd.to_numeric(df["power_kwh"], errors="coerce")
    else:
        raise KeyError(
            "Missing production column: expected one of 'production_kwh', 'power_kwh_forecast', or 'power_kwh'"
        )

    df_dropdup = df.drop_duplicates(subset=["malo", "time_berlin", "production_kwh"])

    df_dropdup["deltaspot_eur"] = (
        df_dropdup["production_kwh"] * df_dropdup["dayaheadprice"] / 1000
    ) - (
        df_dropdup["production_kwh"]
        * df_dropdup["monthly_reference_market_price_eur_mwh"]
        / 1000
    )

    df_weighted_delta_by_malo = (
        df_dropdup.groupby(["malo"])
        .agg(
            total_prod_kwh_malo=("production_kwh", "sum"),
            spot_rmv_eur_malo=("deltaspot_eur", "sum"),
        )
        .reset_index()
    )
    df_weighted_delta_by_malo["weighted_delta_permalo"] = (
        df_weighted_delta_by_malo["spot_rmv_eur_malo"]
        / (df_weighted_delta_by_malo["total_prod_kwh_malo"] / 1000)
    ).round(2)

    total_prod = df_dropdup.groupby(["malo"])["production_kwh"].sum()

    df_monthly_delta = (
        df_dropdup.groupby(["year", "month", "malo"])
        .agg(
            deltaspot_eur_monthly=("deltaspot_eur", "sum"),
            available_months=("available_months", "first"),
            available_years=("available_years", "first"),
        )
        .reset_index()
    )
    df_monthly_delta["total_prod_kwh"] = df_monthly_delta["malo"].map(total_prod)
    df_monthly_delta["total_prod_mwh"] = df_monthly_delta["total_prod_kwh"] / 1000

    df_capacity_inputs_by_malo = (
        df_monthly_delta.groupby(["malo"], dropna=False)
        .agg(
            available_months=("available_months", "first"),
            available_years=("available_years", "first"),
            total_prod_mwh=("total_prod_mwh", "first"),
        )
        .reset_index()
    )

    return {
        "weighted_delta_permalo": df_weighted_delta_by_malo,
        "year_agg": df_capacity_inputs_by_malo,
    }


def build_monthly_weighted_delta_comparison(
    df_original: pd.DataFrame,
    df_forecast: pd.DataFrame,
    output_path: Path,
) -> pd.DataFrame:
    """Compute monthly weighted_delta_permalo per malo for original vs forecast production and save as Excel."""

    def _monthly_weighted_delta(df_input: pd.DataFrame) -> pd.DataFrame:
        df = df_input.copy()

        if "production_kwh" in df.columns:
            df["production_kwh"] = pd.to_numeric(df["production_kwh"], errors="coerce")
        elif "power_kwh_forecast" in df.columns:
            df["production_kwh"] = pd.to_numeric(
                df["power_kwh_forecast"], errors="coerce"
            )
        elif "power_kwh" in df.columns:
            df["production_kwh"] = pd.to_numeric(df["power_kwh"], errors="coerce")
        else:
            raise KeyError(
                "Missing production column for monthly weighted delta: expected one of "
                "'production_kwh', 'power_kwh_forecast', or 'power_kwh'"
            )

        if "time_berlin" in df.columns:
            df["time_berlin"] = pd.to_datetime(df["time_berlin"], errors="coerce")
            if "year" not in df.columns:
                df["year"] = df["time_berlin"].dt.year
            if "month" not in df.columns:
                df["month"] = df["time_berlin"].dt.month

        required_cols = [
            "malo",
            "year",
            "month",
            "dayaheadprice",
            "monthly_reference_market_price_eur_mwh",
            "production_kwh",
        ]
        missing_cols = [c for c in required_cols if c not in df.columns]
        if missing_cols:
            raise KeyError(
                f"Missing required columns for monthly weighted delta: {missing_cols}"
            )

        df = df.drop_duplicates(subset=["malo", "time_berlin", "production_kwh"])

        df["deltaspot_eur"] = (
            pd.to_numeric(df["production_kwh"], errors="coerce")
            * pd.to_numeric(df["dayaheadprice"], errors="coerce")
            / 1000
        ) - (
            pd.to_numeric(df["production_kwh"], errors="coerce")
            * pd.to_numeric(
                df["monthly_reference_market_price_eur_mwh"], errors="coerce"
            )
            / 1000
        )

        monthly = (
            df.groupby(["malo", "year", "month"], dropna=False)
            .agg(
                monthly_prod_kwh=("production_kwh", "sum"),
                monthly_deltaspot_eur=("deltaspot_eur", "sum"),
            )
            .reset_index()
        )

        monthly["weighted_delta_permalo_monthly"] = (
            monthly["monthly_deltaspot_eur"] / (monthly["monthly_prod_kwh"] / 1000)
        ).round(2)

        monthly["date_month"] = pd.to_datetime(
            {
                "year": pd.to_numeric(monthly["year"], errors="coerce"),
                "month": pd.to_numeric(monthly["month"], errors="coerce"),
                "day": 1,
            },
            errors="coerce",
        )
        return monthly

    monthly_original = _monthly_weighted_delta(df_original).rename(
        columns={
            "monthly_prod_kwh": "monthly_prod_kwh_original",
            "monthly_deltaspot_eur": "monthly_deltaspot_eur_original",
            "weighted_delta_permalo_monthly": "weighted_delta_permalo_monthly_original",
        }
    )

    monthly_forecast = _monthly_weighted_delta(df_forecast).rename(
        columns={
            "monthly_prod_kwh": "monthly_prod_kwh_forecast",
            "monthly_deltaspot_eur": "monthly_deltaspot_eur_forecast",
            "weighted_delta_permalo_monthly": "weighted_delta_permalo_monthly_forecast",
        }
    )

    monthly_comparison = monthly_original.merge(
        monthly_forecast,
        on=["malo", "year", "month", "date_month"],
        how="outer",
    )

    monthly_comparison["different_forecast-original"] = (
        monthly_comparison["weighted_delta_permalo_monthly_forecast"]
        - monthly_comparison["weighted_delta_permalo_monthly_original"]
    ).round(2)

    monthly_comparison = monthly_comparison.sort_values(
        ["malo", "date_month"]
    ).reset_index(drop=True)

    save_multisheet_excel(monthly_comparison, str(output_path))
    print(f"✅ Saved monthly weighted-delta comparison: {output_path}")
    return monthly_comparison


# =============================================================================
# CURTAILMENT FORECASTING (ML) (ported from code_test_refactored.py)
# =============================================================================


def feature_engineering_classification(
    df: pd.DataFrame,
    feature_names: List[str],
) -> Tuple[pd.DataFrame, List[str], bool]:
    """Feature engineering for classification – must match training."""

    df = df.copy()

    if "volume__mw_imbalance" in df.columns:
        df["volume__mw_imbalance"] = pd.to_numeric(
            df["volume__mw_imbalance"], errors="coerce"
        ).fillna(0)
    else:
        df["volume__mw_imbalance"] = 0.0

    if (
        "curtailment_kWh_per_kw" in df.columns
        and df["curtailment_kWh_per_kw"].notna().any()
    ):
        actual_mask = df["curtailment_kWh_per_kw"].notna()
        df["curtailment_flag"] = np.where(
            actual_mask,
            (pd.to_numeric(df["curtailment_kWh_per_kw"], errors="coerce") > 0).astype(
                int
            ),
            np.nan,
        )
        has_actual_values = True
    else:
        has_actual_values = False

    df = df.rename(columns={"dayaheadprice": "dayaheadprice_eur_mwh"})

    for col in ["dayaheadprice_eur_mwh", "rebap_euro_per_mwh"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "net_power_kw_unit" not in df.columns:
        capacity_candidate_cols = [
            "sum_power_kw_malo",
            "net_power_kw",
            "sum_power_kw_malo_x",
            "sum_power_kw_malo_y",
            "net_power_kw_x",
            "net_power_kw_y",
            "Power in MW",
            "Power in MW_x",
            "Power in MW_y",
        ]
        capacity_col = next(
            (c for c in capacity_candidate_cols if c in df.columns), None
        )

        if capacity_col is None:
            raise KeyError(
                "Missing installed-capacity column for forecasting: expected one of "
                "'net_power_kw_unit', 'sum_power_kw_malo', 'net_power_kw', or merged variants (_x/_y)."
            )

        if capacity_col.startswith("Power in MW"):
            df["net_power_kw_unit"] = (
                pd.to_numeric(df[capacity_col], errors="coerce") * 1000
            )
        else:
            df["net_power_kw_unit"] = pd.to_numeric(df[capacity_col], errors="coerce")

        print(
            f"⚠️ 'net_power_kw_unit' missing; using '{capacity_col}' as fallback for installed capacity"
        )

    denom_kw = pd.to_numeric(df["net_power_kw_unit"], errors="coerce").replace(
        0, np.nan
    )
    df["quarterly_energy_kWh_per_kw"] = (
        pd.to_numeric(df["power_kwh"], errors="coerce") / denom_kw
    )

    df["DA_negative_flag"] = (df["dayaheadprice_eur_mwh"] < 0).astype(int)
    df["DA_negative_flag_lag_1"] = df["DA_negative_flag"].shift(1)

    df["rebap_negative_flag"] = (df["rebap_euro_per_mwh"] < 0).astype(int)
    df["rebap_negative_flag_lag_1"] = df["rebap_negative_flag"].shift(1)

    missing_features = [f for f in feature_names if f not in df.columns]
    if missing_features:
        print(
            "⚠️ Missing classification features from model metadata; "
            f"creating 0.0 fallback columns: {missing_features}"
        )
        for f in missing_features:
            df[f] = 0.0

    for f in feature_names:
        df[f] = pd.to_numeric(df[f], errors="coerce")

    df_clean = df.dropna(subset=feature_names).copy()
    if df_clean.empty:
        raise ValueError(
            "No valid rows after classification cleaning (NaNs in features)."
        )

    return df_clean, feature_names, has_actual_values


def predict_curtailment_classification(
    df_new_prediction: pd.DataFrame,
    model_path: str,
    metadata_path: str,
    plot: bool = False,
) -> Optional[Dict[str, Any]]:
    """Run classification model on new data."""

    if joblib is None:
        print("⚠️ joblib not available; cannot run classification forecasting.")
        return None

    print_header("CLASSIFICATION – LOADING MODEL & METADATA")

    try:
        best_model = joblib.load(model_path)
        with open(metadata_path, "r") as f:
            metadata = json.load(f)
    except FileNotFoundError as e:
        print(f"❌ Error loading classification files: {e}")
        return None

    if not isinstance(metadata, dict):
        raise ValueError("Classification metadata must be a JSON object.")

    feature_names = metadata["feature_names"]
    optimal_threshold = float(
        pd.to_numeric(metadata["80/20_optimal_threshold"], errors="coerce")
    )

    print(f"Using optimal threshold: {optimal_threshold:.4f}")

    print_header("CLASSIFICATION – FEATURE ENGINEERING")
    df_clean, available_features, has_actual_values = (
        feature_engineering_classification(
            df_new_prediction,
            feature_names,
        )
    )

    X_new = df_clean[available_features]
    print(f"Classification rows: {len(X_new)}, features used: {available_features}")

    print_header("CLASSIFICATION – PREDICTION")
    y_proba = best_model.predict_proba(X_new)[:, 1]
    y_pred = (y_proba >= optimal_threshold).astype(int)

    df_clean["predicted_curtailment_probability"] = y_proba
    df_clean["predicted_curtailment_flag"] = y_pred
    df_clean["prediction_timestamp_cls"] = (
        pd.Timestamp.now(tz="Europe/Berlin")
        .tz_localize(None)
        .strftime("%Y-%m-%d %H:%M")
    )

    print(
        f"Predicted curtailment == 1 for {y_pred.sum():,} rows "
        f"({y_pred.mean()*100:.1f}% of classified rows)."
    )

    if (
        has_actual_values
        and "curtailment_flag" in df_clean.columns
        and accuracy_score is not None
    ):
        actual_mask = df_clean["curtailment_flag"].notna()
        if actual_mask.any():
            y_actual = df_clean.loc[actual_mask, "curtailment_flag"].astype(int)
            y_pred_actual = y_pred[actual_mask.values]
            y_proba_actual = y_proba[actual_mask.values]

            accuracy = accuracy_score(y_actual, y_pred_actual)
            precision = precision_score(y_actual, y_pred_actual, zero_division=0)
            recall = recall_score(y_actual, y_pred_actual, zero_division=0)
            f1 = f1_score(y_actual, y_pred_actual, zero_division=0)
            roc_auc = roc_auc_score(y_actual, y_proba_actual)
            avg_precision = average_precision_score(y_actual, y_proba_actual)
        else:
            accuracy = precision = recall = f1 = roc_auc = avg_precision = None

        print_header("CLASSIFICATION – METRICS (ACTUALS AVAILABLE)")
        print(f"Accuracy:      {accuracy:.4f}")
        print(f"Precision:     {precision:.4f}")
        print(f"Recall:        {recall:.4f}")
        print(f"F1-Score:      {f1:.4f}")
        print(f"ROC AUC:       {roc_auc:.4f}")
        print(f"Avg Precision: {avg_precision:.4f}")
    else:
        accuracy = precision = recall = f1 = roc_auc = avg_precision = None
        if (
            has_actual_values
            and "curtailment_flag" in df_clean.columns
            and accuracy_score is None
        ):
            print("ℹ️ sklearn.metrics not available; skipping classification metrics.")
        else:
            print("ℹ️ No actual curtailment available for classification metrics.")

    if plot and plt is not None and sns is not None:
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))
        sns.histplot(y_proba, bins=30, kde=True, ax=ax1)
        ax1.axvline(optimal_threshold, color="red", linestyle="--", label="Threshold")
        ax1.set_title("Linear Scale")
        ax1.set_xlabel("P(curtailment=1)")
        ax1.set_ylabel("Frequency")
        ax1.legend()

        sns.histplot(y_proba, bins=30, kde=True, ax=ax2)
        ax2.axvline(optimal_threshold, color="red", linestyle="--", label="Threshold")
        ax2.set_yscale("log")
        ax2.set_title("Log Scale")
        ax2.set_xlabel("P(curtailment=1)")
        ax2.set_ylabel("Frequency (log scale)")
        ax2.legend()

        plt.suptitle("Predicted Probability Distribution", fontsize=14)
        plt.tight_layout()
        plt.show()
    elif plot:
        print("ℹ️ Plot requested but matplotlib/seaborn not available; skipping plots.")

    return {
        "predictions": df_clean,
        "model": best_model,
        "features_used": available_features,
        "optimal_threshold": optimal_threshold,
        "prediction_metrics": {
            "accuracy": accuracy,
            "precision": precision,
            "recall": recall,
            "f1": f1,
            "roc_auc": roc_auc,
            "avg_precision": avg_precision,
        },
        "prediction_stats": {
            "positive_predictions": int(y_pred.sum()),
            "negative_predictions": int(len(y_pred) - y_pred.sum()),
            "positive_rate": float(y_pred.mean()),
            "mean_probability": float(y_proba.mean()),
            "std_probability": float(y_proba.std()),
            "total_predictions": int(len(y_pred)),
        },
    }


def feature_engineering_regression(
    df: pd.DataFrame,
    reg_features: List[str],
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """Feature engineering for regression – must match training."""

    df = df.copy()

    exo_features = [
        "quarterly_energy_kWh_per_kw",
        "enwex_percentage",
        "dayaheadprice_eur_mwh",
        "rebap_euro_per_mwh",
        "volume__mw_imbalance",
        "id500_eur_mwh",
        "rmv_eur_per_mwh",
    ]

    for col in exo_features:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").ffill().bfill()
        else:
            print(f"⚠️ Regression: missing feature {col} – filled with 0.")
            df[col] = 0.0

    # if "curtailment_kWh_per_kw" in df.columns:
    #     df["curt_lag_1"] = df["curtailment_kWh_per_kw"].shift(1)
    #     df["curt_lag_2"] = df["curtailment_kWh_per_kw"].shift(2)

    available_features = [f for f in reg_features if f in df.columns]
    missing_features = [f for f in reg_features if f not in df.columns]

    if missing_features:
        print(f"⚠️ Missing regression features: {missing_features}")
        if not available_features:
            raise ValueError("No required regression features available.")

    df_clean = df.dropna(subset=available_features).copy()
    if df_clean.empty:
        raise ValueError("No valid rows after regression cleaning (NaNs in features).")

    X_new = df_clean[available_features].apply(pd.to_numeric, errors="coerce")
    return df_clean, X_new, available_features


def plot_regression_predictions(df_clean: pd.DataFrame):
    """Plot regression prediction distribution + time plot."""

    if plt is None:
        print("ℹ️ matplotlib not available; skipping regression plots.")
        return

    y_pred = df_clean["predicted_curtailment_kWh_per_kw"].values
    has_actual = "curtailment_kWh_per_kw" in df_clean.columns
    y_actual = df_clean["curtailment_kWh_per_kw"].values if has_actual else None

    fig, axes = plt.subplots(1, 2, figsize=(12, 5))

    if has_actual:
        y_actual_plot = df_clean[df_clean["curtailment_kWh_per_kw"] > 0][
            "curtailment_kWh_per_kw"
        ].values

        axes[0].hist(
            y_actual_plot,
            bins=30,
            alpha=0.6,
            color="blue",
            edgecolor="black",
            label="Actual",
        )

    axes[0].hist(
        y_pred,
        bins=30,
        alpha=0.6 if has_actual else 0.7,
        color="green",
        edgecolor="black",
        label="Predicted",
    )

    axes[0].set_xlabel("Curtailment (kWh/kW)")
    axes[0].set_ylabel("Frequency")
    axes[0].set_title("Curtailment Distribution")
    axes[0].grid(True, alpha=0.3)
    if has_actual:
        axes[0].legend()

    time_col = None
    for candidate in ["delivery_start_berlin", "time_berlin", "timestamp"]:
        if candidate in df_clean.columns:
            time_col = candidate
            break

    if time_col:
        df_sorted = df_clean.sort_values(time_col)
        x_vals = df_sorted[time_col]
        y_pred_sorted = df_sorted["predicted_curtailment_kWh_per_kw"]
        axes[1].plot(x_vals, y_pred_sorted, color="red", linewidth=1, label="Predicted")

        if has_actual and y_actual is not None:
            axes[1].plot(
                x_vals,
                df_sorted["curtailment_kWh_per_kw"],
                color="blue",
                linewidth=0.8,
                alpha=0.4,
                label="Actual",
            )
        axes[1].set_xlabel("Time")
        axes[1].tick_params(axis="x", rotation=45)
    else:
        x_vals = np.arange(len(y_pred))
        axes[1].plot(x_vals, y_pred, color="red", linewidth=1.2, label="Predicted")
        if has_actual and y_actual is not None:
            axes[1].plot(
                x_vals,
                y_actual,
                color="blue",
                linewidth=1.2,
                alpha=0.8,
                label="Actual",
            )
        axes[1].set_xlabel("Sample Index")

    axes[1].set_ylabel("Curtailment (kWh/kW)")
    axes[1].set_title("Curtailment Over Time")
    if has_actual:
        axes[1].legend()
    axes[1].grid(True, alpha=0.3)

    plt.tight_layout()
    plt.show()


def predict_curtailment_regression(
    df_reg_input: pd.DataFrame,
    model_path: str,
    metadata_path: str,
    plot: bool = False,
) -> Optional[Dict[str, Any]]:
    """Run regression model on subset of rows (already filtered by classification)."""

    if joblib is None:
        print("⚠️ joblib not available; cannot run regression forecasting.")
        return None

    try:
        best_model = joblib.load(model_path)
        with open(metadata_path, "r") as f:
            metadata = json.load(f)
    except FileNotFoundError as e:
        print(f"❌ Error loading regression files: {e}")
        return None

    default_reg_features = [
        "quarterly_energy_kWh_per_kw",
        "enwex_percentage",
        "dayaheadprice_eur_mwh",
        "rebap_euro_per_mwh",
        "volume__mw_imbalance",
        "id500_eur_mwh",
        "rmv_eur_per_mwh",
    ]

    if not isinstance(metadata, dict):
        raise ValueError("Regression metadata must be a JSON object.")

    reg_features = (
        metadata["feature_names"]
        if metadata.get("feature_names")
        else default_reg_features
    )

    if df_reg_input.empty:
        print("ℹ️ No rows passed to regression (no predicted curtailment = 1).")
        return {
            "predictions": df_reg_input.assign(predicted_curtailment_kWh_per_kw=np.nan),
            "model": best_model,
            "features_used": reg_features,
            "prediction_metrics": {"mse": None, "mae": None, "mape": None, "r2": None},
            "prediction_stats": {
                "mean": None,
                "std": None,
                "min": None,
                "max": None,
                "count": 0,
            },
        }

    print_header("REGRESSION – FEATURE ENGINEERING ON FILTERED ROWS")
    df_clean, X_new, used_features = feature_engineering_regression(
        df_reg_input, reg_features
    )
    print(f"Regression rows: {len(X_new)}, features used: {used_features}")

    y_pred = best_model.predict(X_new)
    df_clean["predicted_curtailment_kWh_per_kw"] = y_pred
    # df_clean["prediction_timestamp_reg"] = pd.Timestamp.now(tz="Europe/Berlin").tz_localize(None).strftime("%Y-%m-%d %H:%M")

    mse = mae = mape = r2 = None

    if "curtailment_kWh_per_kw" in df_clean.columns and mean_squared_error is not None:
        actual_mask = df_clean["curtailment_kWh_per_kw"].notna()
        if actual_mask.any():
            y_actual = df_clean.loc[actual_mask, "curtailment_kWh_per_kw"].astype(float)
            y_pred_actual = (
                pd.Series(y_pred, index=df_clean.index).loc[actual_mask].astype(float)
            )
            mse = mean_squared_error(y_actual, y_pred_actual)
            mae = mean_absolute_error(y_actual, y_pred_actual)
            mape = mean_absolute_percentage_error(y_actual, y_pred_actual)
            r2 = r2_score(y_actual, y_pred_actual)
            print_header("REGRESSION – METRICS (ACTUALS AVAILABLE)")
            print(f"MSE:  {mse:.4f}")
            print(f"MAE:  {mae:.4f}")
            print(f"MAPE: {mape:.4f}")
            print(f"R²:   {r2:.4f}")
        else:
            print(
                "ℹ️ Actual curtailment column exists but has no non-null values; skipping regression metrics."
            )
    else:
        if "curtailment_kWh_per_kw" in df_clean.columns and mean_squared_error is None:
            print("ℹ️ sklearn.metrics not available; skipping regression metrics.")
        else:
            print("ℹ️ No actual curtailment available for regression metrics.")

    if plot:
        plot_regression_predictions(df_clean)

    return {
        "predictions": df_clean,
        "model": best_model,
        "features_used": used_features,
        "prediction_metrics": {"mse": mse, "mae": mae, "mape": mape, "r2": r2},
        "prediction_stats": {
            "mean": float(np.mean(y_pred)),
            "std": float(np.std(y_pred)),
            "min": float(np.min(y_pred)),
            "max": float(np.max(y_pred)),
            "count": int(len(y_pred)),
        },
    }


def run_curtailment_forecast(
    df_new_prediction: pd.DataFrame,
    cls_model_path: str,
    cls_meta_path: str,
    reg_model_path: str,
    reg_meta_path: str,
    plot_class: bool = False,
    plot_reg: bool = False,
) -> Optional[Dict[str, Any]]:
    """Classification on all rows, regression on predicted-curtailment subset, then merge back."""

    cls_results = predict_curtailment_classification(
        df_new_prediction,
        model_path=cls_model_path,
        metadata_path=cls_meta_path,
        plot=plot_class,
    )

    if cls_results is None:
        return None

    df_cls = cls_results["predictions"].copy()
    if "predicted_curtailment_flag" not in df_cls.columns:
        print("❌ Classification result missing 'predicted_curtailment_flag'.")
        return {"classification": cls_results, "regression": None, "combined": df_cls}

    df_for_reg = df_cls[df_cls["predicted_curtailment_flag"] == 1].copy()
    print_header("PIPELINE – ROWS FOR REGRESSION")
    print(f"Rows flagged as curtailment (1): {len(df_for_reg)}")

    reg_results = predict_curtailment_regression(
        df_for_reg,
        model_path=reg_model_path,
        metadata_path=reg_meta_path,
        plot=plot_reg,
    )

    df_combined = df_cls.copy()
    df_combined["predicted_curtailment_kWh_per_kw"] = np.nan

    if reg_results is not None and not reg_results["predictions"].empty:
        df_reg_pred = reg_results["predictions"].copy()
        merge_keys = ["malo", "delivery_start_berlin"]
        df_reg_pred = df_reg_pred[merge_keys + ["predicted_curtailment_kWh_per_kw"]]

        df_combined = df_combined.merge(
            df_reg_pred,
            on=merge_keys,
            how="left",
            suffixes=("", "_reg"),
        )

        df_combined["predicted_curtailment_kWh_per_kw"] = df_combined[
            "predicted_curtailment_kWh_per_kw_reg"
        ]
        df_combined.drop(columns=["predicted_curtailment_kWh_per_kw_reg"], inplace=True)
        df_combined["predicted_curtailment_kWh_per_kw"] = pd.to_numeric(
            df_combined["predicted_curtailment_kWh_per_kw"],
            errors="coerce",
        ).fillna(0)

    return {
        "classification": cls_results,
        "regression": reg_results,
        "combined": df_combined,
    }


def set_paths_for_category(category: str) -> Dict[str, str]:
    base_path = os.getenv("CURTAILMENT_MODEL_BASE_PATH", MODEL_BASE_PATH)
    base_path = (base_path or "").rstrip("/")

    category_mapping = {
        "PV_rules": "PV_rules",
        "PV_no_rules": "PV_no_rules",
        "WIND_rules": "WIND_rules",
        "WIND_no_rules": "WIND_no_rules",
    }

    if category not in category_mapping:
        raise ValueError(f"Unknown category: {category}")

    if not base_path:
        raise ValueError(
            "MODEL_BASE_PATH (or CURTAILMENT_MODEL_BASE_PATH) is not set; cannot locate curtailment models."
        )

    folder_name = category_mapping[category]
    category_dir = f"{base_path}/{folder_name}"
    # model_result_dir = f"{category_dir}/model result"

    paths = {
        "CLASS_MODEL_PATH": f"{category_dir}/classification_best_model_{folder_name}.joblib",
        "CLASS_META_PATH": f"{category_dir}/classification_metadata_{folder_name}.json",
        "REG_MODEL_PATH": f"{category_dir}/regression_best_model_{folder_name}.joblib",
        "REG_META_PATH": f"{category_dir}/regression_metadata_{folder_name}.json",
    }

    missing = [key for key, path in paths.items() if not os.path.exists(path)]
    if missing:
        missing_details = ", ".join([f"{key}={paths[key]}" for key in missing])
        raise FileNotFoundError(
            "Updated curtailment artifacts not found. Expected only new model-result files: "
            f"{missing_details}"
        )

    return paths


def run_curtailment_forecast_multi_category(
    df_ts: pd.DataFrame,
    plot_class: bool = False,
    plot_reg: bool = False,
) -> Optional[Dict[str, Any]]:
    """Run the full classification->regression pipeline for each EEG category."""

    df_ts = df_ts.copy()

    if "category" not in df_ts.columns:
        if "tech" in df_ts.columns and "EEG" in df_ts.columns:
            df_ts["category"] = np.where(
                (df_ts["tech"] == "PV") & (df_ts["EEG"] == "no rules"),
                "PV_no_rules",
                np.where(
                    df_ts["tech"] == "PV",
                    "PV_rules",
                    np.where(
                        (df_ts["tech"] == "WIND") & (df_ts["EEG"] == "no rules"),
                        "WIND_no_rules",
                        "WIND_rules",
                    ),
                ),
            )
        else:
            raise ValueError(
                "df_ts must contain 'category' (or at least 'tech' and 'EEG' to derive it)."
            )

    if df_ts["category"].isna().any():
        missing_malo = df_ts.loc[df_ts["category"].isna(), "malo"].unique()
        raise ValueError(
            f"Some malos in time series do not have a category: {missing_malo}"
        )

    all_combined: List[pd.DataFrame] = []
    all_results_by_category: Dict[str, Any] = {}

    for category, df_cat in df_ts.groupby("category"):
        print_header(f"🍫🍫 RUNNING CATEGORY: {category}")
        paths = set_paths_for_category(category)

        res = run_curtailment_forecast(
            df_new_prediction=df_cat,
            cls_model_path=paths["CLASS_MODEL_PATH"],
            cls_meta_path=paths["CLASS_META_PATH"],
            reg_model_path=paths["REG_MODEL_PATH"],
            reg_meta_path=paths["REG_META_PATH"],
            plot_class=plot_class,
            plot_reg=plot_reg,
        )

        if res is not None:
            combined_cat = res["combined"].copy()
            combined_cat["category"] = category
            all_combined.append(combined_cat)
            all_results_by_category[category] = res

    if not all_combined:
        return None

    df_all = pd.concat(all_combined, ignore_index=True)
    return {"by_category": all_results_by_category, "combined": df_all}


def generate_output_sheets(
    df: pd.DataFrame, has_production: bool = False, has_forecast: bool = False
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Generate three output sheets for customer report."""

    df = normalize_nan_strings(df.copy())

    # Only include Enervis year columns that exist and have at least one non-null value.
    enervis_year_cols = [
        c
        for c in [f"enervis_{y}" for y in TARGET_YEARS]
        if c in df.columns and not df[c].isna().all()
    ]

    sheet1_order = [
        "malo",
        "Projekt",
        "Technology",
        "Power in MW",
        "INB",
        "EEG",
        "AW in EUR/MWh",
        "weighted_2021_eur_mwh_blindleister",
        "weighted_2023_eur_mwh_blindleister",
        "weighted_2024_eur_mwh_blindleister",
        "weighted_2025_eur_mwh_blindleister",
        "average_weighted_eur_mwh_blindleister",
        *enervis_year_cols,
        "avg_enervis",
        "weighted_delta_permalo",
        "forecast_weighted_delta_permalo" if has_forecast else None,
        "Curtailment & redispatch included",
        "available_months",
        "available_years",
        "capacity_factor_percent",
        "forecast_capacity_factor_percent" if has_forecast else None,
        "Balancing Cost",
        "Curtailment_value_weighted",
        "AMW/RWM Delta",
        "Fair Value",
        "Trading Convenience",
    ]

    sheet2_order = [
        "malo",
        "Projekt",
        "Technology",
        "Power in MW",
        "INB",
        "EEG",
        "AW in EUR/MWh",
        "average_weighted_eur_mwh_blindleister",
        "avg_enervis",
        "weighted_delta_permalo",
        "forecast_weighted_delta_permalo" if has_forecast else None,
        "Curtailment & redispatch included",
        "available_months",
        "available_years",
        "capacity_factor_percent",
        "Balancing Cost",
        "Curtailment_value_weighted",
        "AMW/RWM Delta",
        "Fair Value",
        "Trading Convenience",
        "Fee EUR/MWh",
    ]

    sheet3_order = [
        "malo",
        "Projekt",
        "Technology",
        "Power in MW",
        "AW in EUR/MWh",
        "Fee EUR/MWh",
    ]

    if not has_production:
        for col in [
            "weighted_delta_permalo",
            "forecast_weighted_delta_permalo",
            "forecast_capacity_factor_percent",
            "available_months",
            "available_years",
            "capacity_factor_percent",
        ]:
            if col in sheet1_order:
                sheet1_order.remove(col)
            if col in sheet2_order:
                sheet2_order.remove(col)

    sheet1 = sanitize_for_excel(
        normalize_nan_strings(ensure_and_reorder(df.copy(), sheet1_order))
    )
    sheet2 = sanitize_for_excel(
        normalize_nan_strings(ensure_and_reorder(df.copy(), sheet2_order))
    )
    sheet3 = sanitize_for_excel(
        normalize_nan_strings(ensure_and_reorder(df.copy(), sheet3_order))
    )
    return sheet1, sheet2, sheet3


def main():
    """Integrated pipeline: fixings + Blindleister + Enervis + (optional) production report."""

    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    input_path = Path(STAMMDATEN_PATH)
    customer_name = input_path.stem.split("_", 1)[0]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = OUTPUT_FOLDER / f"{customer_name}_pricing_{timestamp}.xlsx"

    print_header("STARTING PIPELINE")
    print(f"✅ Input: {input_path}")
    print(f"✅ Output: {out_path}")

    print_header("STEP 1: LOADING STAMMDATEN")
    df_stamm = load_stammdaten(input_path)

    if (
        "net_power_kw_unit" not in df_stamm.columns
        and "net_power_kw" in df_stamm.columns
    ):
        df_stamm["net_power_kw_unit"] = pd.to_numeric(
            df_stamm["net_power_kw"], errors="coerce"
        )

    print_header("STEP 2: EEG CATEGORIES")
    df_stamm = set_category_based_on_conditions(df_stamm)

    print_header("STEP 3: APPLYING FIXINGS")
    try:
        df_fixings = load_bigquery_fixings(PROJECT_ID)
        fixings = extract_all_fixings(df_fixings, year=str("2026"))
        df_stamm = apply_fixings_to_stammdaten(df_stamm, fixings)
    except Exception as e:
        print(f"⚠️ Fixings step skipped/failed: {e}")

    # =============================================================================
    # STEP 4: BLINDLEISTER + ENERVIS
    # =============================================================================

    print_header("STEP 4: BLINDLEISTER + ENERVIS")

    see_ids: List[str] = []
    if "unit_mastr_id" in df_stamm.columns:
        see_ids = [
            str(id_).strip().upper()
            for id_ in df_stamm["unit_mastr_id"].dropna()
            if str(id_).strip().upper().startswith("SEE")
        ]
    see_ids = list(set(see_ids))
    print(f"📊 Found {len(see_ids)} unique SEE IDs")

    see_ids_wind: List[str] = []
    if "unit_mastr_id" in df_stamm.columns and "tech" in df_stamm.columns:
        for id_ in df_stamm["unit_mastr_id"].dropna():
            sid = str(id_).strip().upper()
            if not sid.startswith("SEE"):
                continue
            tech_vals = df_stamm.loc[
                df_stamm["unit_mastr_id"].astype(str).str.strip().str.upper() == sid,
                "tech",
            ]
            tech = (
                tech_vals.astype(str).str.upper().iloc[0] if not tech_vals.empty else ""
            )
            if tech == "WIND":
                see_ids_wind.append(sid)
    see_ids_wind = list(set(see_ids_wind))

    blindleister_results = pd.DataFrame()
    if see_ids:
        try:
            blindleister = BlindleisterAPI(
                BLINDLEISTER_EMAIL, BLINDLEISTER_PASSWORD, BLINDLEISTER_HARDCODED_TOKEN
            )
            if not BLINDLEISTER_HARDCODED_TOKEN:
                blindleister.get_token()
            df_flat = blindleister.get_market_prices(see_ids, TARGET_YEARS)

            # saving monthly result
            df_flat.to_excel(
                OUTPUT_FOLDER / f"blindleister_market_prices_monthly.xlsx", index=False
            )

            blind_units_set = set(df_flat["unit_mastr_id"].unique())
            stamm_units_set = set(df_stamm["unit_mastr_id"].unique())
            missing_in_blind = stamm_units_set - blind_units_set
            if missing_in_blind:
                print(
                    f"⚠️ {len(missing_in_blind)} unit_mastr_id from stamm are missing in blind. {list(missing_in_blind)}"
                )
            else:
                print("✅ All unit_mastr_id from Blindleister are present in Stamm")

            if not df_flat.empty:
                blindleister_results = process_blindleister_data(df_flat)
        except Exception as e:
            print(f"❌ Blindleister processing error: {e}")

    df_wind_from_stammdaten = pd.DataFrame()
    if "tech" in df_stamm.columns and "turbine_model" in df_stamm.columns:
        mask_wind_manual = (
            df_stamm["tech"].astype(str).str.upper().eq("WIND")
            & df_stamm["turbine_model"].notna()
        )
        if not df_stamm[mask_wind_manual].empty:
            df_wind_from_stammdaten = df_stamm[mask_wind_manual].copy()

    df_wind_from_generator_details = pd.DataFrame()
    if see_ids_wind:
        try:
            blindleister_for_mastr = BlindleisterAPI(
                BLINDLEISTER_EMAIL,
                BLINDLEISTER_PASSWORD,
                BLINDLEISTER_HARDCODED_TOKEN,
            )
            if not BLINDLEISTER_HARDCODED_TOKEN:
                blindleister_for_mastr.get_token()

            df_gen_details = blindleister_for_mastr.get_generator_details(
                see_ids_wind, year=2024
            )
            if not df_gen_details.empty:
                df_gen_details = df_gen_details.copy()
                df_gen_details["unit_mastr_id"] = (
                    df_gen_details["unit_mastr_id"].astype(str).str.strip().str.upper()
                )
                df_stamm_join = df_stamm.copy()
                if "unit_mastr_id" in df_stamm_join.columns:
                    df_stamm_join["unit_mastr_id"] = (
                        df_stamm_join["unit_mastr_id"]
                        .astype(str)
                        .str.strip()
                        .str.upper()
                    )

                # Save the raw generator-details enriched with all available stammdaten columns.
                # This is useful for auditing and debugging turbine matching / Enervis inputs.
                try:
                    if "unit_mastr_id" in df_stamm_join.columns:
                        df_gen_details_merged = df_gen_details.merge(
                            df_stamm_join,
                            on="unit_mastr_id",
                            how="left",
                            suffixes=("_gen", ""),
                        )
                    else:
                        df_gen_details_merged = df_gen_details.copy()

                    out_gd_path = (
                        OUTPUT_FOLDER
                        / f"generator_details_merged_stammdaten_{timestamp}.xlsx"
                    )
                    df_gen_details_merged.to_excel(out_gd_path, index=False)
                    print(
                        f"💾 Saved generator-details merged with stammdaten: {out_gd_path}"
                    )
                except Exception as e:
                    print(
                        f"⚠️ Could not save generator-details merged with stammdaten: {e}"
                    )

                if "energy_source" in df_gen_details.columns:
                    df_wind_from_generator_details = df_gen_details[
                        df_gen_details["energy_source"]
                        .astype(str)
                        .str.contains("wind", case=False, na=False)
                    ].copy()
                else:
                    df_wind_from_generator_details = df_gen_details.copy()
                    df_wind_from_generator_details["energy_source"] = "wind"

                if (
                    "unit_mastr_id" in df_stamm_join.columns
                    and "malo" in df_stamm_join.columns
                ):
                    df_wind_from_generator_details = (
                        df_wind_from_generator_details.merge(
                            df_stamm_join[["malo", "unit_mastr_id"]],
                            on="unit_mastr_id",
                            how="left",
                        )
                    )
                if (
                    "net_power_kw" not in df_wind_from_generator_details.columns
                    and "net_power_kw_unit" in df_wind_from_generator_details.columns
                ):
                    df_wind_from_generator_details.rename(
                        columns={"net_power_kw_unit": "net_power_kw"}, inplace=True
                    )
                df_wind_from_generator_details["tech"] = (
                    df_wind_from_generator_details["energy_source"]
                    .str.strip()
                    .str.upper()
                )

        except Exception as e:
            print(f"⚠️ Could not fetch/prepare generator-details for Enervis input: {e}")

    def _run_enervis_for_df(df_input: pd.DataFrame, source_label: str) -> pd.DataFrame:
        if df_input.empty:
            return pd.DataFrame()
        anemos = AnemosAPI(ANEMOS_EMAIL, ANEMOS_PASSWORD)
        anemos.get_token()
        turbine_ref = pd.read_excel(TURBINE_REFERENCE_PATH)
        turbine_ref["id"] = turbine_ref["id"].astype(str).str.strip()
        total_rows = len(df_input)
        num_batches = (total_rows + BATCH_SIZE - 1) // BATCH_SIZE
        all_results: List[pd.DataFrame] = []
        for batch_num in range(num_batches):
            start_idx = batch_num * BATCH_SIZE
            end_idx = min((batch_num + 1) * BATCH_SIZE, total_rows)
            df_batch = df_input.iloc[start_idx:end_idx].copy()
            batch_nan_path = (
                OUTPUT_FOLDER
                / f"batch_{source_label}_{batch_num+1:03d}_turbine_matching.xlsx"
            )
            result_df, _ = run_enervis_calculation(
                df_batch,
                anemos,
                turbine_ref,
                batch_nan_path,
                batch_info=f"{source_label} batch {batch_num+1:03d}",
            )
            if result_df is not None and not result_df.empty:
                all_results.append(result_df)
        return (
            pd.concat(all_results, ignore_index=True) if all_results else pd.DataFrame()
        )

    enervis_results_generator_details = _run_enervis_for_df(
        df_wind_from_generator_details, "generator-details"
    )
    enervis_results_stammdaten = _run_enervis_for_df(
        df_wind_from_stammdaten, "stammdaten"
    )

    enervis_results_fallback = pd.DataFrame()

    if not df_wind_from_generator_details.empty or not df_wind_from_stammdaten.empty:
        malos_to_retry: List[str] = []
        year_cols = [f"enervis_{year}" for year in TARGET_YEARS]
        existing_year_cols = [
            col for col in year_cols if col in enervis_results_generator_details.columns
        ]

        if existing_year_cols and not enervis_results_generator_details.empty:
            mask_all_nan = (
                enervis_results_generator_details[existing_year_cols].isna().all(axis=1)
            )
            malos_to_retry.extend(
                enervis_results_generator_details.loc[mask_all_nan, "malo"]
                .astype(str)
                .tolist()
            )

        if "malo" in df_wind_from_generator_details.columns:
            generator_malos = set(
                df_wind_from_generator_details["malo"].astype(str).str.strip().tolist()
            )
            result_malos = (
                set(
                    enervis_results_generator_details["malo"]
                    .astype(str)
                    .str.strip()
                    .tolist()
                )
                if not enervis_results_generator_details.empty
                else set()
            )
            malos_to_retry.extend(list(generator_malos - result_malos))

        malos_to_retry = list(set(malos_to_retry))
        print(
            f"\n🍌🍌🍌Retrying {len(malos_to_retry)} malo(s) with stammdaten fallback..."
        )
        print(
            f"   Malos to retry: {malos_to_retry[:20]}{'...' if len(malos_to_retry) > 20 else ''}"
        )

        if (
            malos_to_retry
            and not df_wind_from_stammdaten.empty
            and "malo" in df_wind_from_stammdaten.columns
        ):
            df_wind_from_stammdaten["malo"] = (
                df_wind_from_stammdaten["malo"].astype(str).str.strip()
            )
            df_fallback = df_wind_from_stammdaten[
                df_wind_from_stammdaten["malo"].isin(malos_to_retry)
            ].copy()
            enervis_results_fallback = _run_enervis_for_df(
                df_fallback, "stammdaten-fallback"
            )

    all_results: List[pd.DataFrame] = []
    if not enervis_results_generator_details.empty:
        tmp = enervis_results_generator_details.copy()
        tmp["source"] = "generator-details"
        all_results.append(tmp)
    if not enervis_results_fallback.empty:
        tmp = enervis_results_fallback.copy()
        tmp["source"] = "stammdaten-fallback"
        all_results.append(tmp)
    if not enervis_results_stammdaten.empty:
        tmp = enervis_results_stammdaten.copy()
        tmp["source"] = "stammdaten"
        all_results.append(tmp)

    if all_results:
        df_all = pd.concat(all_results, ignore_index=True)
        df_all["malo"] = df_all["malo"].astype(str).str.strip()

        # Define Enervis columns to average
        enervis_cols_to_avg = [f"enervis_{y}" for y in TARGET_YEARS] + ["avg_enervis"]
        existing_enervis_cols = [c for c in enervis_cols_to_avg if c in df_all.columns]

        # Build aggregation dictionary
        agg_dict = {}

        # Each column needs its own lambda to avoid closure issues
        for col in existing_enervis_cols:
            agg_dict[col] = lambda x, col=col: x.mean(skipna=True)

        # Keep first value for other columns
        other_cols = [
            c
            for c in df_all.columns
            if c not in ["malo"] + existing_enervis_cols + ["source"]
        ]
        for col in other_cols:
            agg_dict[col] = "first"

        # Group by malo and aggregate
        df_best = df_all.groupby("malo", dropna=False).agg(agg_dict).reset_index()

        enervis_results = df_best

        # Drop Enervis columns where all rows are null
        enervis_cols_to_check = [f"enervis_{y}" for y in TARGET_YEARS] + ["avg_enervis"]
        for col in enervis_cols_to_check:
            if col in enervis_results.columns and enervis_results[col].isna().all():
                enervis_results = enervis_results.drop(columns=[col])
    else:
        enervis_results = pd.DataFrame()

    df_units_enriched = df_stamm.copy()

    # merging units stammdaten with blindleister per unit
    if not blindleister_results.empty and "unit_mastr_id" in df_units_enriched.columns:
        df_units_enriched = df_units_enriched.merge(
            blindleister_results, on="unit_mastr_id", how="left"
        )

    # merging units stammdaten with enervis per malo
    if not enervis_results.empty and "malo" in df_units_enriched.columns:
        df_units_enriched["malo"] = df_units_enriched["malo"].astype(str).str.strip()
        enervis_results["malo"] = enervis_results["malo"].astype(str).str.strip()
        df_units_enriched = df_units_enriched.merge(
            enervis_results, on="malo", how="left"
        )

    # aggregate per malo
    df_assets_enriched = aggregate_stammdaten_by_malo(df_units_enriched)
    if (
        "tech" in df_assets_enriched.columns
        and "Technology" not in df_assets_enriched.columns
    ):
        df_assets_enriched = df_assets_enriched.rename(columns={"tech": "Technology"})

    # =============================================================================
    # STEP 5: PRODUCTION DATA (if present)
    # =============================================================================

    xls = pd.ExcelFile(input_path, engine="openpyxl")
    sheet_names = xls.sheet_names
    has_production = len(sheet_names) > 1 and sheet_names[0].lower() == "stammdaten"

    if not has_production:
        print_header("STEP 5: NO PRODUCTION DATA - GENERATING REPORT")
        df_assets_enriched = normalize_nan_strings(df_assets_enriched)
        sheet1, sheet2, sheet3 = generate_output_sheets(
            df_assets_enriched, has_production=False, has_forecast=False
        )

        customer_name = input_path.stem.split("_", 1)[0]

        sh1 = sanitize_sheet_name(f"{customer_name}_1")
        sh2 = sanitize_sheet_name(f"{customer_name}_2")
        sh3 = sanitize_sheet_name(f"{customer_name}_3")

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            sheet1.to_excel(writer, sheet_name=sh1, index=False, na_rep="")
            sheet2.to_excel(writer, sheet_name=sh2, index=False, na_rep="")
            sheet3.to_excel(writer, sheet_name=sh3, index=False, na_rep="")
        format_excel_output(str(out_path))
        print_header("PIPELINE COMPLETE")
        print(f"✅ Output saved to {out_path}")
        return

    print_header("STEP 5: PRODUCTION DATA FOUND - PROCESSING")

    if load_workbook is None:
        raise RuntimeError("openpyxl is required for production Excel processing")

    sheets = [
        s
        for s in sheet_names
        if s.lower()
        not in [
            "stammdaten",
            "curtailment",
            "unprocessed",
            "redispatch_wind",
            "redispatch_pv",
            "redispatch",
        ]
    ]

    merged = []
    batch_size = 5
    for i in range(0, len(sheets), batch_size):
        batch = sheets[i : i + batch_size]
        dfs = []
        for sh in batch:
            df_tmp = pd.read_excel(input_path, sheet_name=sh, engine="openpyxl")
            df_tmp.columns = df_tmp.columns.str.strip()
            if "malo" in df_tmp.columns:
                df_tmp["malo"] = df_tmp["malo"].astype(str).str.strip()
            dfs.append(df_tmp)
        merged.append(pd.concat(dfs, ignore_index=True))
        del dfs
        gc.collect()

    df_production_raw = pd.concat(merged, ignore_index=True)
    del merged
    gc.collect()

    df_production_raw.columns = df_production_raw.columns.str.strip()
    df_production_raw["malo"] = df_production_raw["malo"].astype(str).str.strip()
    df_production_raw = process_time_column(df_production_raw)

    df_production_filtered = filter_production_data_by_completeness(
        df_production_raw, debug=True
    )
    if df_production_filtered.empty:
        print("❌ No valid production data after filtering")
        return

    def custom_power_mwh(group):
        if group.nunique() == 1:
            return group.mean()
        return group.sum()

    grouped = df_production_filtered.groupby(
        ["malo", "time_berlin", "available_years", "available_months"]
    )
    df_production_filtered = grouped["power_kwh"].apply(custom_power_mwh).reset_index()
    df_production_filtered.rename(columns={"power_kwh": "infeed_kwh"}, inplace=True)
    df_production_filtered["__adj_kwh"] = 0.0

    wb = load_workbook(input_path, read_only=True)
    available_sheets = wb.sheetnames
    wb.close()

    if "redispatch" in available_sheets:
        df_redispatch = pd.read_excel(
            input_path, sheet_name="redispatch", engine="openpyxl"
        )
        df_redispatch.columns = df_redispatch.columns.str.strip()
        df_redispatch["malo"] = df_redispatch["malo"].astype(str).str.strip()
        df_redispatch = process_time_column(df_redispatch)

        # Apply custom_power_mwh aggregation
        grouped_redispatch = df_redispatch.groupby(["malo", "time_berlin"])
        df_redispatch = (
            grouped_redispatch["redispatch_kwh"].apply(custom_power_mwh).reset_index()
        )

        df_production_filtered = pd.merge(
            df_production_filtered,
            df_redispatch[["malo", "time_berlin", "redispatch_kwh"]],
            on=["malo", "time_berlin"],
            how="left",
        )
        df_production_filtered["__adj_kwh"] = df_production_filtered[
            "__adj_kwh"
        ] + df_production_filtered["redispatch_kwh"].fillna(0)

    if "curtailment" in available_sheets:
        df_curtailment = pd.read_excel(
            input_path, sheet_name="curtailment", engine="openpyxl"
        )
        df_curtailment.columns = df_curtailment.columns.str.strip()
        df_curtailment["malo"] = df_curtailment["malo"].astype(str).str.strip()
        df_curtailment = process_time_column(df_curtailment)

        # Apply custom_power_mwh aggregation
        grouped_curtailment = df_curtailment.groupby(["malo", "time_berlin"])
        df_curtailment = (
            grouped_curtailment["curtailment_kwh"].apply(custom_power_mwh).reset_index()
        )

        df_production_filtered = pd.merge(
            df_production_filtered,
            df_curtailment[["malo", "time_berlin", "curtailment_kwh"]],
            on=["malo", "time_berlin"],
            how="left",
        )
        df_production_filtered["curtailment_kwh"] = df_production_filtered[
            "curtailment_kwh"
        ].fillna(0)
        df_production_filtered["__adj_kwh"] = (
            df_production_filtered["__adj_kwh"]
            + df_production_filtered["curtailment_kwh"]
        )
        df_production_filtered["curtailment"] = df_production_filtered["malo"].isin(
            df_curtailment["malo"]
        )

    # Ensure optional adjustment columns always exist (so forecasting can run for all malos)
    if "redispatch_kwh" not in df_production_filtered.columns:
        df_production_filtered["redispatch_kwh"] = 0.0
    if "curtailment_kwh" not in df_production_filtered.columns:
        df_production_filtered["curtailment_kwh"] = 0.0
    if "curtailment" not in df_production_filtered.columns:
        df_production_filtered["curtailment"] = False

    df_production_filtered["infeed_kwh"] = pd.to_numeric(
        df_production_filtered["infeed_kwh"], errors="coerce"
    )
    df_production_filtered["__adj_kwh"] = pd.to_numeric(
        df_production_filtered["__adj_kwh"], errors="coerce"
    )
    df_production_filtered["redispatch_kwh"] = pd.to_numeric(
        df_production_filtered["redispatch_kwh"], errors="coerce"
    ).fillna(0)
    df_production_filtered["curtailment_kwh"] = pd.to_numeric(
        df_production_filtered["curtailment_kwh"], errors="coerce"
    ).fillna(0)
    df_production_filtered["power_kwh"] = (
        df_production_filtered["infeed_kwh"] + df_production_filtered["__adj_kwh"]
    )

    print_header("STEP 6: MERGING WITH PRICE DATA")
    df_dayahead_prices = load_day_ahead_prices(DAY_AHEAD_PRICE_PATH)
    df_rmv_prices = load_rmv_prices(RMV_PRICE_PATH)
    df_dayahead_prices_qh = expand_hourly_to_quarter_hourly(df_dayahead_prices)

    # Keep base components so forecast can be applied as: infeed + redispatch + forecasted curtailment
    df_production_qh_agg = df_production_filtered[
        [
            "malo",
            "time_berlin",
            "available_years",
            "available_months",
            "infeed_kwh",
            "redispatch_kwh",
            "curtailment_kwh",
            "power_kwh",
            "curtailment",
        ]
    ].copy()

    df_temp = df_stamm.copy()
    pattern = r"\d+.*rules"
    df_temp["_sort_priority"] = (
        df_temp.get("category", "").astype(str).str.contains(pattern, na=False)
    )
    df_sorted = df_temp.sort_values(["malo", "_sort_priority"], ascending=[True, False])

    df_assets_mapping = (
        df_sorted.groupby(["malo"], dropna=False)
        .agg({"tech": "first", "net_power_kw_unit": "sum", "category": "first"})
        .reset_index()
    )
    df_production_qh_agg = df_production_qh_agg.merge(
        df_assets_mapping, on="malo", how="left"
    )
    df_production_qh_agg.rename(
        columns={"net_power_kw_unit": "sum_power_kw_malo"}, inplace=True
    )

    df_prod_with_dayahead = pd.merge(
        df_production_qh_agg, df_dayahead_prices_qh, on="time_berlin", how="inner"
    )
    df_prod_with_dayahead["year"] = df_prod_with_dayahead["time_berlin"].dt.year.astype(
        "int16"
    )
    df_prod_with_dayahead["month"] = df_prod_with_dayahead[
        "time_berlin"
    ].dt.month.astype("int8")
    df_prod_with_dayahead["tech"] = (
        df_prod_with_dayahead["tech"]
        .astype(str)
        .str.strip()
        .str.upper()
        .astype("category")
    )

    df_prod_with_prices = df_prod_with_dayahead.merge(
        df_rmv_prices, on=["tech", "year", "month"], how="left"
    )
    df_prod_with_prices_dedup = df_prod_with_prices.drop_duplicates(
        subset=["malo", "time_berlin", "power_kwh"]
    )

    base_prod_delta = process_production_data(df_prod_with_prices_dedup)
    df_weighted_delta_by_malo = base_prod_delta["weighted_delta_permalo"].copy()
    df_capacity_inputs_by_malo = base_prod_delta["year_agg"].copy()

    for dfx in [
        df_assets_enriched,
        df_weighted_delta_by_malo,
        df_capacity_inputs_by_malo,
    ]:
        dfx["malo"] = dfx["malo"].astype(str).str.strip()

    df_assets_with_weighted_delta = df_assets_enriched.merge(
        df_weighted_delta_by_malo[["malo", "weighted_delta_permalo"]],
        on="malo",
        how="left",
    )
    df_assets_with_production_metrics = df_assets_with_weighted_delta.merge(
        df_capacity_inputs_by_malo, on="malo", how="left"
    )

    df_assets_with_production_metrics["denominator"] = (
        df_assets_with_production_metrics["Power in MW"]
        * df_assets_with_production_metrics["available_months"]
        * 730
    )
    df_assets_with_production_metrics["denominator"] = (
        df_assets_with_production_metrics["denominator"].replace(0, float("nan"))
    )

    df_assets_with_production_metrics["capacity_factor_percent"] = (
        (df_assets_with_production_metrics["total_prod_mwh"])
        / df_assets_with_production_metrics["denominator"]
        * 100
    ).round(2)

    # =========================================================================
    # STEP 8: ML CURTAILMENT FORECASTING (if data for 2023-2025)
    # =========================================================================

    print_header("STEP 8: CHECKING FOR CURTAILMENT FORECASTING")

    df_forecast_weighted_delta_by_malo: Optional[pd.DataFrame] = None
    df_forecast_capacity_inputs_by_malo: Optional[pd.DataFrame] = None
    df_cf_compare: Optional[pd.DataFrame] = None

    has_recent_data = (
        df_prod_with_prices["time_berlin"].dt.year.isin([2023, 2024, 2025]).any()
    )

    if not ENABLE_FORECASTING:
        print("ℹ️ Forecasting disabled via ENABLE_FORECASTING=0")
    elif not has_recent_data:
        print("ℹ️ No recent data (2023-2025) for curtailment forecasting")
    else:
        print("✅ Recent data (2023-2025) found, proceeding with forecasting")

        df_forecast_input = df_prod_with_prices[
            df_prod_with_prices["time_berlin"].dt.year.isin([2023, 2024, 2025, 2026])
        ].copy()

        # Track malos with actual curtailment data (for validation metrics)
        curtailed_malos: List[str] = []
        if "curtailment" in df_production_filtered.columns:
            curtailed_malos = (
                df_production_filtered[df_production_filtered["curtailment"]]["malo"]
                .unique()
                .tolist()
            )

        df_forecast_input["has_actual_curtailment"] = df_forecast_input["malo"].isin(
            curtailed_malos
        )
        print(
            f"ℹ️ Found {len(curtailed_malos)} malos with existing curtailment data (will use for validation)"
        )

        # Provide actual labels only for malos where curtailment is available.
        # This prevents classification/regression metrics from being contaminated by rows without labels.
        if "curtailment_kwh" in df_forecast_input.columns:
            df_forecast_input["curtailment_kwh"] = pd.to_numeric(
                df_forecast_input["curtailment_kwh"], errors="coerce"
            ).fillna(0)
        else:
            df_forecast_input["curtailment_kwh"] = 0.0

        if "sum_power_kw_malo" in df_forecast_input.columns:
            denom_kw = pd.to_numeric(
                df_forecast_input["sum_power_kw_malo"], errors="coerce"
            )
        else:
            denom_kw = pd.Series(np.nan, index=df_forecast_input.index)

        # only for malos with actual curtailment data
        df_forecast_input["curtailment_kWh_per_kw"] = np.where(
            df_forecast_input["has_actual_curtailment"] & denom_kw.gt(0),
            df_forecast_input["curtailment_kwh"] / denom_kw,
            np.nan,
        )

        if df_forecast_input.empty:
            print("ℹ️ No data available for forecasting after filtering")
        else:
            try:
                if pandas_gbq is None:
                    raise RuntimeError(
                        "pandas_gbq not available; cannot fetch forecast features"
                    )

                print("Fetching forecast features from BigQuery...")
                forecast_table_query = f"""
                SELECT *
                FROM `{CURTAILMENT_FORECAST_TABLE}`
                ORDER BY delivery_start_berlin
                """
                df_forecast_features = pandas_gbq.read_gbq(
                    forecast_table_query, project_id=PROJECT_ID
                )
                df_forecast_features["delivery_start_berlin"] = pd.to_datetime(
                    df_forecast_features["delivery_start_berlin"], errors="coerce"
                )

                forecast_feature_cols = [
                    col
                    for col in df_forecast_features.columns
                    if col not in df_forecast_input.columns
                    or col in ["delivery_start_berlin", "tech"]
                ]

                df_forecast_input = df_forecast_input.merge(
                    df_forecast_features[forecast_feature_cols],
                    left_on=["time_berlin", "tech"],
                    right_on=["delivery_start_berlin", "tech"],
                    how="left",
                )

                print(f"✅ Merged forecast features: {len(df_forecast_input)} rows")

                results = run_curtailment_forecast_multi_category(
                    df_ts=df_forecast_input,
                    plot_class=False,
                    plot_reg=False,
                )

                if results is not None:
                    df_curtailment_forecast_predictions = results["combined"].copy()
                    print(
                        f"✅ Curtailment forecasting complete: {len(df_curtailment_forecast_predictions)} predictions"
                    )

                    df_curtailment_forecast_predictions.rename(
                        columns={"dayaheadprice_eur_mwh": "dayaheadprice"},
                        inplace=True,
                    )

                    # Only apply forecast curtailment to malos that DON'T have actual curtailment data
                    # For malos with actual curtailment, keep the original power_kwh
                    if (
                        "has_actual_curtailment"
                        not in df_curtailment_forecast_predictions.columns
                    ):
                        df_curtailment_forecast_predictions[
                            "has_actual_curtailment"
                        ] = False

                    df_curtailment_forecast_predictions[
                        "raw_curtailment_forecast_kwh"
                    ] = (
                        df_curtailment_forecast_predictions[
                            "predicted_curtailment_kWh_per_kw"
                        ]
                        * df_curtailment_forecast_predictions["sum_power_kw_malo"]
                    )

                    df_curtailment_forecast_predictions["curtailment_forecast_kwh"] = (
                        np.minimum(
                            df_curtailment_forecast_predictions[
                                "raw_curtailment_forecast_kwh"
                            ],
                            df_curtailment_forecast_predictions["sum_power_kw_malo"]
                            * 0.8
                            * 0.25,
                        )
                    )

                    df_curtailment_forecast_predictions["curtailment_forecast_kwh"] = (
                        df_curtailment_forecast_predictions[
                            "curtailment_forecast_kwh"
                        ].clip(lower=0)
                    )

                    # Forecasted production is always built from the unadjusted base components:
                    # infeed_kwh (measured) + redispatch_kwh + forecasted curtailment
                    # This matches your requirement: for malos with curtailment data, add forecast to infeed (not to already-adjusted power).
                    if "infeed_kwh" not in df_curtailment_forecast_predictions.columns:
                        raise RuntimeError(
                            "Missing 'infeed_kwh' in forecast predictions; cannot build forecast production"
                        )
                    if (
                        "redispatch_kwh"
                        not in df_curtailment_forecast_predictions.columns
                    ):
                        df_curtailment_forecast_predictions["redispatch_kwh"] = 0.0

                    df_curtailment_forecast_predictions["raw_power_kwh_forecast"] = (
                        pd.to_numeric(
                            df_curtailment_forecast_predictions["infeed_kwh"],
                            errors="coerce",
                        ).fillna(0)
                        + pd.to_numeric(
                            df_curtailment_forecast_predictions["redispatch_kwh"],
                            errors="coerce",
                        ).fillna(0)
                        + pd.to_numeric(
                            df_curtailment_forecast_predictions[
                                "curtailment_forecast_kwh"
                            ],
                            errors="coerce",
                        ).fillna(0)
                    )

                    # Cap forecasted production by 80% of quarter-hour installed capacity
                    df_curtailment_forecast_predictions["power_kwh_forecast"] = (
                        np.minimum(
                            df_curtailment_forecast_predictions[
                                "raw_power_kwh_forecast"
                            ],
                            df_curtailment_forecast_predictions["sum_power_kw_malo"]
                            * 0.85
                            * 0.25,
                        )
                    )

                    # For rows with actual curtailment context, keep original measured power when it is higher
                    df_curtailment_forecast_predictions["power_kwh_forecast"] = (
                        np.where(
                            df_curtailment_forecast_predictions["power_kwh_forecast"]
                            < df_curtailment_forecast_predictions["power_kwh"],
                            df_curtailment_forecast_predictions["power_kwh"],
                            df_curtailment_forecast_predictions["power_kwh_forecast"],
                        )
                    )

                    save_multisheet_excel(
                        df_curtailment_forecast_predictions,
                        str(out_path.parent / "df_out_forecast_results.xlsx"),
                    )

                    if plt is not None:
                        df_plot = df_curtailment_forecast_predictions.copy()
                        df_plot["time_berlin"] = pd.to_datetime(
                            df_plot["time_berlin"], errors="coerce"
                        )
                        df_plot = df_plot[df_plot["time_berlin"].notna()].copy()

                        total_prod_plot_col = None
                        if "total_prod_kwh" in df_plot.columns:
                            total_prod_plot_col = "total_prod_kwh"
                        elif "power_kwh" in df_plot.columns:
                            df_plot["total_prod_kwh"] = pd.to_numeric(
                                df_plot["power_kwh"], errors="coerce"
                            )
                            total_prod_plot_col = "total_prod_kwh"
                            print(
                                "ℹ️ 'total_prod_kwh' not found; using 'power_kwh' as total production for plotting"
                            )

                        if total_prod_plot_col is None:
                            print(
                                "⚠️ Skipping forecast plots: neither 'total_prod_kwh' nor 'power_kwh' available"
                            )
                        else:
                            plot_dir = out_path.parent / "forecast_plots_by_malo"
                            plot_dir.mkdir(parents=True, exist_ok=True)

                            for malo, group in df_plot.groupby("malo", dropna=False):
                                group = group.sort_values("time_berlin").copy()
                                group["power_kwh_forecast"] = pd.to_numeric(
                                    group["power_kwh_forecast"], errors="coerce"
                                )
                                group[total_prod_plot_col] = pd.to_numeric(
                                    group[total_prod_plot_col], errors="coerce"
                                )

                                fig, ax = plt.subplots(figsize=(14, 5))
                                ax.plot(
                                    group["time_berlin"],
                                    group["power_kwh_forecast"],
                                    label="prod_forecast_kwh",
                                    linewidth=1.4,
                                )
                                ax.plot(
                                    group["time_berlin"],
                                    group[total_prod_plot_col],
                                    label="client_data_kwh",
                                    linewidth=1.0,
                                    alpha=0.8,
                                )
                                ax.set_title(
                                    f"malo={malo} | Forecast vs Client data Production"
                                )
                                ax.set_xlabel("time_berlin")
                                ax.set_ylabel("kWh")
                                ax.legend()
                                ax.grid(True, alpha=0.25)
                                fig.autofmt_xdate()
                                fig.tight_layout()

                                safe_malo = re.sub(r"[^A-Za-z0-9._-]+", "_", str(malo))
                                fig.savefig(
                                    plot_dir
                                    / f"forecast_vs_total_prod_{safe_malo}.png",
                                    dpi=140,
                                )
                                plt.close(fig)

                            print(f"✅ Saved per-malo forecast plots to {plot_dir}")
                    else:
                        print(
                            "ℹ️ matplotlib not available; skipping per-malo forecast plots"
                        )

                    # malos_with_actual = int(df_curtailment_forecast_predictions["has_actual_curtailment"].any())

                    n_malos_with_actual = df_curtailment_forecast_predictions.loc[
                        df_curtailment_forecast_predictions["has_actual_curtailment"],
                        "malo",
                    ].nunique()

                    n_malos_without_actual = df_curtailment_forecast_predictions.loc[
                        ~df_curtailment_forecast_predictions["has_actual_curtailment"],
                        "malo",
                    ].nunique()

                    print(
                        f"🦪🦪 Forecast computed for {n_malos_without_actual} malos without actual curtailment (used for forecast outputs)"
                    )
                    print(
                        f"🦪🦪🦪 Forecast computed for {n_malos_with_actual} malos with actual curtailment (used for forecast outputs + validation metrics)"
                    )

                    df_forecast_prod_for_metrics = df_curtailment_forecast_predictions[
                        [
                            "malo",
                            "time_berlin",
                            "power_kwh_forecast",
                            "dayaheadprice",
                            "monthly_reference_market_price_eur_mwh",
                            "available_years",
                            "available_months",
                            "year",
                            "month",
                        ]
                    ]

                    forecast_curt_delta = process_production_data(
                        df_forecast_prod_for_metrics
                    )

                    monthly_compare_path = (
                        out_path.parent
                        / "monthly_weighted_delta_original_vs_forecast.xlsx"
                    )
                    build_monthly_weighted_delta_comparison(
                        df_original=df_prod_with_prices_dedup,
                        df_forecast=df_forecast_prod_for_metrics,
                        output_path=monthly_compare_path,
                    )

                    df_forecast_weighted_delta_by_malo = forecast_curt_delta[
                        "weighted_delta_permalo"
                    ].copy()

                    df_forecast_capacity_inputs_by_malo = forecast_curt_delta[
                        "year_agg"
                    ].copy()

                    df_forecast_weighted_delta_by_malo.rename(
                        columns={
                            "weighted_delta_permalo": "forecast_weighted_delta_permalo",
                            "total_prod_kwh_malo": "forecast_total_prod_kwh_malo",
                        },
                        inplace=True,
                    )

                    df_forecast_capacity_inputs_by_malo.rename(
                        columns={
                            "available_months": "forecast_available_months",
                            "available_years": "forecast_available_years",
                            "total_prod_mwh": "forecast_total_prod_mwh",
                        },
                        inplace=True,
                    )

                    print("✅ Forecast data processed and ready for merge")
                else:
                    print("ℹ️ Forecast pipeline returned no results")

            except Exception as e:
                print(f"⚠️ Curtailment forecasting failed: {str(e)}")
                print("Continuing without forecast data...")

    # =========================================================================
    # STEP 9: Merge forecast results
    # =========================================================================

    if (
        df_forecast_weighted_delta_by_malo is not None
        and df_forecast_capacity_inputs_by_malo is not None
    ):
        print_header("STEP 9: MERGING FORECAST RESULTS")

        for dfx in [
            df_forecast_weighted_delta_by_malo,
            df_forecast_capacity_inputs_by_malo,
        ]:
            dfx["malo"] = dfx["malo"].astype(str).str.strip()

        df_assets_with_forecast_delta = df_assets_with_production_metrics.merge(
            df_forecast_weighted_delta_by_malo[
                [
                    "malo",
                    "forecast_weighted_delta_permalo",
                    "forecast_total_prod_kwh_malo",
                ]
            ],
            on="malo",
            how="left",
        )

        df_assets_with_forecast_metrics = df_assets_with_forecast_delta.merge(
            df_forecast_capacity_inputs_by_malo,
            on="malo",
            how="left",
        )

        df_assets_with_forecast_metrics["denominator_1"] = (
            df_assets_with_forecast_metrics["Power in MW"]
            * df_assets_with_forecast_metrics["forecast_available_months"]
            * 730
        )
        df_assets_with_forecast_metrics["denominator_1"] = (
            df_assets_with_forecast_metrics["denominator_1"].replace(0, float("nan"))
        )

        df_assets_with_forecast_metrics["forecast_capacity_factor_percent"] = (
            (df_assets_with_forecast_metrics["forecast_total_prod_kwh_malo"])
            / 1000
            / df_assets_with_forecast_metrics["denominator_1"]
            * 100
        ).round(2)

        # Capacity factor debug: compare original vs forecast and save to file
        try:
            df_cf_compare = build_capacity_factor_comparison(
                df_base=df_assets_with_production_metrics,
                df_forecast=df_assets_with_forecast_metrics,
            )
            cf_topn = int(os.getenv("CF_DEBUG_TOPN", "25"))
            print_capacity_factor_debug(df_cf_compare, top_n=cf_topn, only_drops=True)
            save_multisheet_excel(
                sanitize_for_excel(df_cf_compare),
                str(out_path.parent / "capacity_factor_comparison.xlsx"),
            )
        except Exception as e:
            print(f"⚠️ Could not generate capacity-factor debug report: {e}")

        df_report_input = df_assets_with_forecast_metrics
        has_forecast = True
        print("✅ Forecast results merged")
    else:
        df_report_input = df_assets_with_production_metrics
        has_forecast = False
        print("ℹ️ No forecast results to merge")

    # =========================================================================
    # STEP 10: Generate final reports
    # =========================================================================

    print_header("STEP 10: GENERATING REPORT")
    df_report_input = normalize_nan_strings(df_report_input)
    sheet1, sheet2, sheet3 = generate_output_sheets(
        df_report_input, has_production=True, has_forecast=has_forecast
    )

    # customer_name = input_path.stem
    customer_name = input_path.stem.split("_", 1)[0]

    sh1 = sanitize_sheet_name(f"{customer_name}_1")
    sh2 = sanitize_sheet_name(f"{customer_name}_2")
    sh3 = sanitize_sheet_name(f"{customer_name}_3")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sheet1.to_excel(writer, sheet_name=sh1, index=False, na_rep="")
        sheet2.to_excel(writer, sheet_name=sh2, index=False, na_rep="")
        sheet3.to_excel(writer, sheet_name=sh3, index=False, na_rep="")

    format_excel_output(str(out_path))
    print_header("PIPELINE COMPLETE")
    print(f"✅ Output saved to {out_path}")


if __name__ == "__main__":
    main()
